#!/usr/bin/env python3
"""
Export Governance Phase 1 - Test Suite
Tests all 11 verification points without modifying any source code.
Uses FastAPI TestClient with an isolated temp SQLite DB.
Run from project root: python test_export_governance_phase1.py
"""
import json
import os
import pathlib
import sys
import tempfile
from unittest.mock import patch

# ---------------------------------------------------------------------------
# 1. Path setup
# ---------------------------------------------------------------------------
PROJECT_ROOT = pathlib.Path(__file__).parent
sys.path.insert(0, str(PROJECT_ROOT))

# ---------------------------------------------------------------------------
# 2. Redirect DB to a fresh temp file BEFORE anything touches data.db
#    get_connection() reads DB_PATH at call time, so this works.
# ---------------------------------------------------------------------------
import data.db as _db_module

_tmp_db_file = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
_TMP_DB_PATH = pathlib.Path(_tmp_db_file.name)
_tmp_db_file.close()
_db_module.DB_PATH = _TMP_DB_PATH

# ---------------------------------------------------------------------------
# 3. Initialise schema in temp DB (idempotent; called again by lifespan)
# ---------------------------------------------------------------------------
from data.models import init_db
init_db()

# ---------------------------------------------------------------------------
# 4. Import app + TestClient after DB is redirected
# ---------------------------------------------------------------------------
from fastapi.testclient import TestClient
from api.app import app

# ---------------------------------------------------------------------------
# 5. Inject test API keys into KEY_ROLE_MAP (no env-var validation needed
#    because we write directly into the already-built dict)
# ---------------------------------------------------------------------------
import core.config as _cfg
_TEST_ADMIN_KEY = "test-admin-api-key-export-governance-check"   # >32 chars
_TEST_USER_KEY  = "test-user-api-key-export-governance-checker"  # >32 chars
_cfg.KEY_ROLE_MAP[_TEST_ADMIN_KEY] = "admin"
_cfg.KEY_ROLE_MAP[_TEST_USER_KEY]  = "user"

# ---------------------------------------------------------------------------
# 6. Helpers
# ---------------------------------------------------------------------------
from auth.jwt_auth import create_access_token
from data.report_service import save_report
from data.db import get_connection


def _jwt(user_id: str, role: str = "user") -> str:
    return create_access_token({"sub": user_id, "role": role})


_MINIMAL_CONTENT = {
    "sections": [
        {
            "type": "text",
            "heading": "Overview",
            "items": ["Test report for export governance verification."],
        }
    ]
}


def _create_report(user_id: str, title: str = "Test Report") -> int:
    return save_report(
        user_id=user_id,
        title=title,
        task_type="generate_dataset_report",
        content=_MINIMAL_CONTENT,
    )


def _query_export_logs(
    report_id: int | None = None,
    user_id: str | None = None,
) -> list[dict]:
    conn = get_connection()
    try:
        conds, params = [], []
        if report_id is not None:
            conds.append("report_id = ?")
            params.append(report_id)
        if user_id is not None:
            conds.append("user_id = ?")
            params.append(user_id)
        where = f"WHERE {' AND '.join(conds)}" if conds else ""
        rows = conn.execute(
            f"SELECT * FROM export_logs {where} ORDER BY id DESC",
            params,
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# 7. Test runner
# ---------------------------------------------------------------------------
PASS_SYM = "PASS"
FAIL_SYM = "FAIL"
_results: list[tuple[str, str, bool, str]] = []  # (id, label, ok, detail)


def check(test_id: str, label: str, ok: bool, detail: str = "") -> bool:
    sym = PASS_SYM if ok else FAIL_SYM
    _results.append((test_id, label, ok, detail))
    line = f"    [{sym}] {label}"
    if detail:
        line += f"  ({detail})"
    print(line)
    return ok


# ===========================================================================
# RUN
# ===========================================================================
print()
print("=" * 72)
print("  Export Governance Phase 1 - Automated Test Suite")
print("  DB:", _TMP_DB_PATH)
print("=" * 72)

USER_ID = "test-user-42"

with TestClient(app) as client:

    # -----------------------------------------------------------------------
    print("\n[T1] PDF export creates correct export_logs row")
    # -----------------------------------------------------------------------
    rid_pdf = _create_report(USER_ID, "PDF Export Test")
    r = client.get(
        f"/v1/reports/{rid_pdf}/export?format=pdf",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    logs = _query_export_logs(report_id=rid_pdf)
    print(f"    HTTP {r.status_code}  Content-Type: {r.headers.get('content-type')}")
    if logs:
        print("    DB row:", json.dumps(logs[0], indent=6, default=str))
    check("T1-01", "HTTP 200",              r.status_code == 200,
          f"got {r.status_code}")
    check("T1-02", "1 export_log row",      len(logs) == 1,
          f"got {len(logs)}")
    if logs:
        row = logs[0]
        check("T1-03", "status='success'",          row["status"] == "success",          repr(row["status"]))
        check("T1-04", "export_format='pdf'",        row["export_format"] == "pdf",        repr(row["export_format"]))
        check("T1-05", "report_id populated",        row["report_id"] == rid_pdf,          repr(row["report_id"]))
        check("T1-06", "user_id populated",           row["user_id"] == USER_ID,            repr(row["user_id"]))
        check("T1-07", "filename populated",          bool(row.get("filename")),            repr(row.get("filename")))
        check("T1-08", "file_size_bytes > 0",         (row.get("file_size_bytes") or 0) > 0, repr(row.get("file_size_bytes")))
        check("T1-09", "exported_at set",             bool(row.get("exported_at")),         repr(row.get("exported_at")))
        check("T1-10", "error_reason is NULL",        row.get("error_reason") is None,      repr(row.get("error_reason")))

    # -----------------------------------------------------------------------
    print("\n[T2] CSV export creates correct export_logs row")
    # -----------------------------------------------------------------------
    rid_csv = _create_report(USER_ID, "CSV Export Test")
    r = client.get(
        f"/v1/reports/{rid_csv}/export?format=csv",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    logs = _query_export_logs(report_id=rid_csv)
    print(f"    HTTP {r.status_code}  Content-Type: {r.headers.get('content-type')}")
    if logs:
        print("    DB row:", json.dumps(logs[0], indent=6, default=str))
    check("T2-01", "HTTP 200",              r.status_code == 200,       f"got {r.status_code}")
    check("T2-02", "1 export_log row",      len(logs) == 1,             f"got {len(logs)}")
    if logs:
        row = logs[0]
        check("T2-03", "status='success'",   row["status"] == "success",  repr(row["status"]))
        check("T2-04", "export_format='csv'", row["export_format"] == "csv", repr(row["export_format"]))
        check("T2-05", "filename ends .csv",  (row.get("filename") or "").endswith(".csv"), repr(row.get("filename")))
        check("T2-06", "file_size_bytes > 0", (row.get("file_size_bytes") or 0) > 0, repr(row.get("file_size_bytes")))
        check("T2-07", "exported_at set",     bool(row.get("exported_at")),             repr(row.get("exported_at")))

    # -----------------------------------------------------------------------
    print("\n[T3] JSON export creates correct export_logs row")
    # -----------------------------------------------------------------------
    rid_json = _create_report(USER_ID, "JSON Export Test")
    r = client.get(
        f"/v1/reports/{rid_json}/export?format=json",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    logs = _query_export_logs(report_id=rid_json)
    print(f"    HTTP {r.status_code}  Content-Type: {r.headers.get('content-type')}")
    if logs:
        print("    DB row:", json.dumps(logs[0], indent=6, default=str))
    check("T3-01", "HTTP 200",               r.status_code == 200,        f"got {r.status_code}")
    check("T3-02", "1 export_log row",       len(logs) == 1,              f"got {len(logs)}")
    if logs:
        row = logs[0]
        check("T3-03", "status='success'",    row["status"] == "success",   repr(row["status"]))
        check("T3-04", "export_format='json'", row["export_format"] == "json", repr(row["export_format"]))
        check("T3-05", "filename ends .json",  (row.get("filename") or "").endswith(".json"), repr(row.get("filename")))
        check("T3-06", "file_size_bytes > 0",  (row.get("file_size_bytes") or 0) > 0, repr(row.get("file_size_bytes")))
        check("T3-07", "exported_at set",      bool(row.get("exported_at")),              repr(row.get("exported_at")))

    # -----------------------------------------------------------------------
    print("\n[T4] X-Forwarded-For: first IP captured")
    # -----------------------------------------------------------------------
    rid_xff = _create_report(USER_ID, "XFF Test")
    r = client.get(
        f"/v1/reports/{rid_xff}/export?format=json",
        headers={
            "Authorization": f"Bearer {_jwt(USER_ID)}",
            "X-Forwarded-For": "10.0.0.1, 172.16.0.5, 192.168.1.1",
        },
    )
    logs = _query_export_logs(report_id=rid_xff)
    stored_ip = logs[0].get("ip_address") if logs else None
    print(f"    HTTP {r.status_code}  ip_address={stored_ip!r}")
    check("T4-01", "HTTP 200",                   r.status_code == 200,        f"got {r.status_code}")
    check("T4-02", "ip_address = '10.0.0.1'",    stored_ip == "10.0.0.1",     repr(stored_ip))
    check("T4-03", "ip_address not full header",  stored_ip != "10.0.0.1, 172.16.0.5, 192.168.1.1", repr(stored_ip))

    # -----------------------------------------------------------------------
    print("\n[T5] User-Agent captured and safely truncated to 512")
    # -----------------------------------------------------------------------
    long_ua = "TestBrowser/1.0 " + "X" * 600   # 616 chars total
    rid_ua = _create_report(USER_ID, "UA Test")
    r = client.get(
        f"/v1/reports/{rid_ua}/export?format=json",
        headers={
            "Authorization": f"Bearer {_jwt(USER_ID)}",
            "User-Agent": long_ua,
        },
    )
    logs = _query_export_logs(report_id=rid_ua)
    stored_ua = (logs[0].get("user_agent") or "") if logs else ""
    print(f"    HTTP {r.status_code}  sent_len={len(long_ua)} stored_len={len(stored_ua)}")
    print(f"    stored_ua[:60]={stored_ua[:60]!r}")
    check("T5-01", "HTTP 200",                      r.status_code == 200,            f"got {r.status_code}")
    check("T5-02", "user_agent stored (non-empty)",  len(stored_ua) > 0,             f"len={len(stored_ua)}")
    check("T5-03", "user_agent <= 512 chars",         len(stored_ua) <= 512,          f"len={len(stored_ua)}")
    check("T5-04", "user_agent truncated from 616",  len(stored_ua) < len(long_ua),  f"{len(stored_ua)} vs {len(long_ua)}")
    check("T5-05", "user_agent starts with 'TestBrowser/1.0'",
          stored_ua.startswith("TestBrowser/1.0"), repr(stored_ua[:20]))

    # -----------------------------------------------------------------------
    print("\n[T6] Nonexistent report -> 404, no export_log row created")
    # -----------------------------------------------------------------------
    nonexistent_id = 999999
    r = client.get(
        f"/v1/reports/{nonexistent_id}/export?format=pdf",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    logs = _query_export_logs(report_id=nonexistent_id)
    print(f"    HTTP {r.status_code}")
    check("T6-01", "HTTP 404",                  r.status_code == 404,   f"got {r.status_code}")
    check("T6-02", "0 export_log rows for 404", len(logs) == 0,         f"got {len(logs)}")

    # -----------------------------------------------------------------------
    print("\n[T7] Forced export failure -> status='failed' row with error_reason")
    # -----------------------------------------------------------------------
    rid_fail = _create_report(USER_ID, "Failure Test")
    with patch("api.v1.routes._build_pdf_bytes",
               side_effect=RuntimeError("PDF engine crashed")):
        r = client.get(
            f"/v1/reports/{rid_fail}/export?format=pdf",
            headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
        )
    logs = _query_export_logs(report_id=rid_fail)
    print(f"    HTTP {r.status_code}")
    if logs:
        print("    DB row:", json.dumps(logs[0], indent=6, default=str))
    check("T7-01", "HTTP 500",                    r.status_code == 500,              f"got {r.status_code}")
    check("T7-02", "1 export_log row created",    len(logs) == 1,                    f"got {len(logs)}")
    if logs:
        row = logs[0]
        check("T7-03", "status='failed'",          row["status"] == "failed",         repr(row["status"]))
        check("T7-04", "error_reason non-empty",   bool(row.get("error_reason")),     repr(row.get("error_reason")))
        check("T7-05", "error_reason has message", "PDF engine crashed" in (row.get("error_reason") or ""),
              repr(row.get("error_reason")))
        check("T7-06", "exported_at is NULL",      row.get("exported_at") is None,   repr(row.get("exported_at")))
        check("T7-07", "export_format='pdf'",      row["export_format"] == "pdf",    repr(row.get("export_format")))
        check("T7-08", "user_id populated",        row.get("user_id") == USER_ID,    repr(row.get("user_id")))
        check("T7-09", "report_id populated",      row.get("report_id") == rid_fail, repr(row.get("report_id")))

    # -----------------------------------------------------------------------
    print("\n[T8] Admin GET /admin/export-logs returns all rows")
    # -----------------------------------------------------------------------
    r = client.get(
        "/v1/admin/export-logs",
        headers={"x-api-key": _TEST_ADMIN_KEY},
    )
    body = r.json()
    print(f"    HTTP {r.status_code}  status={body.get('status')}  count={body.get('count')}")
    check("T8-01", "HTTP 200",              r.status_code == 200,            f"got {r.status_code}")
    check("T8-02", "status='success'",      body.get("status") == "success", repr(body.get("status")))
    check("T8-03", "data is a list",        isinstance(body.get("data"), list),
          type(body.get("data")).__name__)
    check("T8-04", "count matches data len", body.get("count") == len(body.get("data", [])),
          f"count={body.get('count')} len={len(body.get('data', []))}")
    check("T8-05", "count > 0 (prior exports)", (body.get("count") or 0) > 0,
          f"count={body.get('count')}")
    if body.get("data"):
        row0 = body["data"][0]
        check("T8-06", "rows have expected keys",
              all(k in row0 for k in ["id", "user_id", "report_id", "export_format",
                                      "status", "filename", "file_size_bytes",
                                      "exported_at", "created_at"]),
              f"keys={list(row0.keys())}")

    # -----------------------------------------------------------------------
    print("\n[T9] Admin filters by export_format and status")
    # -----------------------------------------------------------------------
    # filter export_format=pdf
    r9a = client.get(
        "/v1/admin/export-logs?export_format=pdf",
        headers={"x-api-key": _TEST_ADMIN_KEY},
    )
    b9a = r9a.json()
    pdf_rows = b9a.get("data", [])
    print(f"    export_format=pdf -> HTTP {r9a.status_code}  count={b9a.get('count')}")
    check("T9-01", "filter pdf: HTTP 200",      r9a.status_code == 200, f"got {r9a.status_code}")
    check("T9-02", "filter pdf: count >= 1",      len(pdf_rows) >= 1,    f"got {len(pdf_rows)}")
    check("T9-03", "filter pdf: all rows are pdf",
          all(r.get("export_format") == "pdf" for r in pdf_rows),
          f"formats={[r.get('export_format') for r in pdf_rows]}")

    # filter export_format=csv
    r9b = client.get(
        "/v1/admin/export-logs?export_format=csv",
        headers={"x-api-key": _TEST_ADMIN_KEY},
    )
    b9b = r9b.json()
    csv_rows = b9b.get("data", [])
    print(f"    export_format=csv -> HTTP {r9b.status_code}  count={b9b.get('count')}")
    check("T9-04", "filter csv: all rows are csv",
          all(r.get("export_format") == "csv" for r in csv_rows),
          f"formats={[r.get('export_format') for r in csv_rows]}")

    # filter status=failed
    r9c = client.get(
        "/v1/admin/export-logs?status=failed",
        headers={"x-api-key": _TEST_ADMIN_KEY},
    )
    b9c = r9c.json()
    failed_rows = b9c.get("data", [])
    print(f"    status=failed -> HTTP {r9c.status_code}  count={b9c.get('count')}")
    check("T9-05", "filter failed: HTTP 200",     r9c.status_code == 200, f"got {r9c.status_code}")
    check("T9-06", "filter failed: count >= 1 (T7 row)", len(failed_rows) >= 1, f"got {len(failed_rows)}")
    check("T9-07", "filter failed: all rows are failed",
          all(r.get("status") == "failed" for r in failed_rows),
          f"statuses={[r.get('status') for r in failed_rows]}")

    # filter status=success
    r9d = client.get(
        "/v1/admin/export-logs?status=success",
        headers={"x-api-key": _TEST_ADMIN_KEY},
    )
    b9d = r9d.json()
    success_rows = b9d.get("data", [])
    print(f"    status=success -> HTTP {r9d.status_code}  count={b9d.get('count')}")
    check("T9-08", "filter success: all rows are success",
          all(r.get("status") == "success" for r in success_rows),
          f"statuses={[r.get('status') for r in success_rows]}")

    # combined filter: export_format=pdf AND status=failed
    r9e = client.get(
        "/v1/admin/export-logs?export_format=pdf&status=failed",
        headers={"x-api-key": _TEST_ADMIN_KEY},
    )
    b9e = r9e.json()
    combined_rows = b9e.get("data", [])
    print(f"    export_format=pdf&status=failed -> count={b9e.get('count')}")
    check("T9-09", "combined filter: count >= 1 (T7 was pdf+failed)", len(combined_rows) >= 1,
          f"got {len(combined_rows)}")
    check("T9-10", "combined filter: all rows match both predicates",
          all(r.get("export_format") == "pdf" and r.get("status") == "failed"
              for r in combined_rows),
          f"rows={[(r.get('export_format'), r.get('status')) for r in combined_rows]}")

    # -----------------------------------------------------------------------
    print("\n[T10] Non-admin user returns 403")
    # -----------------------------------------------------------------------
    r10 = client.get(
        "/v1/admin/export-logs",
        headers={"x-api-key": _TEST_USER_KEY},
    )
    print(f"    HTTP {r10.status_code}")
    check("T10-01", "non-admin -> HTTP 403", r10.status_code == 403, f"got {r10.status_code}")

    # -----------------------------------------------------------------------
    print("\n[T11] init_db() is idempotent - schema and indexes survive second call")
    # -----------------------------------------------------------------------
    init_called_ok = False
    try:
        init_db()
        init_called_ok = True
    except Exception as exc:
        check("T11-01", "second init_db() raises no exception", False, str(exc))

    if init_called_ok:
        check("T11-01", "second init_db() raises no exception", True)

        conn = get_connection()
        try:
            col_rows = conn.execute("PRAGMA table_info(export_logs)").fetchall()
            cols = {r[1] for r in col_rows}
            idx_rows = conn.execute(
                "SELECT name FROM sqlite_master "
                "WHERE type='index' AND tbl_name='export_logs'"
            ).fetchall()
            indexes = {r[0] for r in idx_rows}
        finally:
            conn.close()

        expected_cols = {
            "id", "user_id", "report_id", "export_format", "filename",
            "file_size_bytes", "status", "error_reason", "ip_address",
            "user_agent", "exported_at", "created_at",
        }
        expected_idx = {
            "idx_export_logs_user_id",
            "idx_export_logs_report_id",
            "idx_export_logs_export_format",
            "idx_export_logs_exported_at",
        }

        print(f"    Columns ({len(cols)}): {sorted(cols)}")
        print(f"    Indexes: {sorted(indexes & expected_idx)}")

        missing_cols = expected_cols - cols
        missing_idx  = expected_idx  - indexes
        check("T11-02", "all 12 columns present",  len(missing_cols) == 0,
              f"missing={missing_cols}" if missing_cols else "")
        check("T11-03", "all 4 indexes present",   len(missing_idx) == 0,
              f"missing={missing_idx}"  if missing_idx  else "")
        check("T11-04", "no unexpected extra cols",
              (cols - expected_cols) == set(),
              f"extra={cols - expected_cols}" if cols - expected_cols else "")

        # Verify existing rows were not lost
        conn = get_connection()
        total_rows = conn.execute("SELECT COUNT(*) FROM export_logs").fetchone()[0]
        conn.close()
        print(f"    export_logs row count after second init_db(): {total_rows}")
        check("T11-05", "existing rows not wiped by second init_db()", total_rows > 0,
              f"count={total_rows}")

    # -----------------------------------------------------------------------
    print("\n[T12] XLSX export creates correct export_logs row")
    # -----------------------------------------------------------------------
    rid_xlsx = _create_report(USER_ID, "XLSX Export Test")
    r = client.get(
        f"/v1/reports/{rid_xlsx}/export?format=xlsx",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    logs = _query_export_logs(report_id=rid_xlsx)
    print(f"    HTTP {r.status_code}  Content-Type: {r.headers.get('content-type')}")
    if logs:
        print("    DB row:", json.dumps(logs[0], indent=6, default=str))
    check("T12-01", "HTTP 200",              r.status_code == 200,             f"got {r.status_code}")
    check("T12-02", "1 export_log row",      len(logs) == 1,                   f"got {len(logs)}")
    check("T12-03", "correct Content-Type",
          "spreadsheetml" in (r.headers.get("content-type") or ""),
          repr(r.headers.get("content-type")))
    check("T12-04", "Content-Disposition present",
          "attachment" in (r.headers.get("content-disposition") or "").lower(),
          repr(r.headers.get("content-disposition")))
    check("T12-05", "filename ends .xlsx",
          ".xlsx" in (r.headers.get("content-disposition") or "").lower(),
          repr(r.headers.get("content-disposition")))
    if logs:
        row = logs[0]
        check("T12-06", "status='success'",          row["status"] == "success",          repr(row["status"]))
        check("T12-07", "export_format='xlsx'",       row["export_format"] == "xlsx",      repr(row["export_format"]))
        check("T12-08", "report_id populated",        row["report_id"] == rid_xlsx,        repr(row["report_id"]))
        check("T12-09", "user_id populated",          row["user_id"] == USER_ID,           repr(row["user_id"]))
        check("T12-10", "filename ends .xlsx",        (row.get("filename") or "").endswith(".xlsx"), repr(row.get("filename")))
        check("T12-11", "filename starts toolsmithai_", (row.get("filename") or "").startswith("toolsmithai_"), repr(row.get("filename")))
        check("T12-12", "file_size_bytes > 0",        (row.get("file_size_bytes") or 0) > 0, repr(row.get("file_size_bytes")))
        check("T12-13", "exported_at set",            bool(row.get("exported_at")),        repr(row.get("exported_at")))
        check("T12-14", "error_reason is NULL",       row.get("error_reason") is None,     repr(row.get("error_reason")))

    # -----------------------------------------------------------------------
    print("\n[T13] XLSX workbook structure and document properties")
    # -----------------------------------------------------------------------
    import io as _io
    try:
        from openpyxl import load_workbook as _load_workbook
        _openpyxl_available = True
    except ImportError:
        _openpyxl_available = False

    _RICH_CONTENT = {
        "sections": [
            {"type": "executive_summary", "heading": "Executive Summary",
             "summary": "Strong performance this quarter.",
             "key_takeaways": ["Revenue up 12%"], "risks": ["Churn risk"], "opportunities": ["New market"]},
            {"type": "kpi", "heading": "Key Metrics",
             "kpis": [
                 {"label": "Revenue", "value": 1200000, "format": "currency", "trend": "up",
                  "status": "good", "description": "Total revenue"},
                 {"label": "Churn Rate", "value": 4.2, "format": "percent", "trend": "down",
                  "status": "warning", "description": "Monthly churn"},
             ]},
            {"type": "anomaly", "heading": "Anomalies",
             "anomalies": [{"title": "Missing values", "severity": "high",
                            "category": "missing_data", "description": "15% nulls", "evidence": "Col A"}]},
            {"type": "recommendation", "heading": "Recommendations",
             "recommendations": [{"title": "Clean data", "priority": "high",
                                  "action_type": "clean_data", "confidence": "high",
                                  "reason": "Improve model accuracy"}]},
        ]
    }

    rid_rich = save_report(
        user_id=USER_ID, title="Rich Report",
        task_type="generate_dataset_report", content=_RICH_CONTENT,
    )
    r13 = client.get(
        f"/v1/reports/{rid_rich}/export?format=xlsx",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    print(f"    HTTP {r13.status_code}  bytes={len(r13.content)}")
    check("T13-01", "HTTP 200",          r13.status_code == 200,  f"got {r13.status_code}")
    check("T13-02", "response has bytes", len(r13.content) > 0,   f"len={len(r13.content)}")

    if r13.status_code == 200 and _openpyxl_available:
        try:
            wb13 = _load_workbook(_io.BytesIO(r13.content))
            sheet_names = wb13.sheetnames
            print(f"    Sheets: {sheet_names}")
            check("T13-03", "'Cover' sheet exists",            "Cover" in sheet_names,            f"sheets={sheet_names}")
            check("T13-04", "'Summary Dashboard' is sheet[1]", sheet_names[1] == "Summary Dashboard", f"sheet[1]={sheet_names[1] if len(sheet_names) > 1 else 'n/a'}")
            check("T13-05", "'Executive Summary' sheet exists", "Executive Summary" in sheet_names, f"sheets={sheet_names}")
            check("T13-06", "'Key Metrics' sheet exists",       "Key Metrics" in sheet_names,      f"sheets={sheet_names}")
            check("T13-07", "'Anomalies & Risks' sheet exists", "Anomalies & Risks" in sheet_names, f"sheets={sheet_names}")
            check("T13-08", "'Recommendations' sheet exists",   "Recommendations" in sheet_names,  f"sheets={sheet_names}")
            # Cover sheet: brand name in A1
            cover = wb13["Cover"]
            check("T13-09", "Cover A1 contains brand name",
                  "ToolSmithAI" in str(cover["A1"].value or ""), repr(cover["A1"].value))
            # Cover sheet: title in A2
            check("T13-10", "Cover A2 contains report title",
                  "Rich Report" in str(cover["A2"].value or ""), repr(cover["A2"].value))
            # Key Metrics: freeze panes set (not None)
            kpi_ws = wb13["Key Metrics"]
            check("T13-11", "Key Metrics freeze_panes set",
                  kpi_ws.freeze_panes is not None, repr(kpi_ws.freeze_panes))
            # Document properties
            check("T13-12", "wb.properties.title is set",
                  bool(wb13.properties.title), repr(wb13.properties.title))
            check("T13-13", "wb.properties.creator is set",
                  bool(wb13.properties.creator), repr(wb13.properties.creator))
            check("T13-14", "wb.properties.created is set",
                  wb13.properties.created is not None, repr(wb13.properties.created))
        except Exception as exc13:
            check("T13-03", "workbook opened without error", False, str(exc13))
    elif not _openpyxl_available:
        print("    [SKIP] openpyxl not importable in test environment")

    # -----------------------------------------------------------------------
    print("\n[T14] Partial report (text section only) produces valid XLSX")
    # -----------------------------------------------------------------------
    rid_partial = _create_report(USER_ID, "Partial Report")  # uses _MINIMAL_CONTENT (text only)
    r14 = client.get(
        f"/v1/reports/{rid_partial}/export?format=xlsx",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    print(f"    HTTP {r14.status_code}  bytes={len(r14.content)}")
    check("T14-01", "HTTP 200",           r14.status_code == 200,  f"got {r14.status_code}")
    check("T14-02", "response has bytes", len(r14.content) > 0,    f"len={len(r14.content)}")
    logs14 = _query_export_logs(report_id=rid_partial)
    check("T14-03", "1 export_log row created", len(logs14) == 1,  f"got {len(logs14)}")
    if logs14:
        check("T14-04", "status='success'", logs14[0]["status"] == "success", repr(logs14[0]["status"]))

    if r14.status_code == 200 and _openpyxl_available:
        try:
            wb14 = _load_workbook(_io.BytesIO(r14.content))
            sn14 = wb14.sheetnames
            print(f"    Sheets: {sn14}")
            check("T14-05", "'Cover' present",             "Cover" in sn14,              f"sheets={sn14}")
            check("T14-06", "'Summary Dashboard' present", "Summary Dashboard" in sn14,  f"sheets={sn14}")
            check("T14-07", "'Key Metrics' absent",        "Key Metrics" not in sn14,    f"sheets={sn14}")
            check("T14-08", "'Anomalies & Risks' absent",  "Anomalies & Risks" not in sn14, f"sheets={sn14}")
        except Exception as exc14:
            check("T14-05", "partial workbook opened without error", False, str(exc14))

    # -----------------------------------------------------------------------
    print("\n[T15] Admin export-logs filter works for xlsx")
    # -----------------------------------------------------------------------
    r15 = client.get(
        "/v1/admin/export-logs?export_format=xlsx",
        headers={"x-api-key": _TEST_ADMIN_KEY},
    )
    b15 = r15.json()
    xlsx_rows = b15.get("data", [])
    print(f"    export_format=xlsx -> HTTP {r15.status_code}  count={b15.get('count')}")
    check("T15-01", "HTTP 200",                      r15.status_code == 200,    f"got {r15.status_code}")
    check("T15-02", "count >= 2 (T12 + T13 + T14)",  len(xlsx_rows) >= 2,       f"got {len(xlsx_rows)}")
    check("T15-03", "all rows have export_format=xlsx",
          all(row.get("export_format") == "xlsx" for row in xlsx_rows),
          f"formats={[row.get('export_format') for row in xlsx_rows]}")
    check("T15-04", "no pdf rows in xlsx filter",
          all(row.get("export_format") != "pdf" for row in xlsx_rows),
          f"formats={[row.get('export_format') for row in xlsx_rows]}")

    # -----------------------------------------------------------------------
    print("\n[T16] PDF metadata keys present in raw bytes")
    # -----------------------------------------------------------------------
    rid_meta = _create_report(USER_ID, "PDF Branding Phase1 Test")
    r16 = client.get(
        f"/v1/reports/{rid_meta}/export?format=pdf",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    print(f"    HTTP {r16.status_code}  bytes={len(r16.content)}")
    check("T16-01", "HTTP 200",              r16.status_code == 200,   f"got {r16.status_code}")
    check("T16-02", "response has bytes",    len(r16.content) > 0,     f"len={len(r16.content)}")
    if r16.status_code == 200 and r16.content:
        pdf_bytes = r16.content
        check("T16-03", "/Title key present in PDF",   b"/Title"   in pdf_bytes, "key missing from info dict")
        check("T16-04", "/Author key present in PDF",  b"/Author"  in pdf_bytes, "key missing from info dict")
        check("T16-05", "/Creator key present in PDF", b"/Creator" in pdf_bytes, "key missing from info dict")
        check("T16-06", "/Subject key present in PDF", b"/Subject" in pdf_bytes, "key missing from info dict")

    # -----------------------------------------------------------------------
    print("\n[T17] PDF author and creator value is 'ToolSmithAI'")
    # -----------------------------------------------------------------------
    if r16.status_code == 200 and r16.content:
        # fpdf2 serialises pure-ASCII metadata strings as (literal text) in the
        # PDF info dict, so the brand name appears as plain bytes in the output.
        check("T17-01", "ToolSmithAI literal in PDF bytes",
              b"ToolSmithAI" in r16.content,
              "brand name not found as ASCII literal in PDF output")

    # -----------------------------------------------------------------------
    print("\n[T18] PDF subject value is 'Executive Intelligence Report'")
    # -----------------------------------------------------------------------
    if r16.status_code == 200 and r16.content:
        check("T18-01", "Subject literal in PDF bytes",
              b"Executive Intelligence Report" in r16.content,
              "subject string not found as ASCII literal in PDF output")

    # -----------------------------------------------------------------------
    print("\n[T19] PDF export with Unicode-rich content (euro, arrows, bullets) succeeds")
    # -----------------------------------------------------------------------
    unicode_title = "€ Revenue → Growth: •Best• Metrics — Summary"
    rid_uni = _create_report(USER_ID, unicode_title)
    r19 = client.get(
        f"/v1/reports/{rid_uni}/export?format=pdf",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    print(f"    HTTP {r19.status_code}  bytes={len(r19.content)}")
    check("T19-01", "HTTP 200",                r19.status_code == 200, f"got {r19.status_code}")
    check("T19-02", "response has bytes",      len(r19.content) > 0,   f"len={len(r19.content)}")
    uni_logs = _query_export_logs(report_id=rid_uni)
    check("T19-03", "export_log row created",  len(uni_logs) >= 1,     f"got {len(uni_logs)}")
    if uni_logs:
        check("T19-04", "status='success'",    uni_logs[-1]["status"] == "success",
              repr(uni_logs[-1]["status"]))

    # -----------------------------------------------------------------------
    print("\n[T20] Unicode PDF is substantially larger than core-font PDF (TTF embedded)")
    # -----------------------------------------------------------------------
    if r19.status_code == 200 and r19.content:
        uni_size = len(r19.content)
        print(f"    PDF size: {uni_size} bytes")
        check("T20-01", "PDF > 5000 bytes (Arial TTF subset embedded)",
              uni_size > 5000,
              f"size={uni_size}; core-font PDFs are ~1600 bytes")

    # -----------------------------------------------------------------------
    print("\n[T21] Euro sign U+20AC is encoded as UTF-16 BE (not replaced with '?')")
    # -----------------------------------------------------------------------
    if r19.status_code == 200 and r19.content:
        pdf_bytes = r19.content
        # fpdf2 encodes non-ASCII PDF info-dict strings as UTF-16 BE hex.
        # Euro U+20AC -> hex sequence "20ac" in the PDF byte stream.
        # If _s() had still applied Latin-1 encode/decode, the euro sign
        # would become '?' (U+003F), which would never produce "20ac".
        check("T21-01", "euro U+20AC encoded as hex '20ac' in PDF (not replaced)",
              b"20ac" in pdf_bytes,
              "UTF-16 BE euro hex '20ac' not found; euro may have been replaced with '?'")
        check("T21-02", "ToolSmithAI brand present in PDF (Phase 1 regression)",
              b"ToolSmithAI" in pdf_bytes,
              "ToolSmithAI brand bytes missing from PDF output")

    # -----------------------------------------------------------------------
    print("\n[T22] PDF has cover page: document contains at least 2 pages")
    # -----------------------------------------------------------------------
    rid_cover = _create_report(USER_ID, "Phase 3 Cover Page Test")
    r22 = client.get(
        f"/v1/reports/{rid_cover}/export?format=pdf",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    print(f"    HTTP {r22.status_code}  bytes={len(r22.content)}")
    check("T22-01", "HTTP 200",               r22.status_code == 200,   f"got {r22.status_code}")
    check("T22-02", "response has bytes",     len(r22.content) > 0,     f"len={len(r22.content)}")
    if r22.status_code == 200 and r22.content:
        cover_pdf = r22.content
        # fpdf2 writes /Count N in the page tree; N>=2 means cover + content page
        check("T22-03", "PDF has 2+ pages (/Count >= 2 in page tree)",
              b"/Count 2" in cover_pdf or b"/Count 3" in cover_pdf or b"/Count 4" in cover_pdf,
              f"page count marker not found; PDF size={len(cover_pdf)}")
        check("T22-04", "PDF > 35000 bytes (cover + TTF overhead)",
              len(cover_pdf) > 35000,
              f"size={len(cover_pdf)}; expected > 35000 for multi-page TTF PDF")

    # -----------------------------------------------------------------------
    print("\n[T23] Cover page embeds metadata in PDF info dict (title, subject)")
    # -----------------------------------------------------------------------
    if r22.status_code == 200 and r22.content:
        cover_pdf = r22.content
        # set_title() and set_subject() write plain-text (uncompressed) entries
        # to the PDF info dict, making them searchable as raw bytes.
        check("T23-01", "Report title present in PDF bytes (set_title metadata)",
              b"Phase 3 Cover Page Test" in cover_pdf,
              "report title not found in PDF info dict bytes")
        check("T23-02", "Executive Intelligence Report in PDF (set_subject metadata)",
              b"Executive Intelligence Report" in cover_pdf,
              "set_subject value missing from PDF output")
        check("T23-03", "ToolSmithAI author in PDF (set_author metadata)",
              b"ToolSmithAI" in cover_pdf,
              "ToolSmithAI author bytes missing from PDF output")

    # -----------------------------------------------------------------------
    print("\n[T24] Multi-section report overflows to page 3+ (header/footer on each page)")
    # -----------------------------------------------------------------------
    _multi_sections = [
        {"type": "text", "heading": f"Section {i}", "items": [
            f"Finding {i}-A: This section contains enterprise analysis data point {i}.",
            f"Finding {i}-B: Additional supporting evidence for section {i} conclusions.",
            f"Finding {i}-C: Recommended follow-up actions based on section {i} data.",
        ]} for i in range(1, 16)
    ]
    rid_multi = save_report(
        user_id=USER_ID,
        title="Multi-Page Report with Header and Footer",
        task_type="generate_dataset_report",
        content={"sections": _multi_sections},
    )
    r24 = client.get(
        f"/v1/reports/{rid_multi}/export?format=pdf",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    print(f"    HTTP {r24.status_code}  bytes={len(r24.content)}")
    check("T24-01", "HTTP 200",               r24.status_code == 200,   f"got {r24.status_code}")
    check("T24-02", "response has bytes",     len(r24.content) > 0,     f"len={len(r24.content)}")
    if r24.status_code == 200 and r24.content:
        multi_pdf = r24.content
        # 15 sections of 3 items each forces overflow beyond page 2.
        # Check /Count is > 2, or file is substantially larger than a 2-page PDF.
        has_page3 = (b"/Count 3" in multi_pdf or b"/Count 4" in multi_pdf
                     or b"/Count 5" in multi_pdf or b"/Count 6" in multi_pdf)
        check("T24-03", "multi-section report produces 3+ pages",
              has_page3 or len(multi_pdf) > 80000,
              f"/Count not >=3 and size={len(multi_pdf)}")
    logs24 = _query_export_logs(report_id=rid_multi)
    if logs24:
        check("T24-04", "status='success' for multi-page export",
              logs24[-1]["status"] == "success",
              repr(logs24[-1]["status"]))

    # -----------------------------------------------------------------------
    print("\n[T25] Phase 1-2 regressions: XLSX and Unicode tests still pass")
    # -----------------------------------------------------------------------
    # Spot-check T1 (PDF basic), T12 (XLSX basic), T16 (metadata), T21 (Unicode)
    rid_reg = _create_report(USER_ID, "Phase 3 Regression Check")
    r25_pdf = client.get(
        f"/v1/reports/{rid_reg}/export?format=pdf",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    r25_xlsx = client.get(
        f"/v1/reports/{rid_reg}/export?format=xlsx",
        headers={"Authorization": f"Bearer {_jwt(USER_ID)}"},
    )
    check("T25-01", "PDF export still works (T1 regression)",
          r25_pdf.status_code == 200,  f"got {r25_pdf.status_code}")
    check("T25-02", "XLSX export still works (T12 regression)",
          r25_xlsx.status_code == 200, f"got {r25_xlsx.status_code}")
    check("T25-03", "PDF still has set_subject metadata (T18 regression)",
          r25_pdf.status_code == 200 and b"Executive Intelligence Report" in r25_pdf.content,
          "set_subject value missing")
    check("T25-04", "XLSX Content-Type correct (T12 regression)",
          r25_xlsx.status_code == 200 and
          "spreadsheetml" in r25_xlsx.headers.get("content-type", ""),
          repr(r25_xlsx.headers.get("content-type")))

    # -----------------------------------------------------------------------
    print('\n[T26] business_kpis section: structured KPI data renders in PDF')
    # -----------------------------------------------------------------------
    _t26_content = {
        'sections': [
            {
                'type': 'business_kpis',
                'heading': 'Key Performance Indicators',
                'dataset_label': 'Sales Financial Analytics',
                'kpis': [
                    {'label': 'Total Revenue', 'value': 4200000, 'value_formatted': '$4.2M',
                     'trend': 'up', 'description': 'Record quarter.', 'delta': 12.3,
                     'delta_direction': 'increase', 'status': 'good', 'confidence': 0.92},
                    {'label': 'Gross Margin', 'value': 0.342, 'value_formatted': '34.2%',
                     'trend': 'stable', 'description': 'Flat vs prior period.', 'delta': None,
                     'delta_direction': None, 'status': 'neutral', 'confidence': 0.88},
                ],
                'items': ['Total Revenue: $4.2M', 'Gross Margin: 34.2%'],
            },
            {'type': 'text', 'heading': 'Overview', 'items': ['T26 test section.']},
        ]
    }
    rid26 = save_report(user_id=USER_ID, title='Phase4-T26-business_kpis',
                        task_type='generate_dataset_report', content=_t26_content)
    r26 = client.get(f'/v1/reports/{rid26}/export?format=pdf',
                     headers={'Authorization': f'Bearer {_jwt(USER_ID)}'})
    print(f'    HTTP {r26.status_code}  bytes={len(r26.content)}')
    check('T26-01', 'HTTP 200 for business_kpis PDF export',
          r26.status_code == 200, f'got {r26.status_code}')
    check('T26-02', 'business_kpis PDF has content bytes',
          len(r26.content) > 0, f'len={len(r26.content)}')
    _logs26 = _query_export_logs(report_id=rid26)
    check('T26-03', "business_kpis export log status='success'",
          bool(_logs26) and _logs26[-1]['status'] == 'success',
          repr(_logs26[-1]['status'] if _logs26 else 'no log'))

    # -----------------------------------------------------------------------
    print('\n[T27] segmentation_insights section: segment breakdown renders in PDF')
    # -----------------------------------------------------------------------
    _t27_content = {
        'sections': [
            {
                'type': 'segmentation_insights',
                'heading': 'Segmentation Analysis',
                'segments': [
                    {
                        'metric': 'Revenue', 'dimension': 'Region',
                        'insight_summary': 'North America leads with 42% of total revenue.',
                        'recommended_action': 'Invest in high-performing regions.',
                        'top_segments': [
                            {'label': 'North America', 'value': 1764000, 'pct_of_total': 42.0, 'rank': 1},
                            {'label': 'EMEA',          'value': 1050000, 'pct_of_total': 25.0, 'rank': 2},
                            {'label': 'APAC',          'value':  840000, 'pct_of_total': 20.0, 'rank': 3},
                        ],
                        'confidence': 0.85,
                    },
                ],
                'items': ['North America leads with 42% of total revenue.'],
            },
            {'type': 'text', 'heading': 'Overview', 'items': ['T27 test section.']},
        ]
    }
    rid27 = save_report(user_id=USER_ID, title='Phase4-T27-segmentation',
                        task_type='generate_dataset_report', content=_t27_content)
    r27 = client.get(f'/v1/reports/{rid27}/export?format=pdf',
                     headers={'Authorization': f'Bearer {_jwt(USER_ID)}'})
    print(f'    HTTP {r27.status_code}  bytes={len(r27.content)}')
    check('T27-01', 'HTTP 200 for segmentation_insights PDF export',
          r27.status_code == 200, f'got {r27.status_code}')
    check('T27-02', 'segmentation_insights PDF has content bytes',
          len(r27.content) > 0, f'len={len(r27.content)}')
    _logs27 = _query_export_logs(report_id=rid27)
    check('T27-03', "segmentation_insights export log status='success'",
          bool(_logs27) and _logs27[-1]['status'] == 'success',
          repr(_logs27[-1]['status'] if _logs27 else 'no log'))

    # -----------------------------------------------------------------------
    print('\n[T28] drilldown_table section: columnar table renders in PDF')
    # -----------------------------------------------------------------------
    _t28_content = {
        'sections': [
            {
                'type': 'drilldown_table',
                'heading': 'Revenue Drilldown by Product',
                'tables': [
                    {
                        'metric': 'Revenue', 'dimension': 'Product',
                        'columns': ['Product', 'Total', 'Share %'],
                        'rows': [
                            {'label': 'Enterprise Suite', 'value': 2100000, 'avg': 10500,
                             'count': 200, 'pct_of_total': 50.0, 'rank': 1},
                            {'label': 'Professional',     'value': 1260000, 'avg': 6300,
                             'count': 200, 'pct_of_total': 30.0, 'rank': 2},
                            {'label': 'Starter',          'value':  840000, 'avg': 4200,
                             'count': 200, 'pct_of_total': 20.0, 'rank': 3},
                        ],
                        'summary': 'Enterprise Suite drives 50% of revenue.',
                        'confidence': 0.88,
                    },
                ],
                'items': ['Enterprise Suite drives 50% of revenue.'],
            },
            {'type': 'text', 'heading': 'Overview', 'items': ['T28 test section.']},
        ]
    }
    rid28 = save_report(user_id=USER_ID, title='Phase4-T28-drilldown',
                        task_type='generate_dataset_report', content=_t28_content)
    r28 = client.get(f'/v1/reports/{rid28}/export?format=pdf',
                     headers={'Authorization': f'Bearer {_jwt(USER_ID)}'})
    print(f'    HTTP {r28.status_code}  bytes={len(r28.content)}')
    check('T28-01', 'HTTP 200 for drilldown_table PDF export',
          r28.status_code == 200, f'got {r28.status_code}')
    check('T28-02', 'drilldown_table PDF has content bytes',
          len(r28.content) > 0, f'len={len(r28.content)}')
    _logs28 = _query_export_logs(report_id=rid28)
    check('T28-03', "drilldown_table export log status='success'",
          bool(_logs28) and _logs28[-1]['status'] == 'success',
          repr(_logs28[-1]['status'] if _logs28 else 'no log'))

    # -----------------------------------------------------------------------
    print('\n[T29] forecast section: narrative items and projected outlook render in PDF')
    # -----------------------------------------------------------------------
    _t29_content = {
        'sections': [
            {
                'type': 'forecast',
                'heading': 'Forecast',
                'forecast_ready': True,
                'target_column': 'date',
                'method': 'linear_trend_with_moving_average',
                'horizon_periods': 3,
                'items': [
                    'Data source: 24 monthly records from 2022-01 to 2023-12.',
                    'Method: Linear trend with moving average.',
                    'Trend: Upward with consistent 4.2% monthly growth.',
                    'Projection range: 142 to 167 records per period.',
                ],
                'chart': {
                    'chart_type': 'forecast',
                    'labels': ['2023-10', '2023-11', '2023-12',
                               '2024-01 (F)', '2024-02 (F)', '2024-03 (F)'],
                    'historical': [120, 131, 138, None, None, None],
                    'forecast':   [None, None, None, 142.0, 151.5, 161.0],
                    'upper_band': [None, None, None, 154.0, 163.5, 173.0],
                    'lower_band': [None, None, None, 130.0, 139.5, 149.0],
                    'forecast_start_index': 3,
                    'date_column': 'date',
                },
            },
            {'type': 'text', 'heading': 'Overview', 'items': ['T29 test section.']},
        ]
    }
    rid29 = save_report(user_id=USER_ID, title='Phase4-T29-forecast',
                        task_type='generate_dataset_report', content=_t29_content)
    r29 = client.get(f'/v1/reports/{rid29}/export?format=pdf',
                     headers={'Authorization': f'Bearer {_jwt(USER_ID)}'})
    print(f'    HTTP {r29.status_code}  bytes={len(r29.content)}')
    check('T29-01', 'HTTP 200 for forecast PDF export',
          r29.status_code == 200, f'got {r29.status_code}')
    check('T29-02', 'forecast PDF has content bytes',
          len(r29.content) > 0, f'len={len(r29.content)}')
    _logs29 = _query_export_logs(report_id=rid29)
    check('T29-03', "forecast export log status='success'",
          bool(_logs29) and _logs29[-1]['status'] == 'success',
          repr(_logs29[-1]['status'] if _logs29 else 'no log'))

    # -----------------------------------------------------------------------
    print('\n[T30] ai_dashboard section: key insight and risk blocks render in PDF')
    # -----------------------------------------------------------------------
    _t30_content = {
        'sections': [
            {
                'type': 'ai_dashboard',
                'heading': 'Executive Intelligence',
                'most_important_insight': 'Revenue spike Q3: 45% increase in enterprise segment.',
                'highest_risk': 'Customer churn rate elevated above threshold in EMEA region.',
                'recommended_action': 'Investigate EMEA retention drivers and launch campaign.',
                'watchlist': ['EMEA churn +8.2%', 'Enterprise deals -3 vs target', 'Pipeline 0.8x'],
            },
            {'type': 'text', 'heading': 'Overview', 'items': ['T30 test section.']},
        ]
    }
    rid30 = save_report(user_id=USER_ID, title='Phase4-T30-ai_dashboard',
                        task_type='generate_dataset_report', content=_t30_content)
    r30 = client.get(f'/v1/reports/{rid30}/export?format=pdf',
                     headers={'Authorization': f'Bearer {_jwt(USER_ID)}'})
    print(f'    HTTP {r30.status_code}  bytes={len(r30.content)}')
    check('T30-01', 'HTTP 200 for ai_dashboard PDF export',
          r30.status_code == 200, f'got {r30.status_code}')
    check('T30-02', 'ai_dashboard PDF has content bytes (was empty before Phase 4)',
          len(r30.content) > 40000, f'len={len(r30.content)} (expected >40000 for non-empty PDF)')
    _logs30 = _query_export_logs(report_id=rid30)
    check('T30-03', "ai_dashboard export log status='success'",
          bool(_logs30) and _logs30[-1]['status'] == 'success',
          repr(_logs30[-1]['status'] if _logs30 else 'no log'))

    # -----------------------------------------------------------------------
    print('\n[T31] insight_priority section: severity-ranked insights render in PDF')
    # -----------------------------------------------------------------------
    _t31_content = {
        'sections': [
            {
                'type': 'insight_priority',
                'heading': 'Prioritized Insights',
                'insights': [
                    {'title': 'Revenue Spike Detected', 'severity': 'high',
                     'evidence': '45% YoY increase in enterprise Q3.',
                     'recommended_action': 'Investigate root drivers immediately.',
                     'confidence': 'high'},
                    {'title': 'EMEA Churn Risk', 'severity': 'medium',
                     'evidence': 'Churn rate elevated in EMEA by 8.2%.',
                     'recommended_action': 'Launch targeted retention campaign.',
                     'confidence': 'medium'},
                    {'title': 'Pipeline Gap Q4', 'severity': 'low',
                     'evidence': 'Q4 pipeline coverage at 0.8x target.',
                     'recommended_action': 'Add 5 qualified deals before quarter end.',
                     'confidence': 'medium'},
                ],
            },
            {'type': 'text', 'heading': 'Overview', 'items': ['T31 test section.']},
        ]
    }
    rid31 = save_report(user_id=USER_ID, title='Phase4-T31-insight_priority',
                        task_type='generate_dataset_report', content=_t31_content)
    r31 = client.get(f'/v1/reports/{rid31}/export?format=pdf',
                     headers={'Authorization': f'Bearer {_jwt(USER_ID)}'})
    print(f'    HTTP {r31.status_code}  bytes={len(r31.content)}')
    check('T31-01', 'HTTP 200 for insight_priority PDF export',
          r31.status_code == 200, f'got {r31.status_code}')
    check('T31-02', 'insight_priority PDF has content bytes (was empty before Phase 4)',
          len(r31.content) > 40000, f'len={len(r31.content)} (expected >40000 for non-empty PDF)')
    _logs31 = _query_export_logs(report_id=rid31)
    check('T31-03', "insight_priority export log status='success'",
          bool(_logs31) and _logs31[-1]['status'] == 'success',
          repr(_logs31[-1]['status'] if _logs31 else 'no log'))

    # -----------------------------------------------------------------------
    print('\n[T32] Phase 4 regression: prior PDF/XLSX/metadata tests still pass')
    # -----------------------------------------------------------------------
    rid32 = _create_report(USER_ID, 'Phase 4 Regression Check')
    r32_pdf  = client.get(f'/v1/reports/{rid32}/export?format=pdf',
                          headers={'Authorization': f'Bearer {_jwt(USER_ID)}'})
    r32_xlsx = client.get(f'/v1/reports/{rid32}/export?format=xlsx',
                          headers={'Authorization': f'Bearer {_jwt(USER_ID)}'})
    print(f'    PDF HTTP {r32_pdf.status_code}  XLSX HTTP {r32_xlsx.status_code}')
    check('T32-01', 'PDF export still works after Phase 4 (T1 regression)',
          r32_pdf.status_code == 200, f'got {r32_pdf.status_code}')
    check('T32-02', 'XLSX export still works after Phase 4 (T12 regression)',
          r32_xlsx.status_code == 200, f'got {r32_xlsx.status_code}')
    check('T32-03', 'PDF metadata set_subject preserved (T18 regression)',
          r32_pdf.status_code == 200 and b'Executive Intelligence Report' in r32_pdf.content,
          'set_subject metadata missing from PDF')


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
print()
print("=" * 72)
print("  RESULTS SUMMARY")
print("=" * 72)
passes = sum(1 for *_, ok, _ in _results if ok)
fails  = sum(1 for *_, ok, _ in _results if not ok)
total  = len(_results)
print(f"  Checks:  {total}  |  PASS: {passes}  |  FAIL: {fails}")
print()

if fails:
    print("  FAILED checks:")
    for tid, label, ok, detail in _results:
        if not ok:
            print(f"    [{tid}]  {label}")
            if detail:
                print(f"            -> {detail}")
    print()

# Cleanup temp DB
try:
    os.unlink(_TMP_DB_PATH)
    print(f"  Temp DB removed: {_TMP_DB_PATH}")
except Exception:
    print(f"  Note: could not remove temp DB: {_TMP_DB_PATH}")
print()

sys.exit(0 if fails == 0 else 1)
