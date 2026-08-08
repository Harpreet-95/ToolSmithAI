"""
Tests for Day 4, Capability 5 — Export:
core.answering.export_service and api/v1/composer.py::composer_export.

Unit-level for the builders (no DB needed — pure transforms over an
already-computed EnterpriseAnswer dict), plus a focused integration pass on
the endpoint's success/refusal gating, calling composer_export() directly
as a function the same way tests/test_conversation_context.py already
calls composer_ask() — bypassing FastAPI's DI layer.

Run from the project root:
    python -m pytest tests/test_export_service.py -v
"""
from __future__ import annotations

import csv
import io
import os

from cryptography.fernet import Fernet

os.environ.setdefault("ENCRYPTION_KEY", Fernet.generate_key().decode())
os.environ.setdefault("JWT_SECRET", "test-export-service-secret-long-enough-1234")
os.environ.setdefault("USER_ID_SALT", "test-export-service-salt-long-enough-12345")

import openpyxl
import pytest

from core.answering.export_service import (
    ALLOWED_FORMATS,
    build_csv_bytes,
    build_export,
    build_pdf_bytes,
    build_xlsx_bytes,
    check_exportable,
)

_SCALAR_ANSWER = {
    "answer": "There are 10,918 students in the database.",
    "summary": "10,918 students.",
    "answer_type": "live_query",
    "confidence": 68,
    "actual_value": 10918,
    "result_preview": [],
    "measure": None,
    "business_entity": "students",
    "applied_filters": [],
    "date_context": None,
    "governance_warnings": [],
    "citations": [{"label": "Live query result"}, {"label": "ADF_Student"}],
    "limitations": [],
    "source_tables": ["dbo.ADF_Student"],
    "insight": None,
    "chart": None,
    "truncation_notice": None,
}

_GROUPED_ANSWER = {
    "answer": "TX leads with 11,916 clients, followed by CA and FL.",
    "summary": "Top 10 comp state.",
    "answer_type": "live_query",
    "confidence": 76,
    "actual_value": None,
    "result_preview": [
        {"vw_Clients": 11916, "Comp State": "TX"},
        {"vw_Clients": 3997, "Comp State": "CA"},
        {"vw_Clients": 2426, "Comp State": "FL"},
    ],
    "measure": None,
    "business_entity": "clients",
    "applied_filters": [{"label": "Region", "operator": "=", "value": "South"}],
    "date_context": {"label": "last_quarter", "start": "2026-04-01", "end": "2026-06-30"},
    "governance_warnings": ["Confirmed PII column masked: CompContactEmail"],
    "citations": [{"label": "Live query result"}, {"label": "vw_Clients"}],
    "limitations": [],
    "source_tables": ["dbo.vw_Clients"],
    "insight": {
        "type": "period_comparison", "label": "vs. the previous period",
        "current_value": 6, "previous_value": 2, "percent_change": 200.0, "direction": "up",
    },
    "chart": {
        "chart_type": "bar_horizontal", "labels": ["TX", "CA", "FL"],
        "series": [{"name": "clients", "data": [11916, 3997, 2426]}],
    },
    "truncation_notice": None,
}


# ---------------------------------------------------------------------------
# check_exportable — Day 4's "never export a refused/clarification-required
# answer" rule as a hard backend gate.
# ---------------------------------------------------------------------------

def test_exportable_when_answered():
    assert check_exportable(_SCALAR_ANSWER, "answered") is None


def test_exportable_when_no_agent_status_and_normal_answer_type():
    assert check_exportable(_SCALAR_ANSWER, None) is None


def test_not_exportable_when_clarification_needed():
    reason = check_exportable({"answer_type": "clarification_needed"}, None)
    assert reason is not None
    assert "clarification" in reason.lower()


def test_not_exportable_when_safely_refused():
    reason = check_exportable(_SCALAR_ANSWER, "safely_refused")
    assert reason is not None


def test_not_exportable_when_governance_blocked():
    reason = check_exportable(_SCALAR_ANSWER, "governance_blocked")
    assert reason is not None


def test_not_exportable_when_execution_failed():
    assert check_exportable(_SCALAR_ANSWER, "execution_failed") is not None


def test_not_exportable_when_no_answer():
    assert check_exportable(None, "answered") is not None
    assert check_exportable({}, "answered") is not None


# ---------------------------------------------------------------------------
# CSV — result table only
# ---------------------------------------------------------------------------

def test_csv_grouped_answer_contains_only_table_rows():
    content = build_csv_bytes(_GROUPED_ANSWER)
    text = content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    assert rows[0] == ["vw_Clients", "Comp State"]
    assert rows[1:] == [["11916", "TX"], ["3997", "CA"], ["2426", "FL"]]
    # never invents/reorders — same order and values as result_preview
    assert "There are" not in text and "leads with" not in text  # no prose in a CSV


def test_csv_scalar_answer_synthesizes_a_single_row_from_actual_value():
    content = build_csv_bytes(_SCALAR_ANSWER)
    text = content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    assert rows == [["students"], ["10918"]]


def test_csv_never_changes_numeric_values():
    content = build_csv_bytes(_GROUPED_ANSWER)
    text = content.decode("utf-8-sig")
    assert "11916" in text and "3997" in text and "2426" in text


# ---------------------------------------------------------------------------
# Excel — Business Summary sheet + Result Table sheet
# ---------------------------------------------------------------------------

def test_xlsx_has_business_summary_and_result_table_sheets():
    content = build_xlsx_bytes("Which states have the most clients?", _GROUPED_ANSWER)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert wb.sheetnames == ["Business Summary", "Result Table"]


def test_xlsx_business_summary_contains_question_answer_insight_and_governance():
    content = build_xlsx_bytes("Which states have the most clients?", _GROUPED_ANSWER)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Business Summary"]
    values = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    joined = " | ".join(values)
    assert "Which states have the most clients?" in joined
    assert "TX leads with 11,916 clients" in joined
    assert "+200%" in joined
    assert "CompContactEmail" in joined  # governance preserved
    assert "Region = South" in joined  # filters preserved
    assert "last_quarter" in joined  # date/time context preserved


def test_xlsx_result_table_matches_result_preview_exactly():
    content = build_xlsx_bytes("q", _GROUPED_ANSWER)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Result Table"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["vw_Clients", "Comp State"]
    data_rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2, max_row=4)]
    assert data_rows == [[11916, "TX"], [3997, "CA"], [2426, "FL"]]


def test_xlsx_scalar_answer_builds_without_error():
    content = build_xlsx_bytes("How many students are in the database?", _SCALAR_ANSWER)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Result Table"]
    assert ws["A1"].value == "students"
    assert ws["A2"].value == 10918


# ---------------------------------------------------------------------------
# PDF — question, answer, insight, chart, table
# ---------------------------------------------------------------------------

def test_pdf_builds_valid_bytes():
    content = build_pdf_bytes("Which states have the most clients?", _GROUPED_ANSWER)
    assert content[:4] == b"%PDF"
    assert len(content) > 1000


def test_pdf_builds_for_scalar_answer_with_no_chart_or_insight():
    content = build_pdf_bytes("How many students are in the database?", _SCALAR_ANSWER)
    assert content[:4] == b"%PDF"


def test_pdf_handles_missing_chart_gracefully():
    answer = {**_SCALAR_ANSWER, "chart": None}
    content = build_pdf_bytes("q", answer)
    assert content[:4] == b"%PDF"


# ---------------------------------------------------------------------------
# build_export dispatch
# ---------------------------------------------------------------------------

def test_build_export_dispatches_correct_media_type_and_filename_extension():
    for fmt, media, ext in [
        ("csv", "text/csv", ".csv"),
        ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx"),
        ("pdf", "application/pdf", ".pdf"),
    ]:
        content, filename, media_type = build_export("How many clients?", _SCALAR_ANSWER, fmt)
        assert media_type == media
        assert filename.endswith(ext)
        assert len(content) > 0


def test_build_export_raises_on_unsupported_format():
    with pytest.raises(ValueError):
        build_export("q", _SCALAR_ANSWER, "docx")


def test_all_allowed_formats_are_actually_buildable():
    for fmt in ALLOWED_FORMATS:
        content, _, _ = build_export("q", _SCALAR_ANSWER, fmt)
        assert len(content) > 0


# ---------------------------------------------------------------------------
# api/v1/composer.py::composer_export — endpoint-level gating, called
# directly as a function (bypassing FastAPI's DI), the same convention
# tests/test_conversation_context.py already uses for composer_ask().
# ---------------------------------------------------------------------------

class _FakeClient:
    host = "127.0.0.1"


class _FakeRequest:
    """Minimal stand-in for fastapi.Request — composer_export only reads
    .headers.get(...) and .client.host."""
    def __init__(self, headers=None):
        self.headers = headers or {}
        self.client = _FakeClient()


from auth.api_key import AuthenticatedUser as _ApiKeyAuthenticatedUser  # noqa: E402
from api.v1.composer import ComposerExportRequest, composer_export  # noqa: E402

_USER = _ApiKeyAuthenticatedUser(role="user", user_id="user-export-1")


def test_endpoint_returns_400_for_clarification_needed_answer():
    body = ComposerExportRequest(
        question="q", enterprise_answer={"answer_type": "clarification_needed"}, format="csv",
    )
    resp = composer_export(body, _FakeRequest(), _USER)
    assert resp.status_code == 400


def test_endpoint_returns_400_for_refused_agent_status():
    body = ComposerExportRequest(
        question="q", enterprise_answer=_SCALAR_ANSWER, agent_status="safely_refused", format="pdf",
    )
    resp = composer_export(body, _FakeRequest(), _USER)
    assert resp.status_code == 400


def test_endpoint_returns_200_with_csv_content_type_on_success():
    body = ComposerExportRequest(
        question="Which states have the most clients?", enterprise_answer=_GROUPED_ANSWER,
        agent_status="answered", format="csv",
    )
    resp = composer_export(body, _FakeRequest(), _USER)
    assert resp.status_code == 200
    assert resp.media_type == "text/csv"
    assert "attachment" in resp.headers["content-disposition"]


def test_endpoint_returns_200_with_xlsx_content_type_on_success():
    body = ComposerExportRequest(
        question="How many students are in the database?", enterprise_answer=_SCALAR_ANSWER,
        agent_status="answered", format="xlsx",
    )
    resp = composer_export(body, _FakeRequest(), _USER)
    assert resp.status_code == 200
    assert resp.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_endpoint_returns_200_with_pdf_content_type_on_success():
    body = ComposerExportRequest(
        question="q", enterprise_answer=_GROUPED_ANSWER, agent_status="answered", format="pdf",
    )
    resp = composer_export(body, _FakeRequest(), _USER)
    assert resp.status_code == 200
    assert resp.media_type == "application/pdf"


def test_endpoint_allows_export_for_legacy_path_with_no_agent_status():
    body = ComposerExportRequest(
        question="q", enterprise_answer=_SCALAR_ANSWER, agent_status=None, format="csv",
    )
    resp = composer_export(body, _FakeRequest(), _USER)
    assert resp.status_code == 200
