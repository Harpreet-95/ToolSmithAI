from datetime import datetime, timedelta, timezone
import hashlib
import io
import math
import os
import secrets
import uuid

import pandas as pd

from fastapi import APIRouter, Depends, File, Query, Request, Response, UploadFile
from fastapi.responses import JSONResponse
from pydantic import BaseModel, EmailStr

from auth.api_key import AuthenticatedUser, require_api_key
from auth.jwt_auth import create_access_token, require_auth, require_jwt, require_role
from auth.password import hash_password, verify_password
from core.email import send_admin_invite_email, send_verification_email
from core.tools.report_generator import _build_pdf_bytes
from core.errors.error_response import build_error_response
from core.input.input_handler import handle_input
from core.security.encryption import decrypt
from core.output.output_formatter import format_output
from core.workflows.workflow_runner import run_workflow_by_name
from core.intelligence.semantic_classifier import classify_columns
from core.intelligence.segmentation_engine import compute_segmentation_profile_from_df
from data.dataset_service import create_dataset_summary, list_datasets_for_user, delete_dataset, rename_dataset, get_dataset_by_id, reprofile_dataset, replace_dataset_source_file
from data.scheduled_workflow_service import (
    create_scheduled_workflow,
    list_scheduled_workflows,
    delete_scheduled_workflow,
    pause_scheduled_workflow,
    resume_scheduled_workflow,
    get_schedule_health,
    get_scheduled_workflow_by_id,
    get_scheduled_workflow_by_engine_tool_id,
    update_scheduled_workflow_outcome,
    _classify_result as classify_schedule_result,
)
from data.scheduler_run_service import (
    create_schedule_run,
    complete_schedule_run,
    fail_schedule_run,
    list_runs_for_schedule,
    list_recent_runs_for_user,
)
from core.interpreter.task_interpreter import interpret_task
from data.workflow_service import create_workflow, list_workflows, delete_workflow, ALLOWED_MULTI_STEP_TYPES
from core.config import (
    ALLOWED_DATASET_EXTENSIONS,
    DATASET_UPLOADS_DIR,
    ENABLE_REAL_EMAIL,
    FRONTEND_BASE_URL,
    RETENTION_DAYS,
    ENABLE_AUTO_APPROVE_ENGINE_TOOLS,
)
from data.audit import delete_audit_log_entries, log_audit_event, purge_old_audit_db, purge_old_audit_log_file
from data.db import get_connection
from data.execution_history import enrich_execution_record, get_execution_by_id, get_repeated_intent_suggestions, get_workflow_success_insights, purge_old_execution_history
from data.usage_service import count_usage_events, list_usage_events
from data.report_service import list_reports_for_user, get_report_by_id, delete_report
from data.notification_service import list_notifications_for_user, mark_notification_read, delete_notification
from data.tool_service import (
    list_tools_db,
    get_tool_by_id_db,
    create_tool_db,
    update_tool_db,
    approve_tool_db,
)
from data.workspace_service import (
    create_workspace_draft,
    attach_workspace_proposal,
    attach_workspace_execution,
    save_workspace_db,
    list_workspaces_for_user,
    get_workspace_by_id,
    link_workspace_workflow,
)
from data.invite_service import create_invite, consume_invite
from data.export_log_service import (
    create_export_log,
    list_export_logs_for_user,
    list_all_export_logs,
    count_all_export_logs,
    get_export_log_summary,
)
from data.email_log_service import (
    list_all_email_logs,
    count_all_email_logs,
    get_email_log_summary,
)

# Side-effect imports: each module registers its connector into ConnectorRegistry on import.
import core.connectors.relational.mssql       # noqa: F401
import core.connectors.relational.postgresql  # noqa: F401
import core.connectors.relational.mysql       # noqa: F401
from data.datasource_service import (
    create_data_source,
    delete_data_source,
    get_metadata_job,
    list_data_sources,
    run_metadata_job,
    test_data_source,
)
from data.schema_service import (
    get_latest_snapshot,
    run_discovery,
)
from data.dictionary_service import (
    approve_column_dictionary,
    approve_table_dictionary,
    generate_and_save_dictionary,
    get_table_dictionary,
    list_dictionary_tables,
)
from data.profiling_service import (
    continue_batch_profiling,
    get_latest_profile,
    list_profile_history,
    run_full_profiling,
    run_structural_profiling,
    start_batch_profiling,
)
from data.domain_service import (
    generate_domain_assignments,
    get_domain_summary,
    list_domain_assignments,
)
from data.entity_service import (
    generate_entity_assignments,
    list_entity_assignments,
    get_entity_summary,
)
from data.domain_learning_service import (
    approve_domain_rule,
    generate_domain_rule_suggestions,
    list_domain_rule_suggestions,
    list_domain_rules,
    reject_domain_rule,
)
from data.entity_learning_service import (
    approve_entity_rule,
    generate_entity_rule_suggestions,
    list_entity_rule_suggestions,
    list_entity_rules,
    reject_entity_rule,
)
from data.domain_refinement_service import (
    analyze_rule_refinement,
    approve_refinement_suggestion,
    list_refinement_suggestions,
    reject_refinement_suggestion,
)

import logging
logger = logging.getLogger(__name__)


def _compute_histogram(series, n_bins: int = 10) -> list:
    """Compute up to n_bins equal-width histogram bins from a non-empty numeric series.

    Returns an empty list on any failure. Never raises.
    Bins span [min, max]; the last bin is closed on both ends.
    Zero-count bins are included so consumers always get a full n_bins picture.
    """
    try:
        if len(series) < 2:
            return []
        min_v = float(series.min())
        max_v = float(series.max())
        if not (math.isfinite(min_v) and math.isfinite(max_v)):
            return []
        if min_v == max_v:
            return [{"min": min_v, "max": max_v, "count": int(len(series))}]
        step = (max_v - min_v) / n_bins
        bins = []
        for i in range(n_bins):
            low  = min_v + i * step
            high = min_v + (i + 1) * step
            if i < n_bins - 1:
                mask = (series >= low) & (series < high)
            else:
                mask = (series >= low) & (series <= high)
            bins.append({
                "min":   round(low, 4),
                "max":   round(high, 4),
                "count": int(mask.sum()),
            })
        return bins
    except Exception:
        return []


def _compute_date_profile(
    df: pd.DataFrame,
    numeric_cols: list,
    categorical_cols: list,
) -> dict:
    """
    Detect date columns in a DataFrame and compute trend insights.
    Runs at upload time when the full DataFrame is available.
    Never raises — returns empty structure on any failure.
    """
    DATE_THRESHOLD = 0.70
    date_columns: list[dict] = []

    for col in categorical_cols:
        try:
            raw_series = df[col]
            series     = raw_series.dropna()
            if len(series) == 0:
                continue
            parsed = pd.to_datetime(series, errors="coerce")
            valid  = int(parsed.notna().sum())
            if valid == 0 or valid / len(series) < DATE_THRESHOLD:
                continue
            valid_dates = parsed.dropna().sort_values()
            earliest    = valid_dates.iloc[0]
            latest      = valid_dates.iloc[-1]
            null_count  = int(raw_series.isnull().sum())

            # Inferred granularity from median inter-record gap
            granularity = "unknown"
            try:
                deltas = valid_dates.diff().dropna().dt.days
                median_delta = float(deltas.median())
                if median_delta <= 1.5:
                    granularity = "daily"
                elif median_delta <= 8:
                    granularity = "weekly"
                elif median_delta <= 32:
                    granularity = "monthly"
            except Exception:
                pass

            # Monthly counts capped at 24
            monthly_counts: list = []
            try:
                mc = valid_dates.dt.to_period("M").value_counts().sort_index()
                monthly_counts = [
                    {"month": str(m), "count": int(c)}
                    for m, c in mc.items()
                ][:24]
            except Exception:
                pass

            date_columns.append({
                "column":              col,
                "earliest":            earliest.isoformat(),
                "latest":              latest.isoformat(),
                "min_date":            earliest.isoformat()[:10],
                "max_date":            latest.isoformat()[:10],
                "valid_count":         valid,
                "null_count":          null_count,
                "range_days":          int((latest - earliest).days),
                "inferred_granularity": granularity,
                "monthly_counts":      monthly_counts,
            })
        except Exception:
            continue

    trend_insights: list[dict] = []
    if date_columns and numeric_cols:
        primary_col = date_columns[0]["column"]
        try:
            df_s = df.copy()
            df_s["__date__"] = pd.to_datetime(df_s[primary_col], errors="coerce")
            df_s = df_s.dropna(subset=["__date__"]).sort_values("__date__").reset_index(drop=True)
            n = len(df_s)
            if n >= 4:
                half = n // 2
                for num_col in numeric_cols[:3]:
                    try:
                        f_mean = df_s.iloc[:half][num_col].dropna().mean()
                        s_mean = df_s.iloc[half:][num_col].dropna().mean()
                        if pd.isna(f_mean) or pd.isna(s_mean) or f_mean == 0:
                            continue
                        pct = round(((s_mean - f_mean) / abs(f_mean)) * 100, 1)
                        if abs(pct) < 5:
                            trend, symbol = "stable", "→"
                        elif pct > 0:
                            trend, symbol = "increasing", "↑"
                        else:
                            trend, symbol = "decreasing", "↓"
                        trend_insights.append({
                            "column":           num_col,
                            "trend":            trend,
                            "pct_change":       pct,
                            "symbol":           symbol,
                            "first_half_mean":  round(float(f_mean), 4),
                            "second_half_mean": round(float(s_mean), 4),
                        })
                    except Exception:
                        continue
        except Exception:
            pass

    return {"date_columns": date_columns, "trend_insights": trend_insights}


WORKFLOW_TEMPLATES: list[dict] = [
    {
        "id":          "weekly-dataset-report",
        "name":        "Weekly Dataset Report",
        "description": "Generate a structured report from your uploaded dataset every week.",
        "category":    "Reporting",
        "intent":      "generate a weekly dataset report",
        "frequency":   "weekly",
    },
    {
        "id":          "email-executive-summary",
        "name":        "Email Executive Summary",
        "description": "Generate and email a concise executive summary of your dataset metrics.",
        "category":    "Email",
        "intent":      "email me an executive summary of the dataset",
        "frequency":   "weekly",
    },
    {
        "id":          "monthly-kpi-report",
        "name":        "Monthly KPI Report",
        "description": "Create a detailed KPI analysis report from your dataset metrics each month.",
        "category":    "Reporting",
        "intent":      "generate a monthly KPI report from the dataset",
        "frequency":   "monthly",
    },
    {
        "id":          "daily-operations-digest",
        "name":        "Daily Operations Digest",
        "description": "Get a daily summary of key operational metrics from your current dataset.",
        "category":    "Operations",
        "intent":      "generate a daily operations digest from the dataset",
        "frequency":   "daily",
    },
    {
        "id":          "dataset-health-monitor",
        "name":        "Dataset Health Monitor",
        "description": "Check your dataset for anomalies, missing values, and data quality issues.",
        "category":    "Monitoring",
        "intent":      "analyze the dataset for data quality and missing values",
        "frequency":   "daily",
    },
    {
        "id":          "scheduled-sales-summary",
        "name":        "Scheduled Sales Summary",
        "description": "Generate a recurring sales performance summary from your uploaded sales data.",
        "category":    "Sales",
        "intent":      "generate a sales summary report from the dataset",
        "frequency":   "weekly",
    },
]


router = APIRouter()


# ---------------------------------------------------------------------------
# Auth request models
# ---------------------------------------------------------------------------

class RegisterRequest(BaseModel):
    name: str
    email: EmailStr
    password: str


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class RegisterAdminRequest(BaseModel):
    name: str
    email: EmailStr
    password: str
    invite_token: str


class CreateInviteRequest(BaseModel):
    email: EmailStr


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str
    confirm_password: str


# ---------------------------------------------------------------------------
# Auth endpoints
# ---------------------------------------------------------------------------

@router.post("/auth/register")
def register(request: RegisterRequest) -> dict:
    conn = get_connection()
    try:
        existing = conn.execute(
            "SELECT id FROM users WHERE email = ?", (request.email,)
        ).fetchone()
        if existing:
            return JSONResponse(
                status_code=409,
                content=build_error_response("Email already registered"),
            )

        raw_token = secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
        token_expires_at = (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
        now = datetime.now(timezone.utc).isoformat()

        cursor = conn.execute(
            "INSERT INTO users"
            " (name, email, password_hash, role, is_active, is_verified,"
            "  verification_token_hash, verification_token_expires_at, created_at)"
            " VALUES (?, ?, ?, 'user', 1, 0, ?, ?, ?)",
            (
                request.name,
                request.email,
                hash_password(request.password),
                token_hash,
                token_expires_at,
                now,
            ),
        )
        conn.commit()
        user_id = cursor.lastrowid
    finally:
        conn.close()

    send_verification_email(request.email, raw_token, user_id=str(user_id))

    token = create_access_token({"sub": str(user_id), "email": request.email, "role": "user"})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": user_id,
            "name": request.name,
            "email": request.email,
            "role": "user",
        },
    }


@router.post("/auth/login")
def login(request: LoginRequest) -> dict:
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT id, name, email, password_hash, role, is_active, is_verified FROM users WHERE email = ?",
            (request.email,),
        ).fetchone()

        if row is None or not verify_password(request.password, row["password_hash"]):
            return JSONResponse(
                status_code=401,
                content=build_error_response("Invalid email or password"),
            )

        if not row["is_active"]:
            return JSONResponse(
                status_code=403,
                content=build_error_response("Account is inactive"),
            )

        if ENABLE_REAL_EMAIL and not row["is_verified"]:
            return JSONResponse(
                status_code=403,
                content=build_error_response("Please verify your email before signing in."),
            )

        conn.execute(
            "UPDATE users SET last_login = ? WHERE id = ?",
            (datetime.now(timezone.utc).isoformat(), row["id"]),
        )
        conn.commit()
    finally:
        conn.close()

    token = create_access_token({"sub": str(row["id"]), "email": row["email"], "role": row["role"]})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": row["id"],
            "name": row["name"],
            "email": row["email"],
            "role": row["role"],
        },
    }


@router.post("/auth/register-admin")
def register_admin(request: RegisterAdminRequest) -> dict:
    """Register a new admin user using a valid invite token."""
    try:
        consume_invite(request.email, request.invite_token)
    except ValueError as exc:
        status = 422
        msg = str(exc)
        if "expired" in msg:
            status = 410
        elif "already been used" in msg:
            status = 409
        elif "does not match" in msg:
            status = 403
        return JSONResponse(status_code=status, content=build_error_response(msg))

    conn = get_connection()
    try:
        existing = conn.execute(
            "SELECT id FROM users WHERE email = ?", (request.email,)
        ).fetchone()
        if existing:
            return JSONResponse(
                status_code=409,
                content=build_error_response("Email already registered"),
            )

        raw_token = secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
        token_expires_at = (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
        now = datetime.now(timezone.utc).isoformat()

        cursor = conn.execute(
            "INSERT INTO users"
            " (name, email, password_hash, role, is_active, is_verified,"
            "  verification_token_hash, verification_token_expires_at, created_at)"
            " VALUES (?, ?, ?, 'admin', 1, 1, ?, ?, ?)",
            (
                request.name,
                request.email,
                hash_password(request.password),
                token_hash,
                token_expires_at,
                now,
            ),
        )
        conn.commit()
        user_id = cursor.lastrowid
    finally:
        conn.close()

    token = create_access_token({"sub": str(user_id), "email": request.email, "role": "admin"})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": user_id,
            "name": request.name,
            "email": request.email,
            "role": "admin",
        },
    }


@router.get("/auth/verify-email")
def verify_email(token: str) -> dict:
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT id, verification_token_expires_at FROM users WHERE verification_token_hash = ?",
            (token_hash,),
        ).fetchone()

        if row is None:
            return JSONResponse(
                status_code=400,
                content=build_error_response("Invalid or already used token"),
            )

        if datetime.fromisoformat(row["verification_token_expires_at"]) < datetime.now(timezone.utc):
            return JSONResponse(
                status_code=400,
                content=build_error_response("Verification link expired"),
            )

        conn.execute(
            "UPDATE users SET is_verified = 1, verification_token_hash = NULL,"
            " verification_token_expires_at = NULL WHERE id = ?",
            (row["id"],),
        )
        conn.commit()
    finally:
        conn.close()

    return {"status": "success", "message": "Email verified"}


@router.post("/auth/change-password")
def change_password(
    request: ChangePasswordRequest,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    if not request.current_password or not request.new_password or not request.confirm_password:
        return JSONResponse(status_code=400, content=build_error_response("All fields are required"))
    if request.new_password != request.confirm_password:
        return JSONResponse(status_code=400, content=build_error_response("Passwords do not match"))
    if len(request.new_password) < 8:
        return JSONResponse(status_code=400, content=build_error_response("New password must be at least 8 characters"))

    conn = get_connection()
    try:
        row = conn.execute(
            "SELECT password_hash FROM users WHERE id = ?", (user.user_id,)
        ).fetchone()
        if row is None:
            return JSONResponse(status_code=404, content=build_error_response("User not found"))

        if not verify_password(request.current_password, row["password_hash"]):
            log_audit_event(
                {"task_type": "password_change", "original_input": "change_password_attempt", "status": "failed"},
                user_id=user.user_id,
            )
            return JSONResponse(status_code=401, content=build_error_response("Current password is incorrect"))

        if verify_password(request.new_password, row["password_hash"]):
            return JSONResponse(status_code=400, content=build_error_response("New password must differ from your current password"))

        new_hash = hash_password(request.new_password)
        conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_hash, user.user_id))
        conn.commit()
    finally:
        conn.close()

    log_audit_event(
        {"task_type": "password_change", "original_input": "change_password_request", "status": "success"},
        user_id=user.user_id,
    )
    return {"status": "success", "data": {"message": "Password changed successfully"}}


# ---------------------------------------------------------------------------
# Existing request models
# ---------------------------------------------------------------------------

class InterpretRequest(BaseModel):
    input: str
    dataset_id: int | None = None
    recipient: str | None = None
    selected_sections: list[str] | None = None
    report_type: str | None = None


_VALID_REPORT_TYPES: frozenset[str] = frozenset({
    "executive",
    "risk",
    "forecast",
    "segmentation",
    "operational",
    "data_quality",
})


class WorkflowRunRequest(BaseModel):
    name: str


class CreateWorkflowRequest(BaseModel):
    name: str
    definition: dict


class DeleteMyDataRequest(BaseModel):
    confirm: bool = False


class CreateScheduledWorkflowRequest(BaseModel):
    input_text: str
    dataset_id: int | None = None
    refresh_before_run: bool = False


class CreateToolRequest(BaseModel):
    name: str
    slug: str | None = None
    config_json: dict


class UpdateToolRequest(BaseModel):
    name: str | None = None
    slug: str | None = None
    config_json: dict | None = None
    enabled: bool | None = None


class ComposeIntentRequest(BaseModel):
    intent: str
    dataset_id: int | None = None
    save_workspace: bool = False


@router.post("/interpret")
def interpret(request: InterpretRequest, user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    if not request.input.strip():
        return JSONResponse(status_code=400, content=build_error_response("Input cannot be empty"))
    if request.report_type is not None and request.report_type not in _VALID_REPORT_TYPES:
        return JSONResponse(
            status_code=400,
            content=build_error_response(
                f"Invalid report_type '{request.report_type}'. "
                f"Valid values: {', '.join(sorted(_VALID_REPORT_TYPES))}."
            ),
        )
    try:
        return handle_input(
            request.input,
            user_id=user.user_id,
            dataset_id=request.dataset_id,
            recipient=request.recipient,
            selected_sections=request.selected_sections,
            report_type=request.report_type,
        )
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/workflows")
def list_workflows_route(user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        return {"status": "success", "data": list_workflows(str(user.user_id))}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.delete("/workflows/{workflow_id}")
def delete_workflow_route(workflow_id: int, user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        deleted = delete_workflow(workflow_id, str(user.user_id))
        if not deleted:
            return JSONResponse(status_code=404, content=build_error_response("Workflow not found"))
        return {"status": "success", "data": {"deleted_id": workflow_id}}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/workflows")
def create_workflow_route(request: CreateWorkflowRequest, user: AuthenticatedUser = Depends(require_auth)) -> dict:
    try:
        workflow_id = create_workflow(request.name, request.definition, str(user.user_id))
        return {
            "status": "success",
            "data": {
                "workflow_id": workflow_id,
                "name": request.name,
            },
        }
    except ValueError as e:
        return JSONResponse(status_code=400, content=build_error_response("Invalid workflow", str(e)))
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/workflows/run")
def run_workflow(request: WorkflowRunRequest, user: AuthenticatedUser = Depends(require_auth)) -> dict:
    try:
        result = run_workflow_by_name(request.name, user_id=user.user_id)
        return format_output(result)
    except ValueError as e:
        return JSONResponse(status_code=404, content=build_error_response("Not found", str(e)))
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/workflows/{workflow_id}/run")
def run_workflow_by_id_route(workflow_id: int, user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    from core.workflows.workflow_runner import run_workflow_by_id
    try:
        result = run_workflow_by_id(workflow_id, user_id=user.user_id)
        return format_output(result)
    except ValueError as e:
        return JSONResponse(status_code=404, content=build_error_response("Not found", str(e)))
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/workflows/{workflow_id}/dry-run")
def dry_run_workflow_route(workflow_id: int, user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    """Validate a multi-step workflow plan and return a step-by-step preview.

    Does not execute any tools, write to the DB, or send notifications.
    Only supported for workflows with a workflow_steps definition.
    """
    from core.workflows.workflow_runner import run_multi_step_workflow
    from data.workflow_service import get_workflow_by_id
    workflow = get_workflow_by_id(workflow_id)
    if workflow is None:
        return JSONResponse(status_code=404, content=build_error_response("Workflow not found"))
    if not workflow["definition"].get("workflow_steps"):
        return JSONResponse(
            status_code=400,
            content=build_error_response(
                "Dry run is only supported for multi-step workflows"
            ),
        )
    try:
        result = run_multi_step_workflow(workflow, user_id=str(user.user_id), dry_run=True)
        return {"status": "success", "data": result}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/audit")
def get_audit_logs(
    user: AuthenticatedUser = Depends(require_role("admin")),
    limit: int = 10,
    task_type: str | None = None,
    from_timestamp: str | None = None,
    to_timestamp: str | None = None,
) -> dict:
    try:
        conn = get_connection()
        conditions = []
        params = []
        if task_type is not None:
            conditions.append("task_type = ?")
            params.append(task_type)
        if from_timestamp is not None:
            conditions.append("timestamp >= ?")
            params.append(from_timestamp)
        if to_timestamp is not None:
            conditions.append("timestamp <= ?")
            params.append(to_timestamp)
        where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        params.append(limit)
        rows = conn.execute(
            f"SELECT id, timestamp, task_type, original_input, status FROM audit_logs {where} ORDER BY id DESC LIMIT ?",
            params,
        ).fetchall()
        conn.close()
        data = []
        for row in rows:
            d = dict(row)
            d["original_input"] = decrypt(d["original_input"])
            data.append(d)
        return {"status": "success", "data": data}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/me/data")
def get_my_data(user: AuthenticatedUser = Depends(require_auth)) -> dict:
    try:
        conn = get_connection()

        audit_rows = conn.execute(
            "SELECT id, timestamp, task_type, original_input, status"
            " FROM audit_logs"
            " WHERE user_id = ?"
            " ORDER BY id DESC",
            (user.user_id,),
        ).fetchall()

        execution_rows = conn.execute(
            "SELECT eh.id, eh.plan_id, eh.workflow_id, eh.trigger_source,"
            " eh.task_type, eh.intent, eh.status,"
            " eh.started_at, eh.finished_at, eh.duration_ms,"
            " eh.step_count, eh.failed_step_id, eh.failed_tool, eh.error_message,"
            " eh.dataset_id,"
            " d.filename  AS dataset_name,"
            " d.row_count AS dataset_row_count"
            " FROM execution_history eh"
            " LEFT JOIN datasets d ON d.id = eh.dataset_id"
            " WHERE eh.user_id = ?"
            " ORDER BY eh.id DESC",
            (user.user_id,),
        ).fetchall()

        conn.close()

        audit_data = []
        for row in audit_rows:
            d = dict(row)
            d["original_input"] = decrypt(d["original_input"])
            audit_data.append(d)

        return {
            "status": "success",
            "data": {
                "audit_logs": audit_data,
                "execution_history": [enrich_execution_record(dict(row)) for row in execution_rows],
            },
        }
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.delete("/me/data")
def delete_my_data(
    request: DeleteMyDataRequest,
    user: AuthenticatedUser = Depends(require_api_key),
) -> dict:
    if not request.confirm:
        return JSONResponse(
            status_code=400,
            content=build_error_response('Confirmation required. Send {"confirm": true} to proceed.'),
        )
    try:
        conn = get_connection()
        audit_cur = conn.execute(
            "DELETE FROM audit_logs WHERE user_id = ?",
            (user.user_id,),
        )
        audit_deleted = audit_cur.rowcount
        exec_cur = conn.execute(
            "DELETE FROM execution_history WHERE user_id = ?",
            (user.user_id,),
        )
        exec_deleted = exec_cur.rowcount
        conn.commit()
        conn.close()
        log_deleted = delete_audit_log_entries(user.user_id)
        return {
            "status": "success",
            "data": {
                "deleted": {
                    "audit_logs": audit_deleted,
                    "execution_history": exec_deleted,
                    "audit_log_file": log_deleted,
                }
            },
        }
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/admin/invites")
def create_admin_invite_route(
    request: CreateInviteRequest,
    user: AuthenticatedUser = Depends(require_auth),
) -> dict:
    """Create an admin invite token. Requires an authenticated admin."""
    if user.role != "admin":
        return JSONResponse(
            status_code=403,
            content=build_error_response("Admin access required"),
        )
    try:
        result = create_invite(email=request.email, created_by=user.user_id)
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content=build_error_response("Failed to create invite", str(exc)),
        )
    try:
        send_admin_invite_email(
            recipient_email=request.email,
            invite_token=result["invite_token"],
            created_by_user_id=user.user_id,
        )
    except Exception as exc:
        print(f"[WARN] Admin invite email failed for {request.email}: {exc}")
    return {"status": "success", "data": result}


@router.post("/admin/purge")
def admin_purge(user: AuthenticatedUser = Depends(require_role("admin"))) -> dict:
    cutoff = (datetime.now(timezone.utc) - timedelta(days=RETENTION_DAYS)).isoformat()
    audit_db_deleted = purge_old_audit_db(cutoff)
    exec_deleted = purge_old_execution_history(cutoff)
    log_deleted = purge_old_audit_log_file(cutoff)
    return {
        "status": "success",
        "data": {
            "retention_days": RETENTION_DAYS,
            "cutoff": cutoff,
            "deleted": {
                "audit_logs": audit_db_deleted,
                "execution_history": exec_deleted,
                "audit_log_file": log_deleted,
            },
        },
    }


@router.post("/tools/compose")
def compose_tool_route(
    request: ComposeIntentRequest,
    user: AuthenticatedUser = Depends(require_auth),
) -> dict:
    """Analyse a natural-language intent and return a proposal for admin review.

    Never saves anything. Never executes anything.
    Uses deterministic rules; enriches with AI when ENABLE_AI_PLANNER=true.
    """
    if not request.intent or not request.intent.strip():
        return JSONResponse(status_code=400, content=build_error_response("intent must not be empty"))
    try:
        from core.composer.intent_composer import compose_from_intent
        proposal = compose_from_intent(
            intent=request.intent.strip(),
            dataset_id=request.dataset_id,
        )
        workspace_id = None
        if request.save_workspace:
            workspace_id = create_workspace_draft(
                user_id=str(user.user_id),
                intent_text=request.intent.strip(),
                dataset_id=request.dataset_id,
            )
            attach_workspace_proposal(
                workspace_id=workspace_id,
                user_id=str(user.user_id),
                proposal=proposal,
                proposal_source=proposal.get("source", "rule_based"),
            )
        response_data = dict(proposal)
        if workspace_id is not None:
            response_data["workspace_id"] = workspace_id
        return {"status": "success", "data": response_data}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/tools")
def create_tool_route(
    request: CreateToolRequest,
    user: AuthenticatedUser = Depends(require_role("admin")),
) -> dict:
    """Create a dynamic tool draft. Sets approved=0, enabled=0. Admin-only."""
    try:
        tool_id = create_tool_db(
            name=request.name,
            slug=request.slug,
            config_json=request.config_json,
            created_by=str(user.user_id),
        )
        tool = get_tool_by_id_db(tool_id)
        return {"status": "success", "data": tool}
    except ValueError as e:
        return JSONResponse(status_code=400, content=build_error_response(str(e)))
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.patch("/tools/{tool_id}/approve")
def approve_tool_route(
    tool_id: int,
    user: AuthenticatedUser = Depends(require_role("admin")),
) -> dict:
    """Validate config and set approved=1, enabled=1. Admin-only explicit approval step."""
    try:
        tool = approve_tool_db(tool_id=tool_id, approved_by=str(user.user_id))
        if tool is None:
            return JSONResponse(status_code=404, content=build_error_response("Tool not found"))
        return {"status": "success", "data": tool}
    except ValueError as e:
        return JSONResponse(status_code=400, content=build_error_response(str(e)))
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.patch("/tools/{tool_id}")
def update_tool_route(
    tool_id: int,
    request: UpdateToolRequest,
    user: AuthenticatedUser = Depends(require_role("admin")),
) -> dict:
    """Update a tool's editable fields. Blocked once approved=1. Admin-only."""
    updates = {k: v for k, v in request.model_dump(exclude_none=True).items()}
    try:
        tool = update_tool_db(tool_id=tool_id, updates=updates)
        if tool is None:
            return JSONResponse(status_code=404, content=build_error_response("Tool not found"))
        return {"status": "success", "data": tool}
    except ValueError as e:
        msg = str(e)
        status_code = 409 if "already been approved" in msg else 400
        return JSONResponse(status_code=status_code, content=build_error_response(msg))
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/tools")
def list_tools_route(user: AuthenticatedUser = Depends(require_role("admin"))) -> dict:
    """Return all tool rows from the DB. Admin-only — reveals system internals."""
    try:
        tools = list_tools_db()
        return {"status": "success", "data": tools}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/tools/{tool_id}")
def get_tool_route(tool_id: int, user: AuthenticatedUser = Depends(require_role("admin"))) -> dict:
    """Return a single tool row by ID. Admin-only."""
    try:
        tool = get_tool_by_id_db(tool_id)
        if tool is None:
            return JSONResponse(status_code=404, content=build_error_response("Tool not found"))
        return {"status": "success", "data": tool}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


# ---------------------------------------------------------------------------
# Engine — Tool Planning
# ---------------------------------------------------------------------------

class PlanToolRequest(BaseModel):
    intent: str
    context: dict | None = None


@router.post("/engine/tools/plan")
def plan_tool_route(
    request: PlanToolRequest,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Convert a natural-language intent into a draft ToolDefinition.

    Does not persist, execute, approve, or schedule anything.
    Returns the ToolDefinition as JSON for client review.
    """
    if not request.intent or not request.intent.strip():
        return JSONResponse(
            status_code=400,
            content=build_error_response("intent must not be empty"),
        )
    try:
        import dataclasses
        import json as _json
        from enum import Enum
        from core.engine.planner import plan_tool

        tool_def = plan_tool(request.intent.strip(), context=request.context)

        def _default(obj):
            if isinstance(obj, Enum):
                return obj.value
            if isinstance(obj, datetime):
                return obj.isoformat()
            raise TypeError(f"Not serializable: {type(obj)}")

        data = _json.loads(_json.dumps(dataclasses.asdict(tool_def), default=_default))
        return {"status": "success", "data": data}
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content=build_error_response("Internal server error", str(e)),
        )


# ---------------------------------------------------------------------------
# Engine — Full Lifecycle (save, approve, execute, query)
# ---------------------------------------------------------------------------

def _engine_serial(obj) -> dict:
    """Serialise an engine dataclass (ToolDefinition, RunRecord, etc.) to a
    JSON-safe dict. Enums → .value, datetime → ISO string."""
    import dataclasses
    import json as _json
    from enum import Enum

    def _default(o):
        if isinstance(o, Enum):
            return o.value
        if isinstance(o, datetime):
            return o.isoformat()
        raise TypeError(f"Not serializable: {type(o)}")

    return _json.loads(_json.dumps(dataclasses.asdict(obj), default=_default))


class SaveToolRequest(BaseModel):
    tool_definition: dict
    # "ai_workspace" identifies saves from the AI Workspace "Save as Reusable
    # Workflow" CTA.  All other callers omit this field (defaults to
    # "engine_workspace"), preserving the existing governance path.
    source: str = "engine_workspace"


@router.post("/engine/tools/save")
def save_tool_route(
    request: SaveToolRequest,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Validate and persist a ToolDefinition and its graph. No execution."""
    import sqlite3
    from core.engine.schema import validate_tool_definition
    from core.engine.contracts import SchemaValidationError
    from data.engine.tool_store import create_tool
    from data.engine.graph_store import save_graph

    try:
        tool_def = validate_tool_definition(request.tool_definition)
    except SchemaValidationError as e:
        return JSONResponse(
            status_code=400,
            content=build_error_response(str(e), getattr(e, "field_name", None)),
        )
    except Exception as e:
        return JSONResponse(
            status_code=400,
            content=build_error_response("Invalid tool definition", str(e)),
        )

    try:
        create_tool(tool_def)
    except sqlite3.IntegrityError:
        return JSONResponse(
            status_code=409,
            content=build_error_response(f"Tool '{tool_def.id}' already exists"),
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content=build_error_response("Internal server error", str(e)),
        )

    try:
        save_graph(tool_def.id, tool_def.graph)
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content=build_error_response("Internal server error", str(e)),
        )

    # ── Autonomous mode: auto-approve for immediate AI Workspace execution ────
    #
    # When ENABLE_AUTO_APPROVE_ENGINE_TOOLS=true and the save originates from
    # the AI Workspace "Save as Reusable Workflow" CTA (source="ai_workspace"),
    # the tool is automatically transitioned:
    #   draft → pending_approval → approved
    # so that executeEngineTool() succeeds without any manual governance step.
    #
    # The full approval state machine is still exercised (not bypassed):
    # SUBMITTED + APPROVED events are recorded with actor_id="system:auto_approve"
    # so the audit trail remains complete and enterprise-compatible.
    #
    # Intended for: demo / free-trial / autonomous orchestration mode.
    # Enterprise: set ENABLE_AUTO_APPROVE_ENGINE_TOOLS=false and use the
    # full submit → approve workflow (submit_engine_tool_route / approve_engine_tool_route).
    # Future: replace with per-tenant governance policy and RBAC compliance mode.
    import logging as _logging
    _log = _logging.getLogger(__name__)

    final_status = tool_def.status.value  # "draft" by default

    if ENABLE_AUTO_APPROVE_ENGINE_TOOLS and request.source == "ai_workspace":
        try:
            from core.engine.approval import submit_for_approval, approve_tool
            submit_for_approval(
                tool_def.id,
                actor_id="system:auto_approve",
                notes=(
                    "Autonomous mode — auto-submitted by AI Workspace "
                    "(ENABLE_AUTO_APPROVE_ENGINE_TOOLS=true)."
                ),
            )
            approve_tool(
                tool_def.id,
                actor_id="system:auto_approve",
                notes=(
                    "Autonomous mode — auto-approved for immediate AI Workspace execution. "
                    "Set ENABLE_AUTO_APPROVE_ENGINE_TOOLS=false to require manual governance."
                ),
            )
            final_status = "approved"
            _log.info(
                "Tool '%s' auto-approved [source=ai_workspace, autonomous_mode=true]",
                tool_def.id,
            )
        except Exception as _auto_err:
            # Non-fatal — tool remains in draft; Run Again will gracefully fall back.
            _log.warning(
                "Auto-approval failed for tool '%s': %s — tool remains in draft state",
                tool_def.id, _auto_err,
            )

    # ── Scheduler bridge: create a scheduled_workflows row when schedule.enabled ──
    #
    # Maps ScheduleSpec.schedule_type to APScheduler frequency values.
    # Skips row creation if a row already exists for this engine_tool_id (dedup).
    # Non-fatal — a bridge failure must not block the save response.
    if tool_def.schedule.enabled:
        _SCHED_TYPE_MAP = {
            "daily":     "daily",
            "weekly":    "weekly",
            "monthly":   "monthly",
            "recurring": "daily",
            "automated": "daily",
        }
        try:
            from core.engine.schedule_parser import parse_schedule_intent as _parse_sched
            _stype = (tool_def.schedule.schedule_type or "").lower()
            _frequency = _SCHED_TYPE_MAP.get(_stype, "daily")

            # Derive cron + human_label: prefer explicit schedule fields, then
            # parse the schedule_type string, then use tool name as last resort.
            _cron        = tool_def.schedule.cron or None
            _human_label = tool_def.schedule.human_label or None
            if not _cron:
                # Parse schedule_type ("weekly" → "0 9 * * 1") or tool name
                _parse_src = _stype or tool_def.name
                try:
                    _sm       = _parse_sched(_parse_src)
                    _cron        = _sm["cron"]
                    _human_label = _human_label or _sm["human_label"]
                    _frequency   = _sm["frequency"]
                except ValueError:
                    pass  # leave cron/human_label as None

            if get_scheduled_workflow_by_engine_tool_id(tool_def.id) is None:
                create_scheduled_workflow(
                    user_id=str(user.user_id),
                    dataset_id=None,
                    input_text=tool_def.name,
                    task_type="engine_tool",
                    frequency=_frequency,
                    day_of_week=None,
                    engine_tool_id=tool_def.id,
                    cron=_cron,
                    human_label=_human_label,
                )
                _log.info(
                    "Scheduler bridge: scheduled_workflow created for engine tool '%s' (frequency=%s, cron=%s)",
                    tool_def.id, _frequency, _cron,
                )
            else:
                _log.debug(
                    "Scheduler bridge: scheduled_workflow already exists for engine tool '%s' — skipped",
                    tool_def.id,
                )
        except Exception as _bridge_err:
            _log.warning(
                "Scheduler bridge failed for engine tool '%s': %s — continuing without schedule",
                tool_def.id, _bridge_err,
            )

    # Derive human_label for the response (best-effort — non-fatal)
    _save_human_label: str | None = tool_def.schedule.human_label or None
    if tool_def.schedule.enabled and not _save_human_label:
        try:
            from core.engine.schedule_parser import parse_schedule_intent as _p_sched
            _src = (tool_def.schedule.schedule_type or "").lower()
            if _src:
                _save_human_label = _p_sched(_src)["human_label"]
        except Exception:
            pass

    return {
        "status": "success",
        "data": {
            "tool_id":     tool_def.id,
            "name":        tool_def.name,
            "status":      final_status,
            "human_label": _save_human_label,
        },
    }


@router.post("/engine/tools/{tool_id}/submit")
def submit_engine_tool_route(
    tool_id: str,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Submit a draft tool for approval."""
    from data.engine.tool_store import get_tool
    from core.engine.approval import submit_for_approval
    from core.engine.contracts import EngineError

    if get_tool(tool_id) is None:
        return JSONResponse(
            status_code=404,
            content=build_error_response("Tool not found"),
        )
    try:
        submit_for_approval(tool_id, actor_id=str(user.user_id))
        return {"status": "success", "data": {"tool_id": tool_id, "action": "submitted"}}
    except EngineError as e:
        return JSONResponse(status_code=400, content=build_error_response(str(e)))
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/engine/tools/{tool_id}/approve")
def approve_engine_tool_route(
    tool_id: str,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Approve a pending tool so it can be executed."""
    from data.engine.tool_store import get_tool
    from core.engine.approval import approve_tool
    from core.engine.contracts import EngineError

    if get_tool(tool_id) is None:
        return JSONResponse(
            status_code=404,
            content=build_error_response("Tool not found"),
        )
    try:
        approve_tool(tool_id, actor_id=str(user.user_id))
        return {"status": "success", "data": {"tool_id": tool_id, "action": "approved"}}
    except EngineError as e:
        return JSONResponse(status_code=400, content=build_error_response(str(e)))
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


class ExecuteEngineToolRequest(BaseModel):
    inputs: dict = {}


@router.post("/engine/tools/{tool_id}/execute")
def execute_engine_tool_route(
    tool_id: str,
    request: ExecuteEngineToolRequest,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Execute an approved tool with the supplied inputs."""
    from data.engine.tool_store import get_tool
    from core.engine.runtime import execute_tool
    from core.engine.contracts import ApprovalRequiredError, EngineError

    tool_def = get_tool(tool_id)
    if tool_def is None:
        return JSONResponse(
            status_code=404,
            content=build_error_response("Tool not found"),
        )
    try:
        record = execute_tool(tool_def, request.inputs, user_id=str(user.user_id))
        return {"status": "success", "data": _engine_serial(record)}
    except ApprovalRequiredError as e:
        return JSONResponse(status_code=403, content=build_error_response(str(e)))
    except EngineError as e:
        return JSONResponse(status_code=400, content=build_error_response(str(e)))
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/engine/tools/{tool_id}/runs")
def list_engine_tool_runs_route(
    tool_id: str,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Return all run records for a tool, newest first."""
    from data.engine.run_store import list_runs_for_tool

    records = list_runs_for_tool(tool_id)
    return {"status": "success", "data": [_engine_serial(r) for r in records]}


@router.get("/engine/tools")
def list_engine_tools_route(
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Return a lightweight summary list of all saved engine tools, newest first.

    Does not include the execution graph — use GET /engine/tools/{tool_id} for full detail.
    """
    from data.engine.tool_store import list_tools

    tools = list_tools()
    data = [
        {
            "id":               t.id,
            "name":             t.name,
            "description":      t.description,
            "status":           t.status.value,
            "trigger_type":     t.trigger.type.value,
            "schedule_enabled": t.schedule.enabled,
            "schedule_type":    t.schedule.schedule_type,
            "created_at":       t.metadata.created_at.isoformat(),
            "updated_at":       t.metadata.updated_at.isoformat(),
        }
        for t in tools
    ]
    return {"status": "success", "data": data, "count": len(data)}


@router.get("/engine/tools/{tool_id}")
def get_engine_tool_route(
    tool_id: str,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Return a ToolDefinition by ID."""
    from data.engine.tool_store import get_tool

    tool_def = get_tool(tool_id)
    if tool_def is None:
        return JSONResponse(
            status_code=404,
            content=build_error_response("Tool not found"),
        )
    return {"status": "success", "data": _engine_serial(tool_def)}


@router.get("/engine/runs/{run_id}")
def get_engine_run_route(
    run_id: str,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Return a RunRecord by run_id, including step_results."""
    from data.engine.run_store import get_run

    record = get_run(run_id)
    if record is None:
        return JSONResponse(
            status_code=404,
            content=build_error_response("Run not found"),
        )
    return {"status": "success", "data": _engine_serial(record)}


@router.get("/workspaces")
def list_workspaces_route(user: AuthenticatedUser = Depends(require_auth)) -> dict:
    try:
        workspaces = list_workspaces_for_user(str(user.user_id))
        return {"status": "success", "data": workspaces}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/workspaces/{workspace_id}")
def get_workspace_route(workspace_id: int, user: AuthenticatedUser = Depends(require_auth)) -> dict:
    try:
        workspace = get_workspace_by_id(workspace_id, str(user.user_id))
        if workspace is None:
            return JSONResponse(status_code=404, content=build_error_response("Workspace not found"))
        return {"status": "success", "data": workspace}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


class WorkspaceExecutionRequest(BaseModel):
    execution_summary: dict
    report_id: int | None = None
    selected_sections: list[str] | None = None


@router.patch("/workspaces/{workspace_id}/execution")
def attach_workspace_execution_route(
    workspace_id: int,
    request: WorkspaceExecutionRequest,
    user: AuthenticatedUser = Depends(require_auth),
) -> dict:
    """Attach execution result to a workspace and advance status to 'executed'."""
    try:
        workspace = attach_workspace_execution(
            workspace_id=workspace_id,
            user_id=str(user.user_id),
            execution_summary=request.execution_summary,
            report_id=request.report_id,
            selected_sections=request.selected_sections,
        )
        if workspace is None:
            return JSONResponse(status_code=404, content=build_error_response("Workspace not found"))
        return {"status": "success", "data": workspace}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.patch("/workspaces/{workspace_id}/save")
def save_workspace_route(workspace_id: int, user: AuthenticatedUser = Depends(require_auth)) -> dict:
    """Mark a workspace as saved."""
    try:
        workspace = save_workspace_db(workspace_id, str(user.user_id))
        if workspace is None:
            return JSONResponse(status_code=404, content=build_error_response("Workspace not found"))
        return {"status": "success", "data": workspace}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/workspaces/{workspace_id}/create-workflow-draft")
def create_workflow_draft_route(
    workspace_id: int,
    user: AuthenticatedUser = Depends(require_auth),
) -> dict:
    """Convert a workspace's workflow proposal into a saved workflow draft.

    The draft is NOT executed. The user must explicitly run it from the Workflows panel.

    Validation:
      - User must own the workspace (404 otherwise).
      - Workspace must have an attached proposal (422).
      - Proposal type must be 'workflow'; dynamic_tool proposals are preview-only (422).
      - Every step type must be in ALLOWED_MULTI_STEP_TYPES (422).
      - If 'dataset_id' is a required input, the workspace must have a dataset attached (422).
    """
    # 1. Ownership
    workspace = get_workspace_by_id(workspace_id, str(user.user_id))
    if workspace is None:
        return JSONResponse(status_code=404, content=build_error_response("Workspace not found"))

    # 2. Proposal must exist
    proposal = workspace.get("proposal")
    if not proposal:
        return JSONResponse(
            status_code=422,
            content=build_error_response("No proposal attached to this workspace"),
        )

    # 3. Proposal type must be 'workflow'
    if proposal.get("proposal_type") != "workflow":
        return JSONResponse(
            status_code=422,
            content=build_error_response(
                "Proposal type is not 'workflow'. "
                "Dynamic tool proposals are preview-only and cannot be converted to workflow drafts."
            ),
        )

    # 4. Validate step types
    steps = proposal.get("primitives_or_steps") or []
    if not steps:
        return JSONResponse(
            status_code=422,
            content=build_error_response("Proposal contains no steps"),
        )
    invalid = [s.get("step_type") for s in steps if s.get("step_type") not in ALLOWED_MULTI_STEP_TYPES]
    if invalid:
        return JSONResponse(
            status_code=422,
            content=build_error_response(
                f"Proposal contains disallowed step type(s): {invalid}. "
                f"Allowed: {sorted(ALLOWED_MULTI_STEP_TYPES)}"
            ),
        )

    # 5. Required inputs satisfied
    required = proposal.get("required_inputs") or []
    if "dataset_id" in required and not workspace.get("dataset_id"):
        return JSONResponse(
            status_code=422,
            content=build_error_response(
                "Required input 'dataset_id' is not satisfied. "
                "Attach a dataset to the workspace before creating a workflow draft."
            ),
        )

    # 6. Build workflow definition from proposal steps
    workflow_steps = [
        {
            "type":    step["step_type"],
            "order":   step["order"],
            "purpose": step.get("purpose", ""),
        }
        for step in steps
    ]
    name = (proposal.get("suggested_name") or workspace.get("title") or "Untitled Workflow").strip()
    definition = {
        "workflow_steps": workflow_steps,
        "intent":         workspace.get("intent_text") or "",
        "dataset_id":     workspace.get("dataset_id"),
        "source":         "workspace_proposal",
        "risk_level":     proposal.get("risk_level", "low"),
    }

    try:
        workflow_id = create_workflow(name, definition, str(user.user_id))
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception as exc:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(exc)))

    # 7. Link workflow back to workspace (best-effort — workflow already saved if this fails)
    try:
        link_workspace_workflow(workspace_id, str(user.user_id), workflow_id)
    except Exception:
        pass

    return {
        "status": "success",
        "data": {
            "workflow_id":   workflow_id,
            "workflow_name": name,
            "workspace_id":  workspace_id,
            "definition":    definition,
            "message":       "Workflow draft created. Review and run it from the Workflows panel.",
        },
    }


@router.get("/insights")
def insights(user: AuthenticatedUser = Depends(require_auth)) -> dict:
    try:
        return {"status": "success", "data": get_workflow_success_insights(str(user.user_id))}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/recommendations")
def recommendations(user: AuthenticatedUser = Depends(require_auth)) -> dict:
    try:
        return {"status": "success", "data": get_repeated_intent_suggestions(str(user.user_id))}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/usage")
def usage(user: AuthenticatedUser = Depends(require_auth)) -> dict:
    try:
        return {
            "status": "success",
            "data": {
                "total_events": count_usage_events(user.user_id),
                "by_event_type": {
                    "interpret":     count_usage_events(user.user_id, event_type="interpret"),
                    "workflow_run":  count_usage_events(user.user_id, event_type="workflow_run"),
                },
                "recent_events": list_usage_events(user.user_id, limit=10),
            },
        }
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/datasets/upload")
async def upload_dataset(
    file: UploadFile = File(...),
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    fname_lower = (file.filename or "").lower()
    if not any(fname_lower.endswith(ext) for ext in ALLOWED_DATASET_EXTENSIONS):
        return JSONResponse(
            status_code=400,
            content=build_error_response("Only CSV and Excel files (.csv, .xlsx, .xls) are supported"),
        )
    try:
        contents = await file.read()
        if fname_lower.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(contents))
        elif fname_lower.endswith(".xlsx"):
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
        else:
            df = pd.read_excel(io.BytesIO(contents), engine="xlrd")
    except Exception as e:
        reason = str(e)[:300]
        fmt = "CSV" if fname_lower.endswith(".csv") else "Excel"
        return JSONResponse(
            status_code=400,
            content=build_error_response(f"Could not parse {fmt} file: {reason}"),
        )
    os.makedirs(DATASET_UPLOADS_DIR, exist_ok=True)
    _ext = fname_lower.rsplit(".", 1)[-1]
    _raw_file_path: str | None = os.path.join(DATASET_UPLOADS_DIR, f"{uuid.uuid4().hex}.{_ext}")
    try:
        with open(_raw_file_path, "wb") as _fh:
            _fh.write(contents)
    except Exception:
        _raw_file_path = None

    def _safe_float(val):
        try:
            v = float(val)
            return None if not math.isfinite(v) else round(v, 4)
        except (TypeError, ValueError, OverflowError):
            return None

    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    categorical_cols = df.select_dtypes(exclude="number").columns.tolist()

    # ── Enhanced numeric profile ──────────────────────────────────────────────
    numeric_profile: dict = {}
    for col in numeric_cols:
        raw    = df[col]
        series = raw.dropna()
        n_tot  = len(raw)
        n_val  = len(series)
        if n_val == 0:
            numeric_profile[col] = {
                "min": None, "max": None, "mean": None, "sum": None,
                "std": None, "median": None,
                "p25": None, "p75": None, "p90": None,
                "non_null_count": 0, "null_count": n_tot,
                "zero_count": 0, "negative_count": 0,
                "outlier_count_iqr": 0, "histogram_bins": [],
            }
        else:
            try:
                q25 = _safe_float(series.quantile(0.25))
                q75 = _safe_float(series.quantile(0.75))
                outlier_count = 0
                if q25 is not None and q75 is not None:
                    iqr_v  = q75 - q25
                    lower  = q25 - 1.5 * iqr_v
                    upper  = q75 + 1.5 * iqr_v
                    outlier_count = int(((series < lower) | (series > upper)).sum())
            except Exception:
                q25 = q75 = None
                outlier_count = 0

            numeric_profile[col] = {
                "min":              _safe_float(series.min()),
                "max":              _safe_float(series.max()),
                "mean":             _safe_float(series.mean()),
                "sum":              _safe_float(series.sum()),
                "std":              _safe_float(series.std()),
                "median":           _safe_float(series.median()),
                "p25":              q25,
                "p75":              q75,
                "p90":              _safe_float(series.quantile(0.90)),
                "non_null_count":   n_val,
                "null_count":       n_tot - n_val,
                "zero_count":       int((series == 0).sum()),
                "negative_count":   int((series < 0).sum()),
                "outlier_count_iqr": outlier_count,
                "histogram_bins":   _compute_histogram(series, 10),
            }

    missing_values = {col: int(count) for col, count in df.isnull().sum().items()}

    # ── Categorical profile (top values — structure preserved for report compat) ─
    categorical_profile: dict = {}
    for col in categorical_cols:
        top = df[col].value_counts().head(5)
        categorical_profile[col] = [
            {"value": str(v), "count": int(c)}
            for v, c in zip(top.index, top.values)
        ]

    # ── Categorical meta (enhanced stats in separate structure) ───────────────
    categorical_meta: dict = {}
    for col in categorical_cols:
        try:
            raw    = df[col]
            n_tot  = len(raw)
            n_null = int(raw.isnull().sum())
            n_val  = n_tot - n_null
            unique_count = int(raw.nunique())
            top_vals     = raw.value_counts()
            top_share    = round(float(top_vals.iloc[0]) / n_tot, 4) if len(top_vals) > 0 and n_tot > 0 else 0.0

            entropy_approx = None
            try:
                total = top_vals.sum()
                if total > 0:
                    props = top_vals / total
                    entropy_approx = round(
                        float(-sum(p * math.log(float(p)) for p in props if p > 0)),
                        4,
                    )
            except Exception:
                pass

            categorical_meta[col] = {
                "unique_count":    unique_count,
                "top_value_share": top_share,
                "null_count":      n_null,
                "non_null_count":  n_val,
                "entropy_approx":  entropy_approx,
            }
        except Exception:
            continue

    # ── Correlation profile (pairwise Pearson, capped at 20 strongest) ────────
    correlation_profile: list = []
    if len(numeric_cols) >= 2:
        try:
            corr_mx = df[numeric_cols].corr()
            pairs: list = []
            for i in range(len(numeric_cols)):
                for j in range(i + 1, len(numeric_cols)):
                    try:
                        c = float(corr_mx.iloc[i, j])
                        if not math.isfinite(c):
                            continue
                        abs_c    = abs(c)
                        strength = "strong" if abs_c >= 0.7 else ("moderate" if abs_c >= 0.4 else "weak")
                        pairs.append({
                            "column_a":    numeric_cols[i],
                            "column_b":    numeric_cols[j],
                            "correlation": round(c, 4),
                            "strength":    strength,
                        })
                    except Exception:
                        continue
            pairs.sort(key=lambda x: abs(x["correlation"]), reverse=True)
            correlation_profile = pairs[:20]
        except Exception:
            pass

    date_profile = _compute_date_profile(df, numeric_cols, categorical_cols)

    semantic_profile = classify_columns(
        columns=df.columns.tolist(),
        numeric_profile=numeric_profile,
        categorical_meta=categorical_meta,
        date_profile=date_profile,
        missing_values=missing_values,
        row_count=len(df),
    )

    segmentation_profile = compute_segmentation_profile_from_df(
        df=df,
        semantic_profile=semantic_profile,
        numeric_profile=numeric_profile,
        row_count=len(df),
    )

    dataset_id = create_dataset_summary(
        user_id=str(user.user_id),
        filename=file.filename,
        row_count=len(df),
        column_count=len(df.columns),
        columns=df.columns.tolist(),
        numeric_profile=numeric_profile,
        missing_values=missing_values,
        categorical_profile=categorical_profile,
        date_profile=date_profile,
        correlation_profile=correlation_profile or None,
        categorical_meta=categorical_meta or None,
        semantic_profile=semantic_profile or None,
        segmentation_profile=segmentation_profile if segmentation_profile.get("computed_pairs", 0) > 0 else None,
        file_path=_raw_file_path,
    )

    sample_rows = df.head(5).fillna("").to_dict(orient="records")
    return {
        "status": "success",
        "data": {
            "dataset_id": dataset_id,
            "filename": file.filename,
            "row_count": len(df),
            "column_count": len(df.columns),
            "columns": df.columns.tolist(),
            "numeric_columns": numeric_cols,
            "sample_rows": sample_rows,
            "numeric_profile": numeric_profile,
            "missing_values": missing_values,
            "categorical_profile": categorical_profile,
        },
    }


@router.get("/datasets/{dataset_id}")
def get_dataset_detail_route(dataset_id: int, user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    import json as _json
    row = get_dataset_by_id(dataset_id)
    if row is None or str(row.get("user_id", "")) != str(user.user_id):
        return JSONResponse(status_code=404, content=build_error_response("Dataset not found"))
    numeric_profile     = _json.loads(row["numeric_profile_json"])
    missing_values      = _json.loads(row["missing_values_json"])
    categorical_profile = _json.loads(row["categorical_profile_json"])

    def _safe_json(key: str):
        raw = row.get(key)
        if not raw:
            return None
        try:
            return _json.loads(raw)
        except Exception:
            return None

    def _load_sample_rows() -> list:
        try:
            fp = row.get("file_path")
            if not fp or not os.path.exists(fp):
                return []
            fp_lower = fp.lower()
            if fp_lower.endswith(".csv"):
                df = pd.read_csv(fp, nrows=5)
            elif fp_lower.endswith(".xlsx"):
                df = pd.read_excel(fp, engine="openpyxl", nrows=5)
            else:
                df = pd.read_excel(fp, engine="xlrd", nrows=5)
            return df.fillna("").to_dict(orient="records")
        except Exception:
            return []

    return {
        "status": "success",
        "data": {
            "dataset_id":          row["id"],
            "filename":            row["filename"],
            "row_count":           row["row_count"],
            "column_count":        row["column_count"],
            "uploaded_at":         row.get("uploaded_at"),
            "columns":             _json.loads(row["columns_json"]),
            "numeric_columns":     list(numeric_profile.keys()),
            "numeric_profile":     numeric_profile,
            "missing_values":      missing_values,
            "categorical_profile": categorical_profile,
            "correlation_profile": _safe_json("correlation_profile_json"),
            "date_profile":        _safe_json("date_profile_json"),
            "categorical_meta":    _safe_json("categorical_meta_json"),
            "semantic_profile":    _safe_json("semantic_profile_json"),
            "sample_rows":         _load_sample_rows(),
        },
    }


@router.get("/datasets")
def list_datasets(user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        datasets = list_datasets_for_user(str(user.user_id))
        return {"status": "success", "data": datasets}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


class PatchDatasetRequest(BaseModel):
    filename: str


@router.patch("/datasets/{dataset_id}")
def rename_dataset_route(dataset_id: int, request: PatchDatasetRequest, user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    filename = request.filename.strip()
    if not filename:
        return JSONResponse(status_code=400, content=build_error_response("filename must not be empty"))
    try:
        updated = rename_dataset(dataset_id, str(user.user_id), filename)
        if not updated:
            return JSONResponse(status_code=404, content=build_error_response("Dataset not found"))
        return {"status": "success", "data": {"dataset_id": dataset_id, "filename": filename}}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.delete("/datasets/{dataset_id}")
def delete_dataset_route(dataset_id: int, user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        deleted = delete_dataset(dataset_id, str(user.user_id))
        if not deleted:
            return JSONResponse(status_code=404, content=build_error_response("Dataset not found"))
        return {"status": "success", "data": {"deleted_id": dataset_id}}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/datasets/{dataset_id}/replace-source")
async def replace_dataset_source_route(
    dataset_id: int,
    file: UploadFile = File(...),
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    fname_lower = (file.filename or "").lower()
    if not any(fname_lower.endswith(ext) for ext in ALLOWED_DATASET_EXTENSIONS):
        return JSONResponse(
            status_code=400,
            content=build_error_response("Only CSV and Excel files (.csv, .xlsx, .xls) are supported"),
        )
    try:
        contents = await file.read()
        updated = replace_dataset_source_file(dataset_id, str(user.user_id), contents, file.filename)
        if updated is None:
            return JSONResponse(status_code=404, content=build_error_response("Dataset not found"))
        return {
            "status": "success",
            "data": {
                "dataset_id":   updated["id"],
                "filename":     updated["filename"],
                "row_count":    updated["row_count"],
                "column_count": updated["column_count"],
                "uploaded_at":  updated.get("uploaded_at"),
            },
        }
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/datasets/{dataset_id}/reprofile")
def reprofile_dataset_route(dataset_id: int, user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        updated = reprofile_dataset(dataset_id, str(user.user_id))
        if updated is None:
            return JSONResponse(
                status_code=404,
                content=build_error_response(
                    "Dataset not found or cannot be reprofiled. "
                    "The dataset may not have a stored source file (uploaded before Phase 1)."
                ),
            )
        return {
            "status": "success",
            "data": {
                "dataset_id":   updated["id"],
                "filename":     updated["filename"],
                "row_count":    updated["row_count"],
                "column_count": updated["column_count"],
                "uploaded_at":  updated.get("uploaded_at"),
                "file_path":    updated.get("file_path"),
            },
        }
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/scheduled-workflows")
def create_scheduled_workflow_route(
    request: CreateScheduledWorkflowRequest,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    if not request.input_text.strip():
        return JSONResponse(status_code=400, content=build_error_response("input_text cannot be empty"))
    try:
        from core.engine.schedule_parser import parse_schedule_intent

        # ── Try natural-language parser first ────────────────────────────────
        parsed_cron: str | None = None
        parsed_label: str | None = None
        parsed_frequency: str | None = None
        day_of_week: str | None = None
        task_type: str = "unknown"

        try:
            sched_meta = parse_schedule_intent(request.input_text)
            parsed_cron      = sched_meta["cron"]
            parsed_label     = sched_meta["human_label"]
            parsed_frequency = sched_meta["frequency"]
        except ValueError:
            # NL parser didn't recognise the phrase — fall back to interpret_task
            pass

        if parsed_frequency is None:
            # Fallback: use the existing task interpreter
            plan = interpret_task(request.input_text)
            schedule = plan.get("schedule")
            if not schedule or not schedule.get("frequency"):
                return JSONResponse(
                    status_code=400,
                    content=build_error_response(
                        "No recurring schedule detected. "
                        "Try phrases like 'every Friday at 9 AM', 'daily at 6 AM', or 'every 2 hours'."
                    ),
                )
            parsed_frequency = schedule["frequency"]
            day_of_week = plan.get("metadata", {}).get("entities", {}).get("day_of_week")
            task_type = plan.get("task_type", "unknown")
        else:
            plan = interpret_task(request.input_text)
            task_type = plan.get("task_type", "unknown")

        entry = create_scheduled_workflow(
            user_id=str(user.user_id),
            dataset_id=request.dataset_id,
            input_text=request.input_text,
            task_type=task_type,
            frequency=parsed_frequency,
            day_of_week=day_of_week,
            cron=parsed_cron,
            human_label=parsed_label,
            refresh_before_run=request.refresh_before_run,
        )
        return {"status": "success", "data": entry}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/scheduled-workflows")
def list_scheduled_workflows_route(user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        return {"status": "success", "data": list_scheduled_workflows(str(user.user_id))}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.patch("/scheduled-workflows/{workflow_id}/pause")
def pause_scheduled_workflow_route(
    workflow_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        updated = pause_scheduled_workflow(workflow_id, str(user.user_id))
        if not updated:
            return JSONResponse(status_code=404, content=build_error_response("Scheduled workflow not found"))
        return {"status": "success", "data": updated}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.patch("/scheduled-workflows/{workflow_id}/resume")
def resume_scheduled_workflow_route(
    workflow_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        updated = resume_scheduled_workflow(workflow_id, str(user.user_id))
        if not updated:
            return JSONResponse(status_code=404, content=build_error_response("Scheduled workflow not found"))
        return {"status": "success", "data": updated}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.delete("/scheduled-workflows/{workflow_id}")
def delete_scheduled_workflow_route(
    workflow_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        deleted = delete_scheduled_workflow(workflow_id, str(user.user_id))
        if not deleted:
            return JSONResponse(status_code=404, content=build_error_response("Scheduled workflow not found"))
        return {"status": "success", "data": {"deleted_id": workflow_id}}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/schedule-health")
def schedule_health_route(user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        return {"status": "success", "data": get_schedule_health(str(user.user_id))}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/executions/{execution_id}/retry")
def retry_execution(
    execution_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    from core.input.input_handler import handle_input
    from core.workflows.workflow_runner import run_workflow_by_id

    row = get_execution_by_id(execution_id, str(user.user_id))
    if row is None:
        return JSONResponse(status_code=404, content=build_error_response("Execution not found"))
    if row["status"] not in ("failed", "unknown"):
        return JSONResponse(
            status_code=400,
            content=build_error_response(
                f"Cannot retry an execution with status '{row['status']}'."
                " Use /rerun for completed executions."
            ),
        )
    try:
        intent = row.get("intent")
        workflow_id = row.get("workflow_id")
        if intent:
            return handle_input(intent, user_id=str(user.user_id))
        if workflow_id:
            data = run_workflow_by_id(workflow_id, user_id=str(user.user_id))
            return {"status": "success", "data": data}
        return JSONResponse(
            status_code=400,
            content=build_error_response("Cannot retry: no intent or workflow to replay."),
        )
    except ValueError as e:
        return JSONResponse(status_code=404, content=build_error_response(str(e)))
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/executions/{execution_id}/rerun")
def rerun_execution(
    execution_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    from core.input.input_handler import handle_input
    from core.workflows.workflow_runner import run_workflow_by_id

    row = get_execution_by_id(execution_id, str(user.user_id))
    if row is None:
        return JSONResponse(status_code=404, content=build_error_response("Execution not found"))
    if row["status"] not in ("completed", "success"):
        return JSONResponse(
            status_code=400,
            content=build_error_response(
                f"Cannot re-run an execution with status '{row['status']}'."
                " Use /retry for failed executions."
            ),
        )
    try:
        intent = row.get("intent")
        workflow_id = row.get("workflow_id")
        if intent:
            return handle_input(intent, user_id=str(user.user_id))
        if workflow_id:
            data = run_workflow_by_id(workflow_id, user_id=str(user.user_id))
            return {"status": "success", "data": data}
        return JSONResponse(
            status_code=400,
            content=build_error_response("Cannot re-run: no intent or workflow to replay."),
        )
    except ValueError as e:
        return JSONResponse(status_code=404, content=build_error_response(str(e)))
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/workflow-templates")
def workflow_templates_route(user: AuthenticatedUser = Depends(require_auth)) -> dict:
    return {"status": "success", "data": WORKFLOW_TEMPLATES}


# ---------------------------------------------------------------------------
# AI Operational Assistant
# ---------------------------------------------------------------------------

_ALLOWED_CONTEXT_TYPES = frozenset({
    "workflow_health", "schedule_health", "execution", "report",
})

_AI_ASSISTANT_SYSTEM_PROMPT = """\
You are an operational assistant for ToolSmithAI. Explain the provided operational data concisely.

Rules:
- Explain ONLY what the data shows. Do NOT suggest executing workflows, modifying data, \
creating schedules, or deleting anything.
- Be concise and actionable.
- Return exactly this JSON:
{
  "explanation": "<clear plain-English explanation, max 1000 characters>",
  "recommended_actions": ["<user-facing action>"],
  "confidence": "low" | "medium" | "high"
}
- recommended_actions: 1-5 items, each max 200 characters.
- confidence: "high" if definitive, "medium" if partial, "low" if ambiguous.
- Return ONLY the JSON. No markdown. No code fences.
"""


def _fmt_ms(ms: int | None) -> str:
    if ms is None:
        return "—"
    return f"{ms}ms" if ms < 1000 else f"{ms / 1000:.1f}s"


def _build_workflow_health_context(wf: dict) -> str:
    return "\n".join([
        f"Workflow: {wf.get('workflow_name') or ('Workflow #' + str(wf.get('workflow_id')))}",
        f"Health: {wf.get('health', '—')}",
        f"Total runs: {wf.get('total_runs', '—')}",
        f"Success rate: {round((wf.get('success_rate') or 0) * 100)}%",
        f"Consecutive failures: {wf.get('consecutive_failures', 0)}",
        f"Average duration: {_fmt_ms(wf.get('avg_duration_ms'))}",
        f"Last run: {wf.get('last_run') or '—'}",
        f"Recommendation: {wf.get('recommendation', '—')}",
    ])


def _build_schedule_health_context(sh: dict) -> str:
    return "\n".join([
        f"Task: {(sh.get('input_text') or '—')[:200]}",
        f"Frequency: {sh.get('frequency', '—')}",
        f"State: {'Enabled' if sh.get('enabled') else 'Paused'}",
        f"Health: {sh.get('health', '—')}",
        f"Overdue: {sh.get('overdue_label') or 'None'}",
        f"Next run: {sh.get('next_run_at') or '—'}",
        f"Last run: {sh.get('last_run_at') or '—'}",
        f"Last status: {sh.get('last_status') or '—'}",
        f"Recommendation: {sh.get('recommendation', '—')}",
    ])


def _build_execution_context(row: dict) -> str:
    lines = [
        f"Intent: {(row.get('intent') or '—')[:200]}",
        f"Task type: {row.get('task_type') or '—'}",
        f"Status: {row.get('status', '—')}",
        f"Duration: {row.get('duration_label') or '—'}",
        f"Source: {row.get('source_label') or row.get('trigger_source') or '—'}",
        f"Started: {row.get('started_at') or '—'}",
    ]
    if row.get("error_message"):
        lines.append(f"Error: {row['error_message'][:300]}")
    return "\n".join(lines)


def _build_report_sections_context(sections: list) -> str:
    lines: list[str] = []
    for sec in sections:
        heading = sec.get("heading") or ""
        if heading:
            lines.append(f"\n## {heading}")

        # Plain-text bullets (Overview, Numeric Insights, Missing Data, etc.)
        for item in sec.get("items") or []:
            lines.append(f"  - {item}")

        # KPI cards (kpi, business_kpis sections)
        for kpi in sec.get("kpis") or []:
            label = kpi.get("label") or ""
            value = kpi.get("value_formatted") or str(kpi.get("value", ""))
            desc  = kpi.get("description") or ""
            lines.append(f"  - {label}: {value}" + (f" ({desc})" if desc else ""))

        # Recommendations (recommendation section)
        for rec in sec.get("recommendations") or []:
            priority = rec.get("priority") or ""
            title    = rec.get("title") or ""
            reason   = rec.get("reason") or ""
            lines.append(f"  - [{priority}] {title}: {reason}")

        # Anomalies with evidence (anomaly section)
        for anom in sec.get("anomalies") or []:
            sev      = (anom.get("severity") or "").upper()
            title    = anom.get("title") or ""
            evidence = anom.get("evidence") or anom.get("description") or ""
            lines.append(f"  - [{sev}] {title}: {evidence}")

        # Executive summary fields
        if sec.get("summary"):
            lines.append(f"  {sec['summary']}")
        for tkw in sec.get("key_takeaways") or []:
            lines.append(f"  - Takeaway: {tkw}")
        for risk in sec.get("risks") or []:
            lines.append(f"  - Risk: {risk}")
        for opp in sec.get("opportunities") or []:
            lines.append(f"  - Opportunity: {opp}")

        # Chart sections — prose explanation + full categorical breakdown
        if sec.get("explanation"):
            lines.append(f"  {sec['explanation']}")
        chart  = sec.get("chart") or {}
        labels = chart.get("labels") or []
        series = chart.get("series") or []
        if labels and series:
            data  = (series[0].get("data") or []) if series else []
            total = sum(data) if data else 0
            for lbl, cnt in zip(labels, data):
                pct = f" ({round(cnt / total * 100)}%)" if total > 0 else ""
                lines.append(f"  - {lbl}: {cnt:,}{pct}")

        # Prioritised insights (insight_priority section)
        for ins in sec.get("insights") or []:
            text = ins.get("insight") or ins.get("title") or "" if isinstance(ins, dict) else str(ins)
            lines.append(f"  - {text}")

        # Trend objects (trend section)
        for tr in sec.get("trends") or []:
            title = tr.get("title") or ""
            desc  = tr.get("description") or ""
            lines.append(f"  - {title}: {desc}")

    return "\n".join(lines)


def _build_dataset_profile_context(dataset_row: dict) -> str:
    """Serialise a dataset's stored profile into a compact, GPT-readable context string.

    Reads only pre-computed profile summaries — never the raw source file.
    All JSON parsing is guarded so a missing or corrupt column produces no output
    rather than raising.
    """
    import json as _json

    def _safe_parse(key: str) -> dict | list | None:
        raw = dataset_row.get(key)
        if not raw:
            return None
        try:
            return _json.loads(raw)
        except Exception:
            return None

    lines: list[str] = []
    filename     = dataset_row.get("filename") or "dataset"
    row_count    = dataset_row.get("row_count") or 0
    column_count = dataset_row.get("column_count") or 0
    lines.append(f"## Dataset: {filename} ({row_count:,} rows, {column_count} columns)")

    numeric_profile     = _safe_parse("numeric_profile_json") or {}
    categorical_profile = _safe_parse("categorical_profile_json") or {}
    missing_values      = _safe_parse("missing_values_json") or {}
    categorical_meta    = _safe_parse("categorical_meta_json") or {}

    # Numeric columns — key stats only; skip histogram bins
    if numeric_profile:
        lines.append("\n### Numeric Columns")
        for col, stats in numeric_profile.items():
            if not isinstance(stats, dict):
                continue
            mean    = stats.get("mean")
            median  = stats.get("median")
            mn      = stats.get("min")
            mx      = stats.get("max")
            nn      = stats.get("non_null_count")
            null_c  = stats.get("null_count", 0)
            parts: list[str] = []
            if mean    is not None: parts.append(f"mean={mean}")
            if median  is not None: parts.append(f"median={median}")
            if mn      is not None: parts.append(f"min={mn}")
            if mx      is not None: parts.append(f"max={mx}")
            if nn      is not None: parts.append(f"non_null={nn:,}")
            if null_c:              parts.append(f"missing={null_c:,}")
            lines.append(f"  {col}: " + ", ".join(parts))

    # Categorical columns — top-5 value counts with % of total rows
    if categorical_profile:
        lines.append("\n### Categorical Columns")
        for col, top_vals in categorical_profile.items():
            if not isinstance(top_vals, list):
                continue
            meta         = categorical_meta.get(col) or {}
            unique_count = meta.get("unique_count")
            label        = f"{col} ({unique_count} unique)" if unique_count is not None else col
            lines.append(f"  {label}:")
            for entry in top_vals:
                if not isinstance(entry, dict):
                    continue
                val = entry.get("value", "")
                cnt = entry.get("count", 0)
                pct = f" ({round(cnt / row_count * 100)}%)" if row_count > 0 else ""
                lines.append(f"    - {val}: {cnt:,}{pct}")

    # Missing values — only columns that have gaps
    cols_with_missing = {
        col: cnt for col, cnt in missing_values.items()
        if isinstance(cnt, int) and cnt > 0
    }
    if cols_with_missing:
        lines.append("\n### Missing Values")
        for col, cnt in cols_with_missing.items():
            pct = f" ({round(cnt / row_count * 100)}%)" if row_count > 0 else ""
            lines.append(f"  - {col}: {cnt:,} missing{pct}")

    return "\n".join(lines)


def _deterministic_explain(context_type: str, data: dict) -> dict:
    """Return a safe, useful explanation without any AI call."""
    if context_type == "workflow_health":
        health = data.get("health", "Unknown")
        rec    = data.get("recommendation") or "Review the workflow configuration."
        explanation = f"This workflow is {health}. {rec}"
        actions = ["Review recent execution history for detailed error information."]
        if health in ("Critical", "Warning"):
            actions.append("Check that all required tools and data sources are accessible.")
        if (data.get("consecutive_failures") or 0) >= 2:
            actions.append("Consider pausing this workflow until the root cause is resolved.")

    elif context_type == "schedule_health":
        health = data.get("health", "Unknown")
        rec    = data.get("recommendation") or "Review the schedule configuration."
        explanation = f"This scheduled workflow is {health}. {rec}"
        actions = {
            "Missed":  ["Check if the server was running during the expected execution window.",
                        "Verify the scheduler service is active."],
            "Delayed": ["The scheduler should process this on the next tick.",
                        "Verify the scheduler service is running."],
            "Paused":  ["Resume the schedule when you are ready to restart automatic runs."],
            "Healthy": ["Schedule is operating normally. No action needed."],
        }.get(health, ["Review the schedule configuration."])

    elif context_type == "execution":
        status = data.get("status", "unknown")
        error  = (data.get("error_message") or "")[:200]
        if status == "failed":
            explanation = f"This execution failed.{' Error: ' + error if error else ''}"
            actions = ["Retry from the History tab.",
                       "Ensure required data (e.g. uploaded CSV) is available."]
        elif status in ("completed", "success"):
            explanation = "This execution completed successfully."
            actions = ["Re-run from the History tab if needed."]
        else:
            explanation = f"Execution status: {status}."
            actions = []

    else:  # report
        explanation = (
            "This dataset report was generated from your uploaded CSV file. "
            "Review each section for insights."
        )
        actions = [
            "Check Numeric Insights for key statistics.",
            "Check Missing Data for data quality issues.",
        ]

    return {
        "explanation":         explanation[:1000],
        "recommended_actions": actions[:5],
        "confidence":          "medium",
    }


def _validate_ai_explanation(raw: dict) -> dict:
    if not isinstance(raw, dict):
        raise ValueError("response must be a dict")
    explanation = raw.get("explanation")
    if not isinstance(explanation, str) or not explanation.strip():
        raise ValueError("explanation must be a non-empty string")
    actions = raw.get("recommended_actions", [])
    if not isinstance(actions, list):
        raise ValueError("recommended_actions must be a list")
    confidence = raw.get("confidence", "medium")
    if confidence not in ("low", "medium", "high"):
        confidence = "medium"
    return {
        "explanation":         explanation.strip()[:1000],
        "recommended_actions": [str(a).strip()[:200] for a in actions[:5] if str(a).strip()],
        "confidence":          confidence,
    }


def _ai_explain_context(context_summary: str, question: str | None) -> dict | None:
    """
    Ask the AI assistant to explain the provided operational context.
    The AI receives only sanitized text — no raw DB rows, secrets, or credentials.
    Returns None on any failure; caller always falls back to _deterministic_explain.
    """
    import json as _json
    import logging as _logging
    _logger = _logging.getLogger(__name__)

    from core.config import (
        ENABLE_AI_ASSISTANT, OPENAI_API_KEY, OPENAI_MODEL, OPENAI_TIMEOUT_SECONDS,
    )
    if not ENABLE_AI_ASSISTANT or not OPENAI_API_KEY:
        return None

    try:
        import openai as _openai
    except ImportError:
        _logger.warning("[ai_assistant] openai not installed; using deterministic fallback")
        return None

    user_msg = context_summary
    if question:
        user_msg += f"\n\nQuestion: {question[:500]}"

    try:
        client = _openai.OpenAI(api_key=OPENAI_API_KEY, timeout=OPENAI_TIMEOUT_SECONDS)
        resp   = client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=[
                {"role": "system", "content": _AI_ASSISTANT_SYSTEM_PROMPT},
                {"role": "user",   "content": user_msg},
            ],
            max_tokens=500,
            temperature=0.2,
            response_format={"type": "json_object"},
        )
        content = resp.choices[0].message.content or ""
        raw     = _json.loads(content.strip())
        result  = _validate_ai_explanation(raw)
        _logger.info("[ai_assistant] success: confidence=%s", result["confidence"])
        return result
    except Exception as exc:
        _logger.warning(
            "[ai_assistant] failed (%s: %s); using deterministic fallback",
            type(exc).__name__, exc,
        )
        return None


class ExplainRequest(BaseModel):
    context_type: str
    context_id:   int | None = None
    question:     str | None = None


@router.post("/assistant/explain")
def assistant_explain(
    request: ExplainRequest,
    user:    AuthenticatedUser = Depends(require_jwt),
) -> dict:
    if request.context_type not in _ALLOWED_CONTEXT_TYPES:
        return JSONResponse(
            status_code=400,
            content=build_error_response(
                f"context_type must be one of: {', '.join(sorted(_ALLOWED_CONTEXT_TYPES))}"
            ),
        )

    context_data:    dict = {}
    context_summary: str  = ""

    try:
        if request.context_type == "workflow_health":
            rows   = get_workflow_success_insights(str(user.user_id))
            target = next((r for r in rows if r["workflow_id"] == request.context_id), None)
            if target is None:
                return JSONResponse(status_code=404, content=build_error_response("Workflow not found"))
            context_data    = target
            context_summary = _build_workflow_health_context(target)

        elif request.context_type == "schedule_health":
            rows   = get_schedule_health(str(user.user_id))
            target = next((r for r in rows if r["id"] == request.context_id), None)
            if target is None:
                return JSONResponse(status_code=404, content=build_error_response("Scheduled workflow not found"))
            context_data    = target
            context_summary = _build_schedule_health_context(target)

        elif request.context_type == "execution":
            row = get_execution_by_id(request.context_id, str(user.user_id))
            if row is None:
                return JSONResponse(status_code=404, content=build_error_response("Execution not found"))
            row             = enrich_execution_record(row)
            context_data    = row
            context_summary = _build_execution_context(row)

        elif request.context_type == "report":
            from data.dataset_service import get_dataset_by_id
            from core.tools.report_generator import generate_dataset_report
            if request.context_id is None:
                return JSONResponse(status_code=400, content=build_error_response("context_id required for report"))
            dataset = get_dataset_by_id(request.context_id)
            if dataset is None or str(dataset.get("user_id", "")) != str(user.user_id):
                return JSONResponse(status_code=404, content=build_error_response("Dataset not found"))
            report          = generate_dataset_report(dataset)
            context_data    = {"dataset_id": request.context_id, "filename": dataset.get("filename")}
            context_summary = _build_report_sections_context(report.get("sections", []))

    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))

    ai_result = _ai_explain_context(context_summary, request.question)
    if ai_result:
        return {"status": "success", "data": {**ai_result, "source": "ai"}}

    fallback = _deterministic_explain(request.context_type, context_data)
    return {"status": "success", "data": {**fallback, "source": "standard"}}


@router.get("/report-metric-snapshots")
def list_recent_metric_snapshots_route(
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Return the 50 most recent metric snapshots for the authenticated user."""
    try:
        from data.report_metric_snapshot_service import list_recent_snapshots_for_user
        snapshots = list_recent_snapshots_for_user(str(user.user_id), limit=50)
        return {"status": "success", "data": snapshots, "count": len(snapshots)}
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content=build_error_response("Internal server error", str(e)),
        )


@router.get("/datasets/{dataset_id}/metric-snapshots")
def list_dataset_metric_snapshots_route(
    dataset_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Return up to 20 metric snapshots for one dataset owned by the authenticated user."""
    try:
        from data.report_metric_snapshot_service import list_snapshots_for_dataset
        snapshots = list_snapshots_for_dataset(str(user.user_id), dataset_id, limit=20)
        return {"status": "success", "data": snapshots, "count": len(snapshots)}
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content=build_error_response("Internal server error", str(e)),
        )


@router.get("/reports")
def list_reports_route(user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        reports = list_reports_for_user(str(user.user_id))
        return {"status": "success", "data": reports}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/reports/{report_id}")
def get_report_route(report_id: int, user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        report = get_report_by_id(report_id, str(user.user_id))
        if report is None:
            return JSONResponse(status_code=404, content=build_error_response("Report not found"))
        return {"status": "success", "data": report}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


class AskReportRequest(BaseModel):
    question: str


_ASK_REPORT_SYSTEM_PROMPT = """\
You are an AI analyst answering questions about a saved business report.
You have access to two sources: the report sections and the dataset profile provided below.
Do NOT invent data or make up statistics. Use only what is explicitly stated in the provided content.
Return exactly this JSON:
{
  "answer": "<clear, concise answer based solely on the report sections and dataset profile, max 600 characters>",
  "cited_sections": ["<section heading 1>", "<up to 3 section headings cited>"],
  "confidence": "low" | "medium" | "high"
}
- If the question cannot be answered from the provided content, say so in the answer field.
- confidence: "high" if the answer is directly in the text, "medium" if inferred, "low" if uncertain.
- Return ONLY the JSON. No markdown. No code fences.
"""


def _ai_answer_report_question(
    sections: list,
    question: str,
    dataset_row: dict | None = None,
) -> dict | None:
    """Ask GPT to answer a question using report sections and, when available, the dataset profile.

    Safety: GPT receives only sanitized section text and pre-computed profile summaries —
    never raw source files, never secrets.
    Returns None on any failure so the caller falls back to the deterministic answer.
    """
    import json as _json
    import logging as _logging
    _log = _logging.getLogger(__name__)

    from core.config import (
        ENABLE_AI_REPORT_NARRATIVE, OPENAI_API_KEY, OPENAI_MODEL, OPENAI_TIMEOUT_SECONDS,
    )
    if not ENABLE_AI_REPORT_NARRATIVE or not OPENAI_API_KEY:
        return None

    try:
        import openai as _openai
    except ImportError:
        return None

    sections_context = _build_report_sections_context(sections)
    profile_context  = _build_dataset_profile_context(dataset_row) if dataset_row else ""

    if profile_context:
        context = f"{profile_context}\n\n---\n\n## Report Sections\n{sections_context}"
    else:
        context = sections_context

    user_msg = f"Report content:\n\n{context}\n\nQuestion: {question[:500]}"

    try:
        client = _openai.OpenAI(api_key=OPENAI_API_KEY, timeout=OPENAI_TIMEOUT_SECONDS)
        resp = client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=[
                {"role": "system", "content": _ASK_REPORT_SYSTEM_PROMPT},
                {"role": "user", "content": user_msg},
            ],
            max_tokens=400,
            temperature=0.2,
            response_format={"type": "json_object"},
        )
        raw = _json.loads(resp.choices[0].message.content or "")
        if not isinstance(raw, dict) or not raw.get("answer"):
            return None
        return {
            "answer":              str(raw["answer"]).strip()[:600],
            "cited_sections_used": [str(s) for s in (raw.get("cited_sections") or [])[:3]],
            "confidence":          str(raw.get("confidence", "medium")) if raw.get("confidence") in ("low", "medium", "high") else "medium",
        }
    except Exception as exc:
        _log.warning("[ask_report] ai failed: %s: %s", type(exc).__name__, exc)
        return None


@router.post("/reports/{report_id}/ask")
def ask_report_route(
    report_id: int,
    request: AskReportRequest,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Ask a natural-language question about a saved report.

    GPT answers from report content only — no invented data, no secrets exposed.
    Falls back to a deterministic message when AI is disabled or unavailable.
    """
    question = (request.question or "").strip()
    if not question:
        return JSONResponse(status_code=400, content=build_error_response("question must not be empty"))

    try:
        report = get_report_by_id(report_id, str(user.user_id))
        if report is None:
            return JSONResponse(status_code=404, content=build_error_response("Report not found"))

        sections = (report.get("content") or {}).get("sections", [])

        dataset_id  = report.get("dataset_id")
        dataset_row = get_dataset_by_id(dataset_id) if dataset_id else None

        ai_result = _ai_answer_report_question(sections, question, dataset_row=dataset_row)
        if ai_result:
            return {
                "status": "success",
                "data": {
                    **ai_result,
                    "enhanced_mode": True,
                    "report_id": report_id,
                },
            }

        return {
            "status": "success",
            "data": {
                "answer":              "In-depth Q&A is not available for this report. Please review the report sections for insights.",
                "cited_sections_used": [],
                "confidence":          "low",
                "enhanced_mode":       False,
                "report_id":           report_id,
            },
        }
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.delete("/reports/{report_id}")
def delete_report_route(report_id: int, user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        deleted = delete_report(report_id, str(user.user_id))
        if not deleted:
            return JSONResponse(status_code=404, content=build_error_response("Report not found"))
        return {"status": "success", "data": {"deleted_id": report_id}}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


# ─── Dataset Ask AI ───────────────────────────────────────────────────────────

_ASK_DATASET_SYSTEM_PROMPT = """\
You are a data analyst assistant for ToolSmithAI.
You are given a dataset profile containing pre-computed statistical summaries only — never raw rows.
Answer the user's question factually and concisely using only what is present in the profile.
Return exactly this JSON:
{
  "answer": "<plain-English answer based solely on the profile, max 600 characters>",
  "confidence": "low" | "medium" | "high",
  "suggested_actions": ["<imperative composer-ready phrase, max 80 characters>"]
}
- suggested_actions: 1–3 items. Each should be a short prompt the user could run next in the workspace.
- confidence: "high" if directly derivable from the profile, "medium" if inferred, "low" if uncertain.
- Do NOT invent statistics. Do NOT reference raw CSV data or file contents.
- Return ONLY the JSON. No markdown. No code fences.
"""


def _ai_answer_dataset_question(
    dataset_row: dict,
    question: str,
    composer_text: str | None = None,
) -> dict | None:
    """Answer a question about a dataset using its pre-computed profile.

    Safety: GPT receives only sanitized profile summaries — never raw CSV rows or secrets.
    Returns None on any failure so the caller falls back to the deterministic answer.
    """
    import json as _json
    import logging as _logging
    _log = _logging.getLogger(__name__)

    from core.config import (
        ENABLE_AI_REPORT_NARRATIVE, OPENAI_API_KEY, OPENAI_MODEL, OPENAI_TIMEOUT_SECONDS,
    )
    if not ENABLE_AI_REPORT_NARRATIVE or not OPENAI_API_KEY:
        return None

    try:
        import openai as _openai
    except ImportError:
        return None

    profile_context = _build_dataset_profile_context(dataset_row)
    user_msg = f"Dataset profile:\n\n{profile_context}"
    if composer_text and composer_text.strip():
        user_msg += f"\n\nUser is currently composing: {composer_text.strip()[:200]}"
    user_msg += f"\n\nQuestion: {question[:500]}"

    try:
        client = _openai.OpenAI(api_key=OPENAI_API_KEY, timeout=OPENAI_TIMEOUT_SECONDS)
        resp = client.chat.completions.create(
            model=OPENAI_MODEL,
            messages=[
                {"role": "system", "content": _ASK_DATASET_SYSTEM_PROMPT},
                {"role": "user",   "content": user_msg},
            ],
            max_tokens=400,
            temperature=0.2,
            response_format={"type": "json_object"},
        )
        raw = _json.loads(resp.choices[0].message.content or "")
        if not isinstance(raw, dict) or not raw.get("answer"):
            return None
        actions = [
            str(a)[:80]
            for a in (raw.get("suggested_actions") or [])[:3]
            if isinstance(a, str) and a.strip()
        ]
        return {
            "answer":            str(raw["answer"]).strip()[:600],
            "confidence":        str(raw.get("confidence", "medium")) if raw.get("confidence") in ("low", "medium", "high") else "medium",
            "suggested_actions": actions,
        }
    except Exception as exc:
        _log.warning("[ask_dataset] ai failed: %s: %s", type(exc).__name__, exc)
        return None


def _deterministic_dataset_answer(dataset_row: dict) -> dict:
    """Return a useful answer from the stored profile without any AI call."""
    import json as _json

    filename     = dataset_row.get("filename") or "dataset"
    row_count    = dataset_row.get("row_count") or 0
    column_count = dataset_row.get("column_count") or 0

    def _safe_parse(key: str) -> dict:
        raw = dataset_row.get(key)
        if not raw:
            return {}
        try:
            return _json.loads(raw)
        except Exception:
            return {}

    missing_values    = _safe_parse("missing_values_json")
    cols_with_missing = [c for c, v in missing_values.items() if isinstance(v, int) and v > 0]

    answer = f"{filename} has {row_count:,} rows and {column_count} columns."
    if cols_with_missing:
        listed = ", ".join(cols_with_missing[:3])
        answer += f" {len(cols_with_missing)} column(s) have missing values: {listed}."

    actions = [
        f"Analyze {filename}",
        "Generate a dataset report",
    ]
    if cols_with_missing:
        actions.append(f"Identify missing values in {cols_with_missing[0]}")

    return {
        "answer":            answer[:600],
        "confidence":        "low",
        "suggested_actions": actions[:3],
    }


class AskDatasetRequest(BaseModel):
    question:      str
    composer_text: str | None = None


@router.post("/datasets/{dataset_id}/ask")
def ask_dataset_route(
    dataset_id: int,
    request:    AskDatasetRequest,
    user:       AuthenticatedUser = Depends(require_jwt),
) -> dict:
    """Ask a natural-language question about a dataset using its pre-computed profile.

    GPT receives only sanitized profile summaries — never raw CSV rows or secrets.
    Falls back to a deterministic answer when AI is disabled or unavailable.
    """
    question = (request.question or "").strip()
    if not question:
        return JSONResponse(status_code=400, content=build_error_response("question must not be empty"))

    try:
        row = get_dataset_by_id(dataset_id)
        if row is None or str(row.get("user_id", "")) != str(user.user_id):
            return JSONResponse(status_code=404, content=build_error_response("Dataset not found"))

        ai_result = _ai_answer_dataset_question(row, question, composer_text=request.composer_text)
        if ai_result:
            return {
                "status": "success",
                "data": {**ai_result, "dataset_id": dataset_id, "enhanced_mode": True},
            }

        fallback = _deterministic_dataset_answer(row)
        return {
            "status": "success",
            "data": {**fallback, "dataset_id": dataset_id, "enhanced_mode": False},
        }
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


_XLSX_BRANDING = {
    'brand_name':    'ToolSmithAI',
    'primary':       '6366F1',
    'primary_dark':  '4338CA',
    'primary_light': 'E0E7FF',
    'success':       '10B981',
    'warning':       'F59E0B',
    'danger':        'F87171',
    'neutral':       '64748B',
    'text_dark':     '1E2A3A',
    'text_muted':    '64748B',
    'bg_white':      'FFFFFF',
    'bg_row_alt':    'F8FAFC',
}


def _build_xlsx_bytes(report: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle

    b        = _XLSX_BRANDING
    title    = report.get('title', 'Untitled Report')
    dataset  = report.get('dataset_filename') or 'Not specified'
    created  = (report.get('created_at') or '')[:19].replace('T', ' ')
    stamp    = datetime.now(timezone.utc).strftime('%b %d, %Y %H:%M UTC')
    brand    = b['brand_name']
    sections = (report.get('content') or {}).get('sections', [])

    by_type: dict = {}
    for sec in sections:
        t = sec.get('type', 'text')
        if t not in by_type:
            by_type[t] = sec

    wb = Workbook()
    wb.remove(wb.active)

    wb.properties.title       = title
    wb.properties.creator     = brand
    wb.properties.description = 'Generated by ' + brand
    wb.properties.created     = datetime.now(timezone.utc)

    def _ns(name, bold=False, size=9, color='000000', fill_hex=None,
            h='left', v='center', wrap=False):
        style = NamedStyle(name=name)
        style.font = Font(bold=bold, size=size, color=color)
        if fill_hex:
            style.fill = PatternFill('solid', fgColor=fill_hex)
        style.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        wb.add_named_style(style)

    _ns('xl_heading',  bold=True,  size=14, color='FFFFFF',          fill_hex=b['primary'],       h='center')
    _ns('xl_subtitle', bold=True,  size=11, color=b['text_dark'],                                  h='center')
    _ns('xl_colhdr',   bold=True,  size=9,  color=b['primary_dark'], fill_hex=b['primary_light'],  h='center')
    _ns('xl_label',    bold=True,  size=9,  color=b['neutral'],                                    h='right',  v='top')
    _ns('xl_data',                 size=9,  color=b['text_dark'],                                  h='left',   v='top', wrap=True)
    _ns('xl_data_alt',             size=9,  color=b['text_dark'],    fill_hex=b['bg_row_alt'],     h='left',   v='top', wrap=True)
    _ns('xl_good',     bold=True,  size=9,  color='065F46',          fill_hex='D1FAE5',            h='center')
    _ns('xl_warning',  bold=True,  size=9,  color='92400E',          fill_hex='FEF3C7',            h='center')
    _ns('xl_risk',     bold=True,  size=9,  color='991B1B',          fill_hex='FEE2E2',            h='center')
    _ns('xl_neutral',              size=9,  color=b['text_muted'],   fill_hex='F1F5F9',            h='center')

    def _widths(ws, widths):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    def _status_sty(s):
        v = (s or '').lower()
        if v in ('good', 'healthy', 'on_track', 'success'):
            return 'xl_good'
        if v in ('warning', 'caution', 'at_risk'):
            return 'xl_warning'
        if v in ('bad', 'critical', 'danger', 'risk', 'error'):
            return 'xl_risk'
        return 'xl_neutral'

    def _severity_sty(s):
        v = (s or '').lower()
        if v == 'high':   return 'xl_risk'
        if v == 'medium': return 'xl_warning'
        return 'xl_neutral'

    def _priority_sty(s):
        v = (s or '').lower()
        if v == 'high':   return 'xl_risk'
        if v == 'medium': return 'xl_warning'
        return 'xl_neutral'

    TREND = {'up': 'Up', 'down': 'Down', 'stable': 'Stable', 'volatile': 'Volatile'}

    def _fmt_val(val, fmt):
        try:
            if fmt == 'currency': return f'${float(val):,.0f}'
            if fmt == 'percent':  return f'{float(val):.1f}%'
            if fmt == 'number':   return f'{float(val):,.2f}'
        except Exception:
            pass
        return str(val) if val is not None else ''

    # ── Cover sheet (always index 0) ──────────────────────────────────────────
    def _sheet_cover():
        ws = wb.create_sheet('Cover')
        ws.sheet_properties.tabColor = b['primary']
        ws.sheet_view.showGridLines  = False
        ws['A1'] = brand
        ws.merge_cells('A1:F1')
        ws['A1'].style = 'xl_heading'
        ws.row_dimensions[1].height = 32
        ws['A2'] = title
        ws.merge_cells('A2:F2')
        ws['A2'].style = 'xl_subtitle'
        ws.row_dimensions[2].height = 24
        for i, (lbl, val) in enumerate([
            ('Dataset',   dataset),
            ('Generated', stamp),
            ('Created',   created or '-'),
        ], start=4):
            ws.cell(row=i, column=1, value=lbl).style = 'xl_label'
            ws.cell(row=i, column=2, value=val).style = 'xl_data'
        _widths(ws, {'A': 14, 'B': 52, 'C': 12, 'D': 12, 'E': 12, 'F': 12})

    # ── Summary Dashboard (always index 1) ────────────────────────────────────
    def _sheet_summary():
        ws = wb.create_sheet('Summary Dashboard')
        ws.sheet_properties.tabColor = b['primary_dark']
        ws.sheet_view.showGridLines  = False
        ws['A1'] = 'Summary Dashboard'
        ws.merge_cells('A1:D1')
        ws['A1'].style = 'xl_heading'
        ws.row_dimensions[1].height = 28
        row = 3
        kpi_sec = by_type.get('kpi')
        if kpi_sec:
            ws.cell(row=row, column=1, value='Key Metrics Snapshot').style = 'xl_colhdr'
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            for kpi in (kpi_sec.get('kpis') or [])[:3]:
                ws.cell(row=row, column=1, value=kpi.get('label', '')).style = 'xl_label'
                ws.cell(
                    row=row, column=2,
                    value=_fmt_val(kpi.get('value'), kpi.get('format', '')),
                ).style = _status_sty(kpi.get('status', ''))
                ws.cell(row=row, column=3, value=TREND.get(kpi.get('trend', ''), '')).style = 'xl_neutral'
                row += 1
            row += 1
        anom_sec = by_type.get('anomaly')
        if anom_sec:
            anoms = anom_sec.get('anomalies') or []
            ws.cell(row=row, column=1, value='Anomalies Detected').style = 'xl_label'
            ws.cell(row=row, column=2, value=len(anoms)).style = (
                'xl_risk' if anoms else 'xl_good'
            )
            row += 2
        rec_sec = by_type.get('recommendation')
        if rec_sec:
            recs = rec_sec.get('recommendations') or []
            ws.cell(row=row, column=1, value='Recommendations').style = 'xl_label'
            ws.cell(row=row, column=2, value=len(recs)).style = 'xl_neutral'
        _widths(ws, {'A': 28, 'B': 20, 'C': 14, 'D': 28})

    # ── Executive Summary sheet ───────────────────────────────────────────────
    def _sheet_exec_summary(sec):
        ws = wb.create_sheet('Executive Summary')
        ws.sheet_properties.tabColor = b['primary']
        ws.sheet_view.showGridLines  = False
        ws['A1'] = sec.get('heading', 'Executive Summary')
        ws.merge_cells('A1:C1')
        ws['A1'].style = 'xl_heading'
        ws.row_dimensions[1].height = 28
        row = 3
        summary = sec.get('summary', '')
        if summary:
            ws.cell(row=row, column=1, value='Summary').style = 'xl_label'
            ws.cell(row=row, column=2, value=summary).style = 'xl_data'
            ws.merge_cells(f'B{row}:C{row}')
            ws.row_dimensions[row].height = 60
            row += 2
        for group_key, group_label in [
            ('key_takeaways', 'Key Takeaway'),
            ('risks',         'Risk'),
            ('opportunities', 'Opportunity'),
        ]:
            items = sec.get(group_key) or []
            for item in items:
                ws.cell(row=row, column=1, value=group_label).style = 'xl_label'
                ws.cell(row=row, column=2, value=str(item)).style = 'xl_data'
                ws.merge_cells(f'B{row}:C{row}')
                row += 1
            if items:
                row += 1
        _widths(ws, {'A': 18, 'B': 60, 'C': 20})

    # ── Key Metrics sheet ─────────────────────────────────────────────────────
    def _sheet_kpi(sec):
        ws = wb.create_sheet('Key Metrics')
        ws.sheet_properties.tabColor = b['success']
        ws.sheet_view.showGridLines  = False
        ws['A1'] = sec.get('heading', 'Key Metrics')
        ws.merge_cells('A1:F1')
        ws['A1'].style = 'xl_heading'
        ws.row_dimensions[1].height = 28
        for col, h in enumerate(
            ['Metric', 'Value', 'Format', 'Trend', 'Status', 'Description'], 1
        ):
            ws.cell(row=2, column=col, value=h).style = 'xl_colhdr'
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = 'A2:F2'
        row = 3
        for i, kpi in enumerate(sec.get('kpis') or []):
            sty = 'xl_data' if i % 2 == 0 else 'xl_data_alt'
            ws.cell(row=row, column=1, value=kpi.get('label', '')).style = sty
            ws.cell(
                row=row, column=2,
                value=_fmt_val(kpi.get('value'), kpi.get('format', '')),
            ).style = _status_sty(kpi.get('status', ''))
            ws.cell(row=row, column=3, value=kpi.get('format', '')).style = sty
            ws.cell(
                row=row, column=4,
                value=TREND.get(kpi.get('trend', ''), kpi.get('trend', '')),
            ).style = sty
            ws.cell(row=row, column=5, value=kpi.get('status', '')).style = _status_sty(kpi.get('status', ''))
            ws.cell(row=row, column=6, value=kpi.get('description', '')).style = sty
            row += 1
        _widths(ws, {'A': 28, 'B': 16, 'C': 12, 'D': 10, 'E': 12, 'F': 40})

    # ── Anomalies & Risks sheet ───────────────────────────────────────────────
    def _sheet_anomaly(sec):
        ws = wb.create_sheet('Anomalies & Risks')
        ws.sheet_properties.tabColor = b['danger']
        ws.sheet_view.showGridLines  = False
        ws['A1'] = sec.get('heading', 'Anomalies & Risks')
        ws.merge_cells('A1:E1')
        ws['A1'].style = 'xl_heading'
        ws.row_dimensions[1].height = 28
        for col, h in enumerate(
            ['Title', 'Severity', 'Category', 'Description', 'Evidence'], 1
        ):
            ws.cell(row=2, column=col, value=h).style = 'xl_colhdr'
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = 'A2:E2'
        row = 3
        for i, anom in enumerate(sec.get('anomalies') or []):
            sty = 'xl_data' if i % 2 == 0 else 'xl_data_alt'
            ws.cell(row=row, column=1, value=anom.get('title', '')).style = sty
            ws.cell(row=row, column=2, value=anom.get('severity', '')).style = _severity_sty(anom.get('severity', ''))
            ws.cell(row=row, column=3, value=anom.get('category', '')).style = sty
            ws.cell(row=row, column=4, value=anom.get('description', '')).style = sty
            ws.cell(row=row, column=5, value=anom.get('evidence', '')).style = sty
            row += 1
        _widths(ws, {'A': 30, 'B': 12, 'C': 18, 'D': 50, 'E': 30})

    # ── Recommendations sheet ─────────────────────────────────────────────────
    def _sheet_recommendation(sec):
        ws = wb.create_sheet('Recommendations')
        ws.sheet_properties.tabColor = b['warning']
        ws.sheet_view.showGridLines  = False
        ws['A1'] = sec.get('heading', 'Recommendations')
        ws.merge_cells('A1:E1')
        ws['A1'].style = 'xl_heading'
        ws.row_dimensions[1].height = 28
        for col, h in enumerate(
            ['Title', 'Priority', 'Action Type', 'Confidence', 'Reason'], 1
        ):
            ws.cell(row=2, column=col, value=h).style = 'xl_colhdr'
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = 'A2:E2'
        row = 3
        for i, rec in enumerate(sec.get('recommendations') or []):
            sty = 'xl_data' if i % 2 == 0 else 'xl_data_alt'
            ws.cell(row=row, column=1, value=rec.get('title', '')).style = sty
            ws.cell(row=row, column=2, value=rec.get('priority', '')).style = _priority_sty(rec.get('priority', ''))
            ws.cell(row=row, column=3, value=rec.get('action_type', '')).style = sty
            ws.cell(row=row, column=4, value=rec.get('confidence', '')).style = sty
            ws.cell(row=row, column=5, value=rec.get('reason', '')).style = sty
            row += 1
        _widths(ws, {'A': 32, 'B': 12, 'C': 18, 'D': 14, 'E': 50})

    # ── Build workbook — sheet order is fixed ─────────────────────────────────
    _sheet_cover()    # index 0
    _sheet_summary()  # index 1
    if 'executive_summary' in by_type:
        _sheet_exec_summary(by_type['executive_summary'])
    if 'kpi' in by_type:
        _sheet_kpi(by_type['kpi'])
    if 'anomaly' in by_type:
        _sheet_anomaly(by_type['anomaly'])
    if 'recommendation' in by_type:
        _sheet_recommendation(by_type['recommendation'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

_EXPORT_FORMATS = {"json", "pdf", "csv", "xlsx"}


def _get_client_ip(request: Request) -> str | None:
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()[:64]
    if request.client:
        return request.client.host
    return None


@router.get("/reports/{report_id}/export")
def export_report_route(
    report_id: int,
    request: Request,
    format: str = Query(default="json"),
    user: AuthenticatedUser = Depends(require_jwt),
) -> Response:
    if format not in _EXPORT_FORMATS:
        return JSONResponse(
            status_code=400,
            content=build_error_response(
                f"Unsupported export format '{format}'. "
                f"Supported formats: {', '.join(sorted(_EXPORT_FORMATS))}"
            ),
        )
    _ip = _get_client_ip(request)
    _ua = (request.headers.get("User-Agent") or "")[:512]
    try:
        report = get_report_by_id(report_id, str(user.user_id))
        if report is None:
            return JSONResponse(status_code=404, content=build_error_response("Report not found"))

        safe = "".join(
            c if (c.isalnum() or c in " .-_") else "_"
            for c in report.get("title", "")
        )[:60].strip().replace(" ", "_")

        if format == "pdf":
            filename   = f"{safe}_report.pdf" if safe.strip("_") else f"report_{report_id}.pdf"
            content    = _build_pdf_bytes(report)
            media_type = "application/pdf"
        elif format == "csv":
            import csv, io
            filename   = f"{safe}_report.csv" if safe.strip("_") else f"report_{report_id}.csv"
            buf = io.StringIO()
            w = csv.writer(buf)
            # Metadata block at the top so the file is self-contained
            w.writerow(["# title",            report.get("title", "")])
            w.writerow(["# task_type",         report.get("task_type", "")])
            w.writerow(["# dataset_filename",  report.get("dataset_filename") or ""])
            w.writerow(["# created_at",        (report.get("created_at") or "")[:19].replace("T", " ")])
            w.writerow([])
            w.writerow(["section", "item"])
            for section in (report.get("content") or {}).get("sections", []):
                sec_type = section.get("type", "text")
                heading  = section.get("heading", "")
                if sec_type == "text":
                    for item in section.get("items", []):
                        w.writerow([heading, item])
                elif sec_type == "kpi":
                    for kpi in section.get("kpis", []):
                        try:
                            w.writerow([
                                heading,
                                kpi.get("label", ""),
                                kpi.get("value", ""),
                                kpi.get("format", ""),
                                kpi.get("trend", ""),
                                kpi.get("description", ""),
                            ])
                        except Exception:
                            pass
                elif sec_type == "executive_summary":
                    summary = section.get("summary", "")
                    if summary:
                        try:
                            w.writerow([heading, "summary", summary])
                        except Exception:
                            pass
                    for item in section.get("key_takeaways", []):
                        try:
                            w.writerow([heading, "key_takeaway", item])
                        except Exception:
                            pass
                    for item in section.get("risks", []):
                        try:
                            w.writerow([heading, "risk", item])
                        except Exception:
                            pass
                    for item in section.get("opportunities", []):
                        try:
                            w.writerow([heading, "opportunity", item])
                        except Exception:
                            pass
                elif sec_type == "recommendation":
                    for rec in section.get("recommendations", []):
                        try:
                            w.writerow([
                                heading,
                                rec.get("priority",    ""),
                                rec.get("action_type", ""),
                                rec.get("confidence",  ""),
                                rec.get("title",       ""),
                                rec.get("reason",      ""),
                            ])
                        except Exception:
                            pass
                elif sec_type == "chart":
                    chart      = section.get("chart", {})
                    chart_type = chart.get("chart_type", "bar")
                    labels     = chart.get("labels", [])
                    for s_entry in chart.get("series", []):
                        s_name = s_entry.get("name", "")
                        s_data = s_entry.get("data", [])
                        for i, label in enumerate(labels):
                            try:
                                val = s_data[i] if i < len(s_data) else ""
                                w.writerow([heading, chart_type, s_name, label, val])
                            except Exception:
                                pass
                elif sec_type == "drift_detection":
                    for drift in section.get("drifts", []):
                        try:
                            w.writerow([
                                heading,
                                drift.get("metric",         ""),
                                drift.get("baseline_value", ""),
                                drift.get("current_value",  ""),
                                drift.get("drift_percent",  ""),
                                drift.get("severity",       ""),
                                drift.get("direction",      ""),
                                drift.get("description",    ""),
                            ])
                        except Exception:
                            pass
                elif sec_type == "historical_comparison":
                    for comp in section.get("comparisons", []):
                        try:
                            w.writerow([
                                heading,
                                comp.get("metric",         ""),
                                comp.get("current_value",  ""),
                                comp.get("previous_value", ""),
                                comp.get("change",         ""),
                                comp.get("change_type",    ""),
                                comp.get("severity",       ""),
                                comp.get("description",    ""),
                            ])
                        except Exception:
                            pass
                elif sec_type == "predictive_readiness":
                    pr_level = section.get("readiness_level", "")
                    pr_score = section.get("readiness_score", "")
                    for sig in section.get("signals", []):
                        try:
                            w.writerow([
                                heading,
                                pr_level,
                                pr_score,
                                sig.get("name",        ""),
                                sig.get("status",      ""),
                                sig.get("description", ""),
                                sig.get("evidence",    ""),
                            ])
                        except Exception:
                            pass
                    for step in section.get("next_steps", []):
                        try:
                            w.writerow([heading, "next_step", step])
                        except Exception:
                            pass
                elif sec_type == "trend":
                    for trend in section.get("trends", []):
                        try:
                            w.writerow([
                                heading,
                                trend.get("direction",   ""),
                                trend.get("strength",    ""),
                                trend.get("category",    ""),
                                trend.get("title",       ""),
                                trend.get("description", ""),
                                trend.get("evidence",    ""),
                            ])
                        except Exception:
                            pass
                elif sec_type == "anomaly":
                    for anomaly in section.get("anomalies", []):
                        try:
                            w.writerow([
                                heading,
                                anomaly.get("severity",    ""),
                                anomaly.get("category",    ""),
                                anomaly.get("title",       ""),
                                anomaly.get("description", ""),
                                anomaly.get("evidence",    ""),
                            ])
                        except Exception:
                            pass
                # Unknown future types: omitted until a row format is defined.
            content    = buf.getvalue().encode("utf-8-sig")  # utf-8-sig adds BOM for Excel
            media_type = "text/csv"
        elif format == 'xlsx':
            safe_prefix = safe if safe.strip('_') else ('report_' + str(report_id))
            filename   = 'toolsmithai_' + safe_prefix + '_report.xlsx'
            content    = _build_xlsx_bytes(report)
            media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            import json as _json
            filename   = f"{safe}_report.json" if safe.strip("_") else f"report_{report_id}.json"
            payload = {
                "export_format": "json",
                "exported_at":   datetime.now(timezone.utc).isoformat(),
                "report": {
                    "id":               report["id"],
                    "title":            report["title"],
                    "task_type":        report["task_type"],
                    "status":           report["status"],
                    "dataset_filename": report.get("dataset_filename"),
                    "created_at":       report["created_at"],
                    "content":          report.get("content"),
                },
            }
            content    = _json.dumps(payload, indent=2).encode()
            media_type = "application/json"

        try:
            create_export_log(
                user_id=str(user.user_id),
                report_id=report_id,
                export_format=format,
                filename=filename,
                file_size_bytes=len(content),
                status="success",
                ip_address=_ip,
                user_agent=_ua,
            )
        except Exception:
            pass

        return Response(
            content=content,
            media_type=media_type,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        try:
            create_export_log(
                user_id=str(user.user_id),
                report_id=report_id,
                export_format=format,
                status="failed",
                error_reason=str(e)[:500],
                ip_address=_ip,
                user_agent=_ua,
            )
        except Exception:
            pass
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/admin/export-logs")
def list_admin_export_logs_route(
    user: AuthenticatedUser = Depends(require_role("admin")),
    limit: int = Query(default=100, le=500),
    offset: int = Query(default=0, ge=0),
    export_format: str | None = None,
    status: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    user_id: str | None = None,
) -> dict:
    try:
        _fmt    = export_format or None
        _status = status or None
        _dfrom  = date_from or None
        _dto    = date_to or None
        _uid    = user_id or None
        logs = list_all_export_logs(
            limit=limit,
            offset=offset,
            export_format=_fmt,
            status=_status,
            date_from=_dfrom,
            date_to=_dto,
            user_id=_uid,
        )
        total = count_all_export_logs(
            export_format=_fmt,
            status=_status,
            date_from=_dfrom,
            date_to=_dto,
            user_id=_uid,
        )
        return {"status": "success", "data": logs, "count": len(logs), "total": total}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/admin/export-logs/summary")
def admin_export_logs_summary_route(
    user: AuthenticatedUser = Depends(require_role("admin")),
) -> dict:
    try:
        summary = get_export_log_summary()
        return {"status": "success", "data": summary}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/admin/email-logs")
def list_admin_email_logs_route(
    user: AuthenticatedUser = Depends(require_role("admin")),
    limit: int = Query(default=100, le=500),
    offset: int = Query(default=0, ge=0),
    email_type: str | None = None,
    status: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    user_id: str | None = None,
) -> dict:
    try:
        _type   = email_type or None
        _status = status or None
        _dfrom  = date_from or None
        _dto    = date_to or None
        _uid    = user_id or None
        logs = list_all_email_logs(
            limit=limit,
            offset=offset,
            email_type=_type,
            status=_status,
            date_from=_dfrom,
            date_to=_dto,
            user_id=_uid,
        )
        total = count_all_email_logs(
            email_type=_type,
            status=_status,
            date_from=_dfrom,
            date_to=_dto,
            user_id=_uid,
        )
        return {"status": "success", "data": logs, "count": len(logs), "total": total}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/admin/email-logs/summary")
def admin_email_logs_summary_route(
    user: AuthenticatedUser = Depends(require_role("admin")),
) -> dict:
    try:
        summary = get_email_log_summary()
        return {"status": "success", "data": summary}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


class EmailReportRequest(BaseModel):
    recipient_email: str


@router.post("/reports/{report_id}/email")
def email_report_route(
    report_id: int,
    request: EmailReportRequest,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    to = request.recipient_email.strip()
    if not to:
        return JSONResponse(
            status_code=400,
            content=build_error_response("recipient_email is required"),
        )
    try:
        from core.email import send_real_email
        from core.tools.report_generator import (
            render_report_as_html_email,
            render_report_as_plain_text,
        )
        report = get_report_by_id(report_id, str(user.user_id))
        if report is None:
            return JSONResponse(status_code=404, content=build_error_response("Report not found"))
        subject      = f"ToolSmithAI Intelligence Report — {report.get('title', 'Report')}"
        _content     = report.get("content") or {}
        _title       = report.get("title", "")
        _dataset     = report.get("dataset_filename") or ""
        _report_url  = f"{FRONTEND_BASE_URL}/reports/{report_id}"
        body         = render_report_as_plain_text(
            _content,
            title=_title,
            dataset_filename=_dataset,
            report_url=_report_url,
        )
        html_body    = render_report_as_html_email(
            _content,
            title=_title,
            dataset_filename=_dataset,
            report_url=_report_url,
        )
        _pdf_bytes: bytes | None = None
        _pdf_filename: str | None = None
        try:
            _safe = "".join(
                c if (c.isalnum() or c in " .-_") else "_"
                for c in (_title or _dataset)
            )[:60].strip().replace(" ", "_")
            _pdf_bytes    = _build_pdf_bytes(report)
            _pdf_filename = f"{_safe}_report.pdf" if _safe.strip("_") else f"report_{report_id}.pdf"
        except Exception as _pdf_exc:
            import logging as _log_mod
            _log_mod.getLogger(__name__).warning(
                "[email_report] PDF generation failed — sending without attachment: %s", _pdf_exc
            )
        result       = send_real_email(to=to, subject=subject, body=body,
                                       html_body=html_body,
                                       attachment_bytes=_pdf_bytes,
                                       attachment_filename=_pdf_filename,
                                       user_id=str(user.user_id), report_id=report_id,
                                       email_type="report")
        sent     = result["sent"]
        reason   = result.get("reason", "")
        # ENABLE_REAL_EMAIL=false returns sent=False with a "disabled" reason.
        # Treat that as simulated success so the UI reflects the intended action.
        simulated = not sent and "disabled" in reason.lower()
        return {
            "status": "success",
            "data": {
                "sent":      sent or simulated,
                "to":        to,
                "message": (
                    f"Report emailed to {to}" if sent
                    else f"Report emailed to {to}." if simulated
                    else reason or "Email could not be sent"
                ),
            },
        }
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/schedules/runs")
def list_all_schedule_runs_route(user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        runs = list_recent_runs_for_user(str(user.user_id))
        return {"status": "success", "data": runs}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/schedules/{schedule_id}/runs")
def list_schedule_runs_route(
    schedule_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    wf = get_scheduled_workflow_by_id(schedule_id)
    if wf is None or str(wf["user_id"]) != str(user.user_id):
        return JSONResponse(status_code=404, content=build_error_response("Schedule not found"))
    try:
        runs = list_runs_for_schedule(schedule_id, str(user.user_id))
        return {"status": "success", "data": runs}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/schedules/{schedule_id}/run-now")
def run_schedule_now_route(
    schedule_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    from core.input.input_handler import handle_input

    wf = get_scheduled_workflow_by_id(schedule_id)
    if wf is None or str(wf["user_id"]) != str(user.user_id):
        return JSONResponse(status_code=404, content=build_error_response("Schedule not found"))

    run_id = None
    try:
        run_id = create_schedule_run(schedule_id, str(user.user_id), trigger_type="manual")
    except Exception:
        pass

    try:
        result = handle_input(
            wf["input_text"],
            user_id=str(user.user_id),
            dataset_id=wf.get("dataset_id"),
        )
        status, warn_msg = classify_schedule_result(result)
        update_scheduled_workflow_outcome(schedule_id, status=status, error=warn_msg)

        if run_id is not None:
            try:
                related_report_id = None
                data = result.get("data", {}) if isinstance(result, dict) else {}
                related_report_id = data.get("report_id")
                complete_schedule_run(run_id, related_report_id=related_report_id)
            except Exception:
                pass

        return {
            "status": "success",
            "data": result.get("data") if isinstance(result, dict) else result,
        }
    except Exception as exc:
        err_msg = str(exc)[:500]
        try:
            update_scheduled_workflow_outcome(schedule_id, status="failed", error=err_msg)
        except Exception:
            pass
        if run_id is not None:
            try:
                fail_schedule_run(run_id, error_message=err_msg)
            except Exception:
                pass
        return JSONResponse(
            status_code=500,
            content=build_error_response("Run failed", err_msg),
        )


@router.get("/notifications")
def list_notifications_route(user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    try:
        notifications = list_notifications_for_user(str(user.user_id))
        return {"status": "success", "data": notifications}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.post("/notifications/{notification_id}/read")
def mark_notification_read_route(
    notification_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        updated = mark_notification_read(notification_id, str(user.user_id))
        if not updated:
            return JSONResponse(status_code=404, content=build_error_response("Notification not found"))
        return {"status": "success", "data": {"id": notification_id, "read": True}}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.delete("/notifications/{notification_id}")
def delete_notification_route(
    notification_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        deleted = delete_notification(notification_id, str(user.user_id))
        if not deleted:
            return JSONResponse(status_code=404, content=build_error_response("Notification not found"))
        return {"status": "success", "data": {"deleted_id": notification_id}}
    except Exception as e:
        return JSONResponse(status_code=500, content=build_error_response("Internal server error", str(e)))


@router.get("/health")
def health() -> dict:
    return {"status": "healthy"}


# ---------------------------------------------------------------------------
# Data Sources  (/v1/sources)
# ---------------------------------------------------------------------------

class DataSourceCreateRequest(BaseModel):
    display_name: str
    source_type: str
    config: dict = {}
    metadata: dict = {}


@router.post("/sources")
def create_source(
    body: DataSourceCreateRequest,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        record = create_data_source(
            user_id=user.user_id,
            payload={
                "display_name": body.display_name,
                "source_type": body.source_type,
                "config": body.config,
                "metadata": body.metadata,
            },
        )
    except ValueError as exc:
        return JSONResponse(status_code=400, content=build_error_response(str(exc)))
    return {"status": "success", "data": record}


@router.get("/sources")
def list_sources(user: AuthenticatedUser = Depends(require_jwt)) -> dict:
    return {"status": "success", "data": list_data_sources(user.user_id)}


@router.delete("/sources/{source_id}")
def delete_source(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = delete_data_source(source_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=409, content=build_error_response(str(exc)))
    except Exception as exc:
        logger.exception("delete_source failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Data source deletion failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.post("/sources/{source_id}/test")
def test_source(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    result = test_data_source(source_id, user.user_id)
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return result


@router.post("/sources/{source_id}/discover")
def discover_source_schema(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = run_discovery(source_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        return JSONResponse(status_code=500, content=build_error_response("Schema discovery failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/schema")
def get_source_schema(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    result = get_latest_snapshot(source_id, user.user_id)
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("No schema snapshot found."))
    return {"status": "success", "data": result}


@router.post("/sources/{source_id}/dictionary/generate")
def generate_dictionary_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = generate_and_save_dictionary(source_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        return JSONResponse(status_code=500, content=build_error_response("Dictionary generation failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/dictionary")
def list_dictionary_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    result = list_dictionary_tables(source_id, user.user_id)
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/dictionary/tables/{table_fqn:path}")
def get_table_dictionary_route(
    source_id: int,
    table_fqn: str,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    result = get_table_dictionary(source_id, user.user_id, table_fqn)
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Table dictionary entry not found."))
    return {"status": "success", "data": result}


@router.post("/sources/{source_id}/dictionary/tables/{table_fqn}/approve")
def approve_table_route(
    source_id: int,
    table_fqn: str,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    result = approve_table_dictionary(source_id, user.user_id, table_fqn)
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Dictionary table entry not found."))
    return {"status": "success", "data": result}


@router.post("/sources/{source_id}/dictionary/tables/{table_fqn}/columns/{column_name}/approve")
def approve_column_route(
    source_id: int,
    table_fqn: str,
    column_name: str,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    result = approve_column_dictionary(source_id, user.user_id, table_fqn, column_name)
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Dictionary column entry not found."))
    return {"status": "success", "data": result}


# ---------------------------------------------------------------------------
# Metadata Profiling  (/v1/sources/{id}/profile/...)
# ---------------------------------------------------------------------------

@router.post("/sources/{source_id}/profile/structural")
def run_structural_profile_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = run_structural_profiling(source_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        return JSONResponse(status_code=500, content=build_error_response("Structural profiling failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/profile")
def get_profile_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    result = get_latest_profile(source_id, user.user_id)
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("No profiling data found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/profile/history")
def get_profile_history_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    result = list_profile_history(source_id, user.user_id)
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.post("/sources/{source_id}/profile/full")
def run_full_profile_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = run_full_profiling(source_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        return JSONResponse(status_code=500, content=build_error_response("Full profiling failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.post("/sources/{source_id}/profile/batch/start")
def start_batch_profile_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = start_batch_profiling(source_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        return JSONResponse(status_code=500, content=build_error_response("Batch profiling start failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": {
        "profiling_snapshot_id":  result.profiling_snapshot_id,
        "next_table_index":        result.next_table_index,
        "total_tables":            result.total_tables,
        "status":                  result.status.value,
    }}


@router.post("/sources/{source_id}/profile/batch/{profiling_snapshot_id}/continue")
def continue_batch_profile_route(
    source_id: int,
    profiling_snapshot_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = continue_batch_profiling(source_id, user.user_id, profiling_snapshot_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        return JSONResponse(status_code=500, content=build_error_response("Batch profiling step failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Profiling snapshot not found."))
    return {"status": "success", "data": {
        "profiling_snapshot_id":          result.profiling_snapshot_id,
        "next_table_index":                result.next_table_index,
        "total_tables":                    result.total_tables,
        "completed_tables":                result.completed_tables,
        "statistical_tables_completed":    result.statistical_tables_completed,
        "structural_tables_completed":     result.structural_tables_completed,
        "is_complete":                     result.status.value == "COMPLETE",
        "status":                          result.status.value,
    }}


# ---------------------------------------------------------------------------
# Domain Detection  (/v1/sources/{id}/domains/...)
# ---------------------------------------------------------------------------

@router.post("/sources/{source_id}/domains/generate")
def generate_domains_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = generate_domain_assignments(source_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        logger.exception("generate_domains_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Domain generation failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/domains")
def list_domains_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = list_domain_assignments(source_id, user.user_id)
    except Exception:
        logger.exception("list_domains_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Failed to retrieve domain assignments."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/domains/summary")
def get_domain_summary_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = get_domain_summary(source_id, user.user_id)
    except Exception:
        logger.exception("get_domain_summary_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Failed to retrieve domain summary."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


# ---------------------------------------------------------------------------
# Entity Detection  (/v1/sources/{id}/entities/...)
# ---------------------------------------------------------------------------

@router.post("/sources/{source_id}/entities/generate")
def generate_entities_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = generate_entity_assignments(source_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        logger.exception("generate_entities_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Entity generation failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/entities")
def list_entities_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = list_entity_assignments(source_id, user.user_id)
    except Exception:
        logger.exception("list_entities_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Failed to retrieve entity assignments."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/entities/summary")
def get_entity_summary_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = get_entity_summary(source_id, user.user_id)
    except Exception:
        logger.exception("get_entity_summary_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Failed to retrieve entity summary."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


# ---------------------------------------------------------------------------
# Domain Learning Rules  (/v1/sources/{id}/domain-rules  &  /v1/domain-rules/{id}/...)
# ---------------------------------------------------------------------------

@router.post("/sources/{source_id}/domain-rules/suggest")
def suggest_domain_rules_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = generate_domain_rule_suggestions(source_id, user.user_id)
    except Exception:
        logger.exception("suggest_domain_rules_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Domain rule suggestion failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/domain-rules/suggestions")
def list_domain_rule_suggestions_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = list_domain_rule_suggestions(source_id, user.user_id)
    except Exception:
        logger.exception("list_domain_rule_suggestions_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Failed to retrieve domain rule suggestions."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/domain-rules")
def list_domain_rules_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = list_domain_rules(source_id, user.user_id)
    except Exception:
        logger.exception("list_domain_rules_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Failed to retrieve domain rules."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.post("/domain-rules/{rule_id}/approve")
def approve_domain_rule_route(
    rule_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = approve_domain_rule(rule_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        logger.exception("approve_domain_rule_route failed for rule_id=%s", rule_id)
        return JSONResponse(status_code=500, content=build_error_response("Domain rule approval failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Domain rule not found."))
    return {"status": "success", "data": result}


@router.post("/domain-rules/{rule_id}/reject")
def reject_domain_rule_route(
    rule_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = reject_domain_rule(rule_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        logger.exception("reject_domain_rule_route failed for rule_id=%s", rule_id)
        return JSONResponse(status_code=500, content=build_error_response("Domain rule rejection failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Domain rule not found."))
    return {"status": "success", "data": result}


# ---------------------------------------------------------------------------
# Entity Learning Rules  (/v1/sources/{id}/entity-rules  &  /v1/entity-rules/{id}/...)
# ---------------------------------------------------------------------------

@router.post("/sources/{source_id}/entity-rules/suggest")
def suggest_entity_rules_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = generate_entity_rule_suggestions(source_id, user.user_id)
    except Exception:
        logger.exception("suggest_entity_rules_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Entity rule suggestion failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/entity-rules/suggestions")
def list_entity_rule_suggestions_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = list_entity_rule_suggestions(source_id, user.user_id)
    except Exception:
        logger.exception("list_entity_rule_suggestions_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Failed to retrieve entity rule suggestions."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/entity-rules")
def list_entity_rules_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = list_entity_rules(source_id, user.user_id)
    except Exception:
        logger.exception("list_entity_rules_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Failed to retrieve entity rules."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.post("/entity-rules/{rule_id}/approve")
def approve_entity_rule_route(
    rule_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = approve_entity_rule(rule_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        logger.exception("approve_entity_rule_route failed for rule_id=%s", rule_id)
        return JSONResponse(status_code=500, content=build_error_response("Entity rule approval failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Entity rule not found."))
    return {"status": "success", "data": result}


@router.post("/entity-rules/{rule_id}/reject")
def reject_entity_rule_route(
    rule_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = reject_entity_rule(rule_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        logger.exception("reject_entity_rule_route failed for rule_id=%s", rule_id)
        return JSONResponse(status_code=500, content=build_error_response("Entity rule rejection failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Entity rule not found."))
    return {"status": "success", "data": result}


# ---------------------------------------------------------------------------
# Domain Rule Refinements  (/v1/sources/{id}/domain-refinements  &
#                           /v1/domain-refinements/{id}/...)
# ---------------------------------------------------------------------------

@router.post("/sources/{source_id}/domain-refinements/analyze")
def analyze_domain_refinements_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = analyze_rule_refinement(source_id, user.user_id)
    except Exception:
        logger.exception("analyze_domain_refinements_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Domain refinement analysis failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.get("/sources/{source_id}/domain-refinements")
def list_domain_refinements_route(
    source_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = list_refinement_suggestions(source_id, user.user_id)
    except Exception:
        logger.exception("list_domain_refinements_route failed for source_id=%s", source_id)
        return JSONResponse(status_code=500, content=build_error_response("Failed to retrieve refinement suggestions."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Data source not found."))
    return {"status": "success", "data": result}


@router.post("/domain-refinements/{suggestion_id}/approve")
def approve_domain_refinement_route(
    suggestion_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = approve_refinement_suggestion(suggestion_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        logger.exception("approve_domain_refinement_route failed for suggestion_id=%s", suggestion_id)
        return JSONResponse(status_code=500, content=build_error_response("Refinement suggestion approval failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Refinement suggestion not found."))
    return {"status": "success", "data": result}


@router.post("/domain-refinements/{suggestion_id}/reject")
def reject_domain_refinement_route(
    suggestion_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    try:
        result = reject_refinement_suggestion(suggestion_id, user.user_id)
    except ValueError as exc:
        return JSONResponse(status_code=422, content=build_error_response(str(exc)))
    except Exception:
        logger.exception("reject_domain_refinement_route failed for suggestion_id=%s", suggestion_id)
        return JSONResponse(status_code=500, content=build_error_response("Refinement suggestion rejection failed."))
    if result is None:
        return JSONResponse(status_code=404, content=build_error_response("Refinement suggestion not found."))
    return {"status": "success", "data": result}


# ---------------------------------------------------------------------------
# Metadata Jobs  (/v1/metadata-jobs)
# ---------------------------------------------------------------------------

@router.post("/metadata-jobs/{job_id}/run")
def run_metadata_job_route(
    job_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    if get_metadata_job(job_id, user.user_id) is None:
        return JSONResponse(status_code=404, content=build_error_response("Metadata job not found."))
    try:
        run_metadata_job(job_id)
    except Exception:
        return JSONResponse(status_code=500, content=build_error_response("Metadata job execution failed."))
    updated = get_metadata_job(job_id, user.user_id)
    return {"status": "success", "data": updated}


@router.get("/metadata-jobs/{job_id}")
def get_metadata_job_route(
    job_id: int,
    user: AuthenticatedUser = Depends(require_jwt),
) -> dict:
    job = get_metadata_job(job_id, user.user_id)
    if job is None:
        return JSONResponse(status_code=404, content=build_error_response("Metadata job not found."))
    return {"status": "success", "data": job}
