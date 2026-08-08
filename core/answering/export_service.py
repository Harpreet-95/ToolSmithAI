"""
Day 4, Capability 5 — Export.

Stateless: builds an .xlsx/.csv/.pdf file directly from an already-computed
EnterpriseAnswer — the caller round-trips the exact dict its own prior
POST /v1/composer/ask response returned (api/v1/composer.py::composer_export
never re-executes SQL, never looks up an execution_id; core.live.query_engine
.LiveQueryEngine itself "does not persist results and does not cache", so
there is nothing to look up after the fact even if it wanted to).

Mirrors the branding conventions already established for report export
(core.tools.report_generator._build_pdf_bytes, api/v1/routes.py
._build_xlsx_bytes) and reuses data.export_log_service's existing audit
table (report_id=None for a Q&A export) — without sharing code with the
report exporter, since a live Q&A answer's export shape (one question, one
natural-language answer, a bounded result table) is a different, much
smaller document than a multi-section dataset-analysis report.
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from typing import Any

_BRAND_RGB = {
    "primary": (99, 102, 241), "primary_dark": (67, 56, 202),
    "text_dark": (30, 42, 58), "text_muted": (100, 116, 139),
    "success": (16, 185, 129), "danger": (248, 113, 113), "warning": (245, 158, 11),
    "rule_light": (210, 220, 235),
}
_BRAND_HEX = {
    "primary": "6366F1", "primary_dark": "4338CA", "primary_light": "E0E7FF",
    "text_dark": "1E2A3A", "text_muted": "64748B", "bg_row_alt": "F8FAFC",
}

ALLOWED_FORMATS = {"xlsx", "csv", "pdf"}
_TERMINAL_SUCCESS_AGENT_STATUS = "answered"
_NON_EXPORTABLE_ANSWER_TYPES = {"clarification_needed"}
_MAX_TABLE_ROWS_PDF = 40  # a PDF is a one-pager-oriented document; xlsx/csv carry the full result_preview


def check_exportable(enterprise_answer: "dict | None", agent_status: "str | None") -> "str | None":
    """Returns None when export is allowed, or a short user-facing reason
    when it must be refused. The one gate every export entry point goes
    through — a hard backend check, not just a frontend affordance, so a
    refused/clarification-required/otherwise-unsuccessful answer can never
    be exported regardless of what the caller sends."""
    if not enterprise_answer:
        return "No answer is available to export."
    if enterprise_answer.get("answer_type") in _NON_EXPORTABLE_ANSWER_TYPES:
        return "This question needs clarification before it can be exported."
    if agent_status is not None and agent_status != _TERMINAL_SUCCESS_AGENT_STATUS:
        return "This question was not successfully answered, so there is nothing to export."
    return None


def _resolve_table(enterprise_answer: dict) -> "tuple[list[str], list[dict]]":
    """(columns, rows) for the Result Table — reuses result_preview
    verbatim, in the SAME order core.answering.result_formatter already
    produced (never re-sorted/re-grouped/re-filtered here). Falls back to a
    single synthesized row from actual_value for a scalar answer, whose
    result_preview is empty by design — still the exact verified value,
    never invented."""
    preview = enterprise_answer.get("result_preview") or []
    if preview:
        columns: list[str] = []
        for row in preview:
            for k in row.keys():
                if k not in columns:
                    columns.append(k)
        return columns, preview

    if enterprise_answer.get("actual_value") is not None:
        label = enterprise_answer.get("measure") or enterprise_answer.get("business_entity") or "Value"
        return [label], [{label: enterprise_answer["actual_value"]}]

    return [], []


def _fmt_filters(enterprise_answer: dict) -> list[str]:
    lines = []
    for f in (enterprise_answer.get("applied_filters") or []):
        value = f.get("value")
        value_str = " – ".join(str(v) for v in value) if isinstance(value, list) else str(value)
        lines.append(f"{f.get('label', '')} {f.get('operator', '')} {value_str}".strip())
    return lines


def _insight_chip_text(enterprise_answer: dict) -> "str | None":
    """Mirrors the exact chip text ChartSection/AIWorkspace.jsx already
    renders ("▲ +200% vs. the previous period") — reuses percent_change/
    direction verbatim, never re-derives them."""
    insight = enterprise_answer.get("insight")
    if not insight:
        return None
    pct = insight.get("percent_change")
    direction = insight.get("direction")
    label = insight.get("label") or "vs. the previous period"
    if pct is None or direction is None:
        return None
    arrow = "▲" if direction == "up" else "▼" if direction == "down" else "▪"
    sign = "+" if pct > 0 else ""
    return f"{arrow} {sign}{pct:g}% {label}"


def _safe_filename(question: str, ext: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9 ._-]", "_", question or "answer")[:60].strip().replace(" ", "_")
    return f"toolsmithai_{safe or 'answer'}.{ext}" if safe.strip("_") else f"toolsmithai_answer.{ext}"


# ---------------------------------------------------------------------------
# CSV — result table only
# ---------------------------------------------------------------------------

def build_csv_bytes(enterprise_answer: dict) -> bytes:
    columns, rows = _resolve_table(enterprise_answer)
    buf = io.StringIO()
    w = csv.writer(buf)
    if columns:
        w.writerow(columns)
        for row in rows:
            w.writerow([row.get(c, "") for c in columns])
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens UTF-8 cleanly


# ---------------------------------------------------------------------------
# Excel — Business Summary sheet + Result Table sheet
# ---------------------------------------------------------------------------

def build_xlsx_bytes(question: str, enterprise_answer: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill

    b = _BRAND_HEX
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = question[:200] or "ToolSmithAI Answer"
    wb.properties.creator = "ToolSmithAI"
    wb.properties.created = datetime.now(timezone.utc)

    def _ns(name, bold=False, size=9, color="000000", fill_hex=None, h="left", v="top", wrap=True):
        style = NamedStyle(name=name)
        style.font = Font(bold=bold, size=size, color=color)
        if fill_hex:
            style.fill = PatternFill("solid", fgColor=fill_hex)
        style.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        wb.add_named_style(style)

    _ns("xl_heading", bold=True, size=14, color="FFFFFF", fill_hex=b["primary"], h="center", wrap=False)
    _ns("xl_label", bold=True, size=9, color=b["text_muted"], h="left")
    _ns("xl_data", size=10, color=b["text_dark"], h="left")
    _ns("xl_colhdr", bold=True, size=9, color=b["primary_dark"], fill_hex=b["primary_light"], h="center", wrap=False)
    _ns("xl_cell", size=9, color=b["text_dark"], h="left", wrap=False)
    _ns("xl_cell_alt", size=9, color=b["text_dark"], fill_hex=b["bg_row_alt"], h="left", wrap=False)

    # ── Business Summary ────────────────────────────────────────────────────
    ws = wb.create_sheet("Business Summary")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "ToolSmithAI — Business Answer"
    ws.merge_cells("A1:B1")
    ws["A1"].style = "xl_heading"
    ws.row_dimensions[1].height = 26

    row = 3

    def _kv(label: str, value: str, height: "int | None" = None) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label).style = "xl_label"
        ws.cell(row=row, column=2, value=value).style = "xl_data"
        if height:
            ws.row_dimensions[row].height = height
        row += 1

    _kv("Question", question or "", height=30)
    _kv("Answer", enterprise_answer.get("answer") or "", height=30)
    insight_text = _insight_chip_text(enterprise_answer)
    if insight_text:
        _kv("Business Insight", insight_text)
    if enterprise_answer.get("date_context"):
        _kv("Time Period", enterprise_answer["date_context"].get("label") or "")
    filters = _fmt_filters(enterprise_answer)
    if filters:
        _kv("Applied Filters", "; ".join(filters))
    _kv("Confidence", f"{round(enterprise_answer.get('confidence') or 0)}%")
    governance = enterprise_answer.get("governance_warnings") or []
    if governance:
        _kv("Governance", "; ".join(str(g) for g in governance), height=24)
    citations = enterprise_answer.get("citations") or []
    if citations:
        _kv("Citations", "; ".join(c.get("label", "") for c in citations if isinstance(c, dict)), height=24)
    limitations = enterprise_answer.get("limitations") or []
    if limitations:
        _kv("Limitations", "; ".join(str(x) for x in limitations), height=24)
    source_tables = enterprise_answer.get("source_tables") or []
    if source_tables:
        _kv("Source table(s)", ", ".join(source_tables))
    _kv("Generated", datetime.now(timezone.utc).strftime("%b %d, %Y %H:%M UTC"))

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80

    # ── Result Table ─────────────────────────────────────────────────────────
    columns, rows = _resolve_table(enterprise_answer)
    ws2 = wb.create_sheet("Result Table")
    ws2.sheet_view.showGridLines = False
    if columns:
        for col_i, col_name in enumerate(columns, start=1):
            ws2.cell(row=1, column=col_i, value=col_name).style = "xl_colhdr"
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:{chr(64 + min(len(columns), 26))}1"
        for r_i, data_row in enumerate(rows, start=2):
            sty = "xl_cell" if r_i % 2 == 0 else "xl_cell_alt"
            for col_i, col_name in enumerate(columns, start=1):
                ws2.cell(row=r_i, column=col_i, value=data_row.get(col_name, "")).style = sty
        for col_i in range(1, len(columns) + 1):
            ws2.column_dimensions[chr(64 + col_i) if col_i <= 26 else "A"].width = 22
    truncation_notice = enterprise_answer.get("truncation_notice")
    if truncation_notice:
        ws2.cell(row=len(rows) + 3, column=1, value=truncation_notice).style = "xl_label"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# PDF — question, answer, insight, chart (simple bar rendering), table
# ---------------------------------------------------------------------------

def _register_fonts(pdf) -> str:
    """Same cross-platform font-loading fallback chain as
    core.tools.report_generator._build_pdf_bytes, kept independent (not
    imported) so this module has no dependency on the report exporter's
    internals. Returns the registered font family name."""
    import os
    font = "Sans"
    windows_dir = "C:/Windows/Fonts"
    styles = {
        "": [os.path.join(windows_dir, "arial.ttf"),
             "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
             "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"],
        "B": [os.path.join(windows_dir, "arialbd.ttf"),
              "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
              "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"],
    }
    fallback = None
    for style, paths in styles.items():
        for p in paths:
            if os.path.isfile(p):
                pdf.add_font(font, style, p)
                if style == "":
                    fallback = p
                break
        else:
            if fallback and style != "":
                pdf.add_font(font, style, fallback)
    return font


def build_pdf_bytes(question: str, enterprise_answer: dict) -> bytes:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    rgb = _BRAND_RGB
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()
    font = _register_fonts(pdf)

    pdf.set_font(font, "B", 16)
    pdf.set_text_color(*rgb["primary"])
    pdf.cell(0, 10, "ToolSmithAI — Business Answer", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_draw_color(*rgb["rule_light"])
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)

    def _section(title: str) -> None:
        pdf.ln(2)
        pdf.set_font(font, "B", 10)
        pdf.set_text_color(*rgb["text_muted"])
        pdf.cell(0, 6, title.upper(), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    _section("Question")
    pdf.set_font(font, "", 11)
    pdf.set_text_color(*rgb["text_dark"])
    pdf.multi_cell(0, 6, question or "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    _section("Answer")
    pdf.set_font(font, "B", 12)
    pdf.set_text_color(*rgb["text_dark"])
    pdf.multi_cell(0, 7, enterprise_answer.get("answer") or "", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    insight_text = _insight_chip_text(enterprise_answer)
    if insight_text:
        _section("Business Insight")
        direction = (enterprise_answer.get("insight") or {}).get("direction")
        color = rgb["success"] if direction == "up" else rgb["danger"] if direction == "down" else rgb["text_muted"]
        pdf.set_font(font, "B", 11)
        pdf.set_text_color(*color)
        pdf.cell(0, 7, insight_text, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    chart = enterprise_answer.get("chart")
    if chart and chart.get("labels") and chart.get("series"):
        _section(f"Chart ({chart.get('chart_type', 'bar')})")
        series = chart["series"][0] if chart["series"] else {"data": []}
        labels = chart["labels"]
        values = [v for v in series.get("data", []) if isinstance(v, (int, float))]
        max_val = max((abs(v) for v in values), default=1) or 1
        pdf.set_font(font, "", 8)
        bar_area_w = 110.0
        for label, value in zip(labels, series.get("data", [])):
            if not isinstance(value, (int, float)):
                continue
            y0 = pdf.get_y()
            pdf.set_text_color(*rgb["text_dark"])
            short_label = str(label)[:22]
            pdf.cell(40, 5, short_label, new_x=XPos.RIGHT, new_y=YPos.TOP)
            bar_w = max(1.0, (abs(value) / max_val) * bar_area_w)
            pdf.set_fill_color(*rgb["primary"])
            pdf.rect(pdf.get_x(), y0 + 0.5, bar_w, 4, style="F")
            pdf.set_xy(pdf.get_x() + bar_area_w + 2, y0)
            pdf.set_text_color(*rgb["text_muted"])
            pdf.cell(0, 5, f"{value:,}" if isinstance(value, int) else f"{value:,.2f}",
                     new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(2)

    columns, rows = _resolve_table(enterprise_answer)
    if columns:
        _section("Result Table")
        shown_rows = rows[:_MAX_TABLE_ROWS_PDF]
        col_w = max(20.0, 190.0 / max(len(columns), 1))
        pdf.set_font(font, "B", 8)
        pdf.set_fill_color(*_BRAND_RGB["primary"])
        pdf.set_text_color(255, 255, 255)
        for col_name in columns:
            pdf.cell(col_w, 6, str(col_name)[:24], border=0, fill=True, align="C")
        pdf.ln(6)
        pdf.set_font(font, "", 8)
        for i, data_row in enumerate(shown_rows):
            if pdf.get_y() > 270:
                pdf.add_page()
            fill = i % 2 == 1
            if fill:
                pdf.set_fill_color(*rgb["rule_light"])
            pdf.set_text_color(*rgb["text_dark"])
            for col_name in columns:
                val = data_row.get(col_name, "")
                pdf.cell(col_w, 6, str(val)[:30], border=0, fill=fill, align="L")
            pdf.ln(6)
        if len(rows) > _MAX_TABLE_ROWS_PDF:
            pdf.set_font(font, "I", 8)
            pdf.set_text_color(*rgb["text_muted"])
            pdf.cell(0, 6, f"+ {len(rows) - _MAX_TABLE_ROWS_PDF} more row(s) — see the Excel/CSV export for the full table.",
                     new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # Printed inline (not pinned to the page bottom via set_y(-15)) — a
    # fixed bottom position combined with auto_page_break can push this
    # onto a spurious trailing blank page for a short, single-page document.
    pdf.ln(6)
    pdf.set_font(font, "", 7)
    pdf.set_text_color(*rgb["text_muted"])
    stamp = datetime.now(timezone.utc).strftime("%b %d, %Y %H:%M UTC")
    pdf.cell(0, 5, f"Generated {stamp} by ToolSmithAI")

    out = pdf.output()
    return bytes(out)


def build_export(question: str, enterprise_answer: dict, fmt: str) -> "tuple[bytes, str, str]":
    """Returns (content_bytes, filename, media_type) for the given format.
    Caller (api/v1/composer.py) is responsible for calling check_exportable()
    first — this function does not re-check."""
    if fmt == "csv":
        return (
            build_csv_bytes(enterprise_answer),
            _safe_filename(question, "csv"),
            "text/csv",
        )
    if fmt == "xlsx":
        return (
            build_xlsx_bytes(question, enterprise_answer),
            _safe_filename(question, "xlsx"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    if fmt == "pdf":
        return (
            build_pdf_bytes(question, enterprise_answer),
            _safe_filename(question, "pdf"),
            "application/pdf",
        )
    raise ValueError(f"Unsupported export format: {fmt!r}")
